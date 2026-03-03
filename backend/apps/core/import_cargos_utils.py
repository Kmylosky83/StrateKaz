"""
Utilidades para importación masiva de Cargos desde Excel.
Sistema de Gestión StrateKaz — Core

Sigue el mismo patrón que talent_hub.colaboradores.import_utils
"""
import io
import logging

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# =============================================================================
# MAPEOS DE VALORES LEGIBLES → CODES DEL MODELO
# =============================================================================

NIVEL_JERARQUICO_MAP = {
    'ESTRATEGICO': 'ESTRATEGICO',
    'ESTRATÉGICO': 'ESTRATEGICO',
    'TACTICO': 'TACTICO',
    'TÁCTICO': 'TACTICO',
    'OPERATIVO': 'OPERATIVO',
    'APOYO': 'APOYO',
    'EXTERNO': 'EXTERNO',
}

NIVEL_EDUCATIVO_MAP = {
    'PRIMARIA': 'PRIMARIA',
    'BACHILLER': 'BACHILLER',
    'TECNICO': 'TECNICO',
    'TÉCNICO': 'TECNICO',
    'TECNOLOGO': 'TECNOLOGO',
    'TECNÓLOGO': 'TECNOLOGO',
    'PROFESIONAL': 'PROFESIONAL',
    'ESPECIALIZACION': 'ESPECIALIZACION',
    'ESPECIALIZACIÓN': 'ESPECIALIZACION',
    'MAESTRIA': 'MAESTRIA',
    'MAESTRÍA': 'MAESTRIA',
    'DOCTORADO': 'DOCTORADO',
}

EXPERIENCIA_MAP = {
    'SIN EXPERIENCIA': 'SIN_EXPERIENCIA',
    'SIN_EXPERIENCIA': 'SIN_EXPERIENCIA',
    '6 MESES': '6_MESES',
    '6_MESES': '6_MESES',
    '1 AÑO': '1_ANO',
    '1 ANO': '1_ANO',
    '1_ANO': '1_ANO',
    '2 AÑOS': '2_ANOS',
    '2 ANOS': '2_ANOS',
    '2_ANOS': '2_ANOS',
    '3 AÑOS': '3_ANOS',
    '3 ANOS': '3_ANOS',
    '3_ANOS': '3_ANOS',
    '5 AÑOS': '5_ANOS',
    '5 ANOS': '5_ANOS',
    '5_ANOS': '5_ANOS',
    '10 AÑOS': '10_ANOS',
    '10+ AÑOS': '10_ANOS',
    '10 ANOS': '10_ANOS',
    '10+ ANOS': '10_ANOS',
    '10_ANOS': '10_ANOS',
}

SI_NO_MAP = {
    'SI': True, 'SÍ': True, 'S': True, '1': True, 'TRUE': True, 'YES': True, 'Y': True,
    'NO': False, 'N': False, '0': False, 'FALSE': False,
}

# Columnas del archivo de importación (índice 0-based)
CARGO_COLUMNAS = [
    'codigo',
    'nombre',
    'nivel_jerarquico',
    'area_nombre',
    'cargo_padre_nombre',
    'descripcion',
    'nivel_educativo',
    'experiencia_requerida',
    'cantidad_posiciones',
    'is_jefatura',
]


# =============================================================================
# PARSER DE EXCEL
# =============================================================================

def parse_cargo_excel(file_content: bytes) -> list[dict]:
    """
    Parsea un archivo Excel y retorna lista de dicts con los datos de cada fila.
    Ignora la primera fila (cabeceras), la fila de ejemplo y la fila de notas.
    Los datos empiezan en la fila 4.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        # Ignorar filas completamente vacías
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        row_data = {'_fila': row_idx}
        for col_idx, col_name in enumerate(CARGO_COLUMNAS):
            if col_idx < len(row):
                val = row[col_idx]
                row_data[col_name] = val if val is not None else ''
            else:
                row_data[col_name] = ''

        rows.append(row_data)

    return rows


def normalizar_valor(valor) -> str:
    """Convierte un valor de celda a string normalizado en mayúsculas."""
    if valor is None:
        return ''
    return str(valor).strip().upper()


def parsear_entero(valor, default: int = 1) -> int:
    """Parsea un valor de celda a entero."""
    if not valor and valor != 0:
        return default
    try:
        return int(float(str(valor).replace(',', '').replace(' ', '')))
    except (ValueError, TypeError):
        return default


def parsear_bool(valor) -> bool:
    """Parsea un valor de celda a booleano."""
    return SI_NO_MAP.get(normalizar_valor(valor), False)


# =============================================================================
# GENERADOR DE PLANTILLA EXCEL
# =============================================================================

# Colores de la plantilla
COLOR_HEADER_BG = '4472C4'
COLOR_HEADER_FONT = 'FFFFFF'
COLOR_EXAMPLE_BG = 'F2F2F2'
COLOR_SECTION_BG = 'D6E4F0'

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

# Definición de columnas para la plantilla
TEMPLATE_COLUMNS = [
    # (header, ancho, requerido, ejemplo, nota)
    ('Código', 18, True, 'COORD-SST', 'Código único del cargo (sin espacios)'),
    ('Nombre', 30, True, 'Coordinador SST', 'Nombre descriptivo del cargo'),
    ('Nivel Jerárquico', 20, True, 'TACTICO', 'ESTRATEGICO, TACTICO, OPERATIVO, APOYO, EXTERNO'),
    ('Área / Proceso', 28, False, 'Seguridad Industrial', 'Nombre exacto del área (ver hoja Referencia)'),
    ('Reporta a (Cargo Padre)', 28, False, 'Gerente General', 'Nombre exacto del cargo superior'),
    ('Descripción', 40, False, 'Coordina el SG-SST', 'Descripción general del cargo'),
    ('Nivel Educativo', 20, False, 'PROFESIONAL', 'PRIMARIA, BACHILLER, TECNICO, TECNOLOGO, PROFESIONAL, ESPECIALIZACION, MAESTRIA, DOCTORADO'),
    ('Experiencia Requerida', 22, False, '3 años', 'Opciones: Sin experiencia, 6 meses, 1 año, 2 años, 3 años, 5 años, 10+ años (también acepta códigos: SIN_EXPERIENCIA, 6_MESES, 1_ANO, etc.)'),
    ('Cantidad Posiciones', 20, False, '1', 'Número de posiciones (default: 1)'),
    ('Es Jefatura', 16, False, 'No', 'Sí o No — Si tiene personal a cargo'),
]


def generate_cargo_import_template(
    areas: list[dict] | None = None,
    cargos_existentes: list[dict] | None = None,
) -> bytes:
    """
    Genera el archivo Excel de plantilla para importación masiva de cargos.
    Incluye:
    - Hoja 1: Plantilla con cabeceras coloreadas, fila de ejemplo y notas
    - Hoja 2: Valores de referencia (áreas, niveles, cargos existentes)
    """
    wb = openpyxl.Workbook()

    # ── HOJA 1: Plantilla ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Cargos'

    # Fila de cabeceras (fila 1)
    for col_idx, (header, width, required, _, _nota) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
        cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # Fila de ejemplo (fila 2)
    for col_idx, (_, _, _, ejemplo, _nota) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=ejemplo)
        cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_EXAMPLE_BG)
        cell.font = Font(italic=True, color='888888', size=9)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = THIN_BORDER

    # Fila de notas (fila 3)
    for col_idx, (_, _, required, _, nota) in enumerate(TEMPLATE_COLUMNS, start=1):
        prefix = '* ' if required else ''
        cell = ws.cell(
            row=3, column=col_idx,
            value=f'{prefix}{nota}' if nota else ('* Requerido' if required else '')
        )
        cell.fill = PatternFill(fill_type='solid', fgColor='FFFBE6' if required else 'F9F9F9')
        cell.font = Font(italic=True, color='777777', size=8)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 25

    # Filas vacías para datos (4 al 103 = 100 filas)
    for row_idx in range(4, 104):
        for col_idx in range(1, len(TEMPLATE_COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx, value='')
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Fijar filas de cabecera
    ws.freeze_panes = 'A4'

    # ── HOJA 2: Valores de Referencia ────────────────────────────────────────
    ws2 = wb.create_sheet('Referencia')

    def write_section(ws, title, headers, rows, start_row, start_col):
        """Escribe una sección de valores en la hoja de referencia."""
        title_cell = ws.cell(row=start_row, column=start_col, value=title)
        title_cell.font = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
        title_cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        title_cell.alignment = Alignment(horizontal='center')
        if len(headers) > 1:
            ws.merge_cells(
                start_row=start_row, start_column=start_col,
                end_row=start_row, end_column=start_col + len(headers) - 1
            )
        start_row += 1
        # Cabeceras
        for h_idx, h in enumerate(headers, start=start_col):
            hc = ws.cell(row=start_row, column=h_idx, value=h)
            hc.font = Font(bold=True, size=9)
            hc.fill = PatternFill(fill_type='solid', fgColor=COLOR_SECTION_BG)
            hc.border = THIN_BORDER
        start_row += 1
        # Datos
        for r in rows:
            for c_idx, val in enumerate(r, start=start_col):
                rc = ws.cell(row=start_row, column=c_idx, value=val)
                rc.border = THIN_BORDER
                rc.font = Font(size=9)
            start_row += 1
        return start_row + 1  # fila de separación

    # Secciones de referencia - columna 1
    row = 1
    row = write_section(ws2, 'NIVELES JERÁRQUICOS', ['Código', 'Descripción'], [
        ['ESTRATEGICO', 'Estratégico — Alta dirección'],
        ['TACTICO', 'Táctico — Mandos medios'],
        ['OPERATIVO', 'Operativo — Ejecución'],
        ['APOYO', 'Apoyo — Funciones transversales'],
        ['EXTERNO', 'Externo — Contratistas, consultores'],
    ], row, 1)

    row = write_section(ws2, 'NIVELES EDUCATIVOS', ['Código', 'Descripción'], [
        ['PRIMARIA', 'Primaria'],
        ['BACHILLER', 'Bachiller'],
        ['TECNICO', 'Técnico'],
        ['TECNOLOGO', 'Tecnólogo'],
        ['PROFESIONAL', 'Profesional'],
        ['ESPECIALIZACION', 'Especialización'],
        ['MAESTRIA', 'Maestría'],
        ['DOCTORADO', 'Doctorado'],
    ], row, 1)

    row = write_section(ws2, 'EXPERIENCIA', ['Código', 'Descripción'], [
        ['SIN_EXPERIENCIA', 'Sin experiencia'],
        ['6_MESES', '6 meses'],
        ['1_ANO', '1 año'],
        ['2_ANOS', '2 años'],
        ['3_ANOS', '3 años'],
        ['5_ANOS', '5 años'],
        ['10_ANOS', '10+ años'],
    ], row, 1)

    # Áreas disponibles en el tenant - columna 4
    area_rows = [[a['name']] for a in (areas or [])]
    if area_rows:
        write_section(ws2, 'ÁREAS / PROCESOS DISPONIBLES', ['Nombre del Área'], area_rows, 1, 4)

    # Cargos existentes en el tenant - columna 6
    cargo_rows = [[c['name'], c.get('code', '')] for c in (cargos_existentes or [])]
    if cargo_rows:
        write_section(
            ws2, 'CARGOS EXISTENTES (para Cargo Padre)',
            ['Nombre', 'Código'], cargo_rows, 1, 6
        )

    # Ajustar anchos hoja 2
    for col_idx in range(1, 9):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 30

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
