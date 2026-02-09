"""
Servicio centralizado de exportación CSV/Excel.

Usado por ExportMixin en ViewSets para generar archivos descargables.
Dependencias: openpyxl (ya en requirements/base.txt)
"""
import csv
import io
import logging
from datetime import date, datetime
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone

logger = logging.getLogger(__name__)

MAX_EXPORT_ROWS = 10_000


class ExportService:
    """Genera archivos CSV y Excel desde querysets de Django."""

    @classmethod
    def to_csv(cls, queryset, fields, filename='export'):
        """
        Genera respuesta HTTP con archivo CSV.

        Args:
            queryset: QuerySet de Django (ya filtrado)
            fields: Lista de dicts con 'field' (attr name) y 'header' (display name)
            filename: Nombre del archivo sin extensión
        """
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        response.write('\ufeff')  # BOM para Excel compatibilidad UTF-8

        writer = csv.writer(response)

        # Encabezados
        headers = [f['header'] for f in fields]
        writer.writerow(headers)

        # Datos
        rows = queryset[:MAX_EXPORT_ROWS]
        for obj in rows:
            row = []
            for f in fields:
                value = cls._get_nested_value(obj, f['field'])
                row.append(cls._format_value(value))
            writer.writerow(row)

        logger.info(
            f'Exportación CSV: {filename}.csv - {min(queryset.count(), MAX_EXPORT_ROWS)} filas'
        )
        return response

    @classmethod
    def to_excel(cls, queryset, fields, filename='export', sheet_name='Datos'):
        """
        Genera respuesta HTTP con archivo Excel (.xlsx).

        Args:
            queryset: QuerySet de Django (ya filtrado)
            fields: Lista de dicts con 'field' y 'header'
            filename: Nombre del archivo sin extensión
            sheet_name: Nombre de la hoja
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error('openpyxl no instalado. Instalar: pip install openpyxl')
            return HttpResponse(
                'Exportación Excel no disponible (openpyxl no instalado)',
                status=501
            )

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Estilos
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # Encabezados
        headers = [f['header'] for f in fields]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Datos
        rows = queryset[:MAX_EXPORT_ROWS]
        for row_idx, obj in enumerate(rows, 2):
            for col_idx, f in enumerate(fields, 1):
                value = cls._get_nested_value(obj, f['field'])
                formatted = cls._format_value(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=formatted)
                cell.border = thin_border

        # Auto-ajustar ancho de columnas
        for col_idx, f in enumerate(fields, 1):
            max_length = len(f['header'])
            for row_idx in range(2, min(ws.max_row + 1, 102)):  # Muestrear primeras 100 filas
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '')
                max_length = max(max_length, len(cell_value))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 3, 50)

        # Filtro automático
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

        # Congelar encabezado
        ws.freeze_panes = 'A2'

        # Pie de página con fecha
        ws.append([])
        ws.append([f'Exportado: {timezone.now().strftime("%Y-%m-%d %H:%M")} | StrateKaz ERP'])

        # Generar respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

        logger.info(
            f'Exportación Excel: {filename}.xlsx - {min(queryset.count(), MAX_EXPORT_ROWS)} filas'
        )
        return response

    @classmethod
    def get_export_fields(cls, serializer_class, custom_fields=None):
        """
        Obtiene campos de exportación desde un serializer o lista personalizada.

        Args:
            serializer_class: Clase del serializer DRF
            custom_fields: Lista de tuplas (field_name, header_label) o None

        Returns:
            Lista de dicts [{'field': 'campo', 'header': 'Encabezado'}]
        """
        if custom_fields:
            return [
                {'field': field, 'header': header}
                for field, header in custom_fields
            ]

        # Inferir desde el serializer
        serializer = serializer_class()
        fields = []
        excluded = {'id', 'created_by', 'updated_by', 'empresa', 'empresa_id'}

        for name, field in serializer.fields.items():
            if name in excluded:
                continue
            label = field.label or name.replace('_', ' ').title()
            fields.append({'field': name, 'header': str(label)})

        return fields

    @classmethod
    def _get_nested_value(cls, obj, field_path):
        """Obtiene valor de un campo, soportando acceso anidado con '__'."""
        parts = field_path.split('__')
        value = obj
        for part in parts:
            if value is None:
                return None
            if hasattr(value, part):
                value = getattr(value, part)
                if callable(value):
                    value = value()
            elif isinstance(value, dict):
                value = value.get(part)
            else:
                return None
        return value

    @classmethod
    def _format_value(cls, value):
        """Formatea un valor para exportación."""
        if value is None:
            return ''
        if isinstance(value, bool):
            return 'Sí' if value else 'No'
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M')
        if isinstance(value, date):
            return value.strftime('%Y-%m-%d')
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, (list, tuple)):
            return ', '.join(str(v) for v in value)
        return str(value)
