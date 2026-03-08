"""
Utilidades para importación masiva de Partes Interesadas desde Excel.
Sistema de Gestión StrateKaz — Contexto Organizacional

Patrón unificado: igual que core.import_cargos_utils,
supply_chain.import_proveedores_utils y sales_crm.import_clientes_utils.
"""
import io
import logging

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# =============================================================================
# CONSTANTES Y MAPEOS
# =============================================================================

NIVEL_INFLUENCIA_MAP = {
    'ALTA': 'alta', 'ALTO': 'alta', 'A': 'alta',
    'MEDIA': 'media', 'MEDIO': 'media', 'M': 'media',
    'BAJA': 'baja', 'BAJO': 'baja', 'B': 'baja',
}

NIVEL_INTERES_MAP = {
    'ALTO': 'alto', 'ALTA': 'alto', 'A': 'alto',
    'MEDIO': 'medio', 'MEDIA': 'medio', 'M': 'medio',
    'BAJO': 'bajo', 'BAJA': 'bajo', 'B': 'bajo',
}

CANAL_MAP = {
    'CORREO ELECTRONICO': 'email', 'CORREO ELECTRÓNICO': 'email', 'EMAIL': 'email',
    'TELEFONO': 'telefono', 'TELÉFONO': 'telefono',
    'REUNION PRESENCIAL': 'reunion', 'REUNIÓN PRESENCIAL': 'reunion', 'REUNION': 'reunion', 'REUNIÓN': 'reunion',
    'VIDEOCONFERENCIA': 'videoconferencia',
    'WHATSAPP': 'whatsapp',
    'PORTAL WEB': 'portal_web',
    'REDES SOCIALES': 'redes_sociales',
    'CORRESPONDENCIA FISICA': 'correspondencia', 'CORRESPONDENCIA FÍSICA': 'correspondencia', 'CORRESPONDENCIA': 'correspondencia',
    'OTRO': 'otro',
}

FRECUENCIA_MAP = {
    'DIARIA': 'diaria',
    'SEMANAL': 'semanal',
    'QUINCENAL': 'quincenal',
    'MENSUAL': 'mensual',
    'BIMESTRAL': 'bimestral',
    'TRIMESTRAL': 'trimestral',
    'SEMESTRAL': 'semestral',
    'ANUAL': 'anual',
    'SEGUN NECESIDAD': 'segun_necesidad',
    'SEGÚN NECESIDAD': 'segun_necesidad',
    'SEGUN NECESIDADES': 'segun_necesidad',
    'SEGÚN NECESIDADES': 'segun_necesidad',
}

# Columnas del archivo de importación (índice 0-based)
PI_COLUMNAS = [
    'grupo_nombre',           # A
    'subgrupo_nombre',        # B
    'nombre',                 # C
    'descripcion',            # D
    'representante',          # E
    'cargo_representante',    # F
    'email',                  # G
    'telefono',               # H
    'direccion',              # I
    'sitio_web',              # J
    'temas_interes_pi',       # K
    'temas_interes_empresa',  # L
    'nivel_influencia_pi',    # M
    'nivel_influencia_empresa', # N
    'nivel_interes',          # O
    'canal_principal',        # P
    'frecuencia_comunicacion', # Q
    'necesidades',            # R
    'expectativas',           # S
    'requisitos_pertinentes', # T
    'es_requisito_legal',     # U
    'relacionado_sst',        # V
    'relacionado_ambiental',  # W
    'relacionado_calidad',    # X
    'relacionado_pesv',       # Y
]


# =============================================================================
# HELPERS
# =============================================================================

def normalizar_valor(valor) -> str:
    """Convierte un valor de celda a string normalizado en mayúsculas."""
    if valor is None:
        return ''
    return str(valor).strip().upper()


def parsear_bool(valor) -> bool:
    """Convierte un valor de celda a booleano."""
    if not valor:
        return False
    norm = normalizar_valor(valor)
    return norm in ('SI', 'SÍ', 'S', '1', 'TRUE', 'YES', 'Y', 'X', '✓', '✔')


# =============================================================================
# PARSER DE EXCEL
# =============================================================================

def parse_partes_interesadas_excel(file_content: bytes) -> list[dict]:
    """
    Parsea un archivo Excel y retorna lista de dicts con los datos de cada fila.
    Los datos empiezan en la fila 4 (fila 1=headers, fila 2=ejemplo, fila 3=notas).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)

    # Intentar leer hoja "Partes Interesadas" o la primera hoja
    if 'Partes Interesadas' in wb.sheetnames:
        ws = wb['Partes Interesadas']
    else:
        ws = wb.active

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        # Ignorar filas completamente vacías
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        row_data = {'_fila': row_idx}
        for col_idx, col_name in enumerate(PI_COLUMNAS):
            if col_idx < len(row):
                val = row[col_idx]
                row_data[col_name] = val if val is not None else ''
            else:
                row_data[col_name] = ''

        rows.append(row_data)

    return rows


# =============================================================================
# GENERADOR DE PLANTILLA EXCEL
# =============================================================================

# Colores (azul corporativo ISO)
COLOR_HEADER_BG = '4472C4'
COLOR_HEADER_FONT = 'FFFFFF'
COLOR_EXAMPLE_BG = 'F2F2F2'
COLOR_SECTION_BG = 'D6E4F0'

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

# Definición de columnas para la plantilla
# (header, ancho, requerido, ejemplo, nota)
TEMPLATE_COLUMNS = [
    ('Grupo', 25, True, 'Personal', 'Grupo macro (ver hoja Referencia)'),
    ('Subgrupo (Tipo)', 25, False, 'Alta Dirección', 'Tipo dentro del grupo (se crea automáticamente si no existe)'),
    ('Nombre Parte Interesada', 30, True, 'Gerente General', 'Nombre único de la parte interesada'),
    ('Descripción', 40, False, 'Máxima autoridad de la empresa', 'Descripción de la parte interesada'),
    ('Representante', 25, False, 'Juan Pérez', 'Nombre de la persona representante'),
    ('Cargo Representante', 22, False, 'Director General', 'Cargo o rol del representante'),
    ('Email Contacto', 28, False, 'contacto@empresa.com', 'Correo electrónico de contacto'),
    ('Teléfono Contacto', 18, False, '+57 300 123 4567', 'Número de teléfono de contacto'),
    ('Dirección', 30, False, 'Calle 123 #45-67, Bogotá', 'Dirección física de la parte interesada'),
    ('Sitio Web', 28, False, 'https://empresa.com', 'URL del sitio web (incluir https://)'),
    ('Temas Interés PI', 35, False, 'Rentabilidad, crecimiento', 'Qué le interesa a la PI de la empresa'),
    ('Temas Interés Empresa', 35, False, 'Liderazgo, compromiso', 'Qué le interesa a la empresa de la PI'),
    ('Impacto PI→Empresa', 18, False, 'Alta', 'Alta / Media / Baja — Poder de la PI'),
    ('Impacto Empresa→PI', 18, False, 'Media', 'Alta / Media / Baja — Poder de la Empresa'),
    ('Nivel Interés', 18, False, 'Alto', 'Alto / Medio / Bajo'),
    ('Canal Comunicación', 22, False, 'Reunión Presencial', 'Canal principal (ver hoja Referencia)'),
    ('Frecuencia Comunicación', 22, False, 'Mensual', 'Diaria / Semanal / Quincenal / Mensual / Trimestral / Anual / Según necesidad'),
    ('Necesidades', 40, False, 'Productos de calidad, información oportuna', '¿Qué necesita esta PI de la organización?'),
    ('Expectativas', 40, False, 'Servicio confiable, mejora continua', '¿Qué espera esta PI de la organización?'),
    ('Requisitos Pertinentes', 40, False, 'Certificación ISO, tiempos de entrega', 'Requisitos que la empresa debe cumplir para esta PI'),
    ('Requisito Legal', 12, False, 'No', 'Sí / No — ¿Tiene requisitos de carácter legal?'),
    ('SST', 10, False, 'Sí', 'Sí / No — ¿Relacionado con SST (ISO 45001)?'),
    ('Ambiental', 10, False, 'No', 'Sí / No — ¿Relacionado con Ambiental (ISO 14001)?'),
    ('Calidad', 10, False, 'Sí', 'Sí / No — ¿Relacionado con Calidad (ISO 9001)?'),
    ('PESV', 10, False, 'No', 'Sí / No — ¿Relacionado con PESV?'),
]


def generate_partes_interesadas_template(
    grupos: list[dict] | None = None,
    tipos: list[dict] | None = None,
    partes_existentes: list[dict] | None = None,
) -> bytes:
    """
    Genera el archivo Excel de plantilla para importación masiva de partes interesadas.
    Incluye:
    - Hoja 1: Plantilla con cabeceras, fila ejemplo, notas, freeze panes
    - Hoja 2: Valores de referencia (grupos, tipos, niveles, canales)
    """
    wb = openpyxl.Workbook()

    # ── HOJA 1: Plantilla ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Partes Interesadas'

    # Fila de cabeceras (fila 1)
    for col_idx, (header, width, required, _, _nota) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
        cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # Fila de ejemplo (fila 2)
    for col_idx, (_, _, _, ejemplo, _nota) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=ejemplo)
        cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_EXAMPLE_BG)
        cell.font = Font(italic=True, color='888888', size=9)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = THIN_BORDER

    # Fila de notas (fila 3)
    for col_idx, (_, _, required, _, nota) in enumerate(TEMPLATE_COLUMNS, start=1):
        prefix = '* ' if required else ''
        cell = ws.cell(
            row=3, column=col_idx,
            value=f'{prefix}{nota}' if nota else ('* Requerido' if required else '')
        )
        cell.fill = PatternFill(fill_type='solid', fgColor='FFFBE6' if required else 'F9F9F9')
        cell.font = Font(italic=True, color='777777', size=8)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 25

    # Filas vacías para datos (4 al 503 = 500 filas)
    for row_idx in range(4, 504):
        for col_idx in range(1, len(TEMPLATE_COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx, value='')
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Fijar filas de cabecera
    ws.freeze_panes = 'A4'

    # ── HOJA 2: Valores de Referencia ────────────────────────────────────────
    ws2 = wb.create_sheet('Referencia')

    def write_section(ws_ref, title, headers, data_rows, start_row, start_col):
        """Escribe una sección de valores en la hoja de referencia."""
        title_cell = ws_ref.cell(row=start_row, column=start_col, value=title)
        title_cell.font = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
        title_cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        title_cell.alignment = Alignment(horizontal='center')
        if len(headers) > 1:
            ws_ref.merge_cells(
                start_row=start_row, start_column=start_col,
                end_row=start_row, end_column=start_col + len(headers) - 1
            )
        start_row += 1
        for h_idx, h in enumerate(headers, start=start_col):
            hc = ws_ref.cell(row=start_row, column=h_idx, value=h)
            hc.font = Font(bold=True, size=9)
            hc.fill = PatternFill(fill_type='solid', fgColor=COLOR_SECTION_BG)
            hc.border = THIN_BORDER
        start_row += 1
        for r in data_rows:
            for c_idx, val in enumerate(r, start=start_col):
                rc = ws_ref.cell(row=start_row, column=c_idx, value=val)
                rc.border = THIN_BORDER
                rc.font = Font(size=9)
            start_row += 1
        return start_row + 1

    # ── Columna 1: Catálogos ──
    row = 1

    # Grupos de Partes Interesadas
    grupo_rows = [[g['nombre']] for g in (grupos or [])]
    if grupo_rows:
        row = write_section(ws2, 'GRUPOS DE PARTES INTERESADAS', ['Nombre'], grupo_rows, row, 1)

    # Tipos (Subgrupos) por Grupo
    tipo_rows = [[t.get('grupo__nombre', ''), t['nombre']] for t in (tipos or [])]
    if tipo_rows:
        row = write_section(ws2, 'TIPOS (SUBGRUPOS)', ['Grupo', 'Nombre'], tipo_rows, row, 1)

    # Niveles de Influencia
    row = write_section(ws2, 'NIVELES DE INFLUENCIA/IMPACTO', ['Valor'], [
        ['Alta'], ['Media'], ['Baja'],
    ], row, 1)

    # Niveles de Interés
    row = write_section(ws2, 'NIVELES DE INTERÉS', ['Valor'], [
        ['Alto'], ['Medio'], ['Bajo'],
    ], row, 1)

    # Canales de Comunicación
    row = write_section(ws2, 'CANALES DE COMUNICACIÓN', ['Valor'], [
        ['Correo Electrónico'],
        ['Teléfono'],
        ['Reunión Presencial'],
        ['Videoconferencia'],
        ['WhatsApp'],
        ['Portal Web'],
        ['Redes Sociales'],
        ['Correspondencia Física'],
        ['Otro'],
    ], row, 1)

    # Frecuencias de Comunicación
    row = write_section(ws2, 'FRECUENCIAS DE COMUNICACIÓN', ['Valor'], [
        ['Diaria'], ['Semanal'], ['Quincenal'], ['Mensual'],
        ['Bimestral'], ['Trimestral'], ['Semestral'], ['Anual'],
        ['Según necesidad'],
    ], row, 1)

    # Sistemas de Gestión (SST/Ambiental/Calidad/PESV)
    write_section(ws2, 'RELACIÓN CON SISTEMAS / SÍ/NO', ['Valor'], [
        ['Sí'], ['No'],
    ], row, 1)

    # ── Columna 4: Partes existentes ──
    pi_rows = [[p.get('nombre', ''), p.get('tipo__grupo__nombre', ''), p.get('tipo__nombre', '')]
               for p in (partes_existentes or [])]
    if pi_rows:
        write_section(
            ws2, 'PARTES INTERESADAS EXISTENTES',
            ['Nombre', 'Grupo', 'Tipo'], pi_rows, 1, 4
        )

    # Ajustar anchos hoja 2
    for col_idx in range(1, 8):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 28

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
