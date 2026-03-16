"""
Views para Contexto Organizacional - Gestión Estratégica
========================================================

ViewSets para los modelos de análisis estratégico del contexto organizacional.

Todos los ViewSets implementan:
- StandardViewSetMixin: Funcionalidad estándar (toggle, filtros, bulk, auditoría)
- Filtros por empresa, estado, tipo, fecha
- Acciones especiales: aprobar(), archivar()
- Optimización de queries con select_related y prefetch_related

Endpoints generados:
- /api/gestion-estrategica/contexto/analisis-dofa/
- /api/gestion-estrategica/contexto/factores-dofa/
- /api/gestion-estrategica/contexto/estrategias-tows/
- /api/gestion-estrategica/contexto/analisis-pestel/
- /api/gestion-estrategica/contexto/factores-pestel/
- /api/gestion-estrategica/contexto/fuerzas-porter/
- /api/gestion-estrategica/contexto/tipos-parte-interesada/
- /api/gestion-estrategica/contexto/partes-interesadas/
- /api/gestion-estrategica/contexto/requisitos-pi/
- /api/gestion-estrategica/contexto/matriz-comunicacion/

Autor: Sistema ERP StrateKaz
Fecha: 2025-12-26
Actualizado: 2026-01-24 - Migrado a app independiente
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from django.db.models import Count
from django.db import transaction
from django.http import HttpResponse
from datetime import date
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from apps.core.mixins import StandardViewSetMixin
from apps.core.permissions import GranularActionPermission
from rest_framework.permissions import IsAuthenticated
from apps.gestion_estrategica.revision_direccion.services.resumen_mixin import ResumenRevisionMixin
from .models import (
    TipoAnalisisDOFA,
    AnalisisDOFA,
    FactorDOFA,
    EstrategiaTOWS,
    TipoAnalisisPESTEL,
    AnalisisPESTEL,
    FactorPESTEL,
    FuerzaPorter,
    GrupoParteInteresada,
    TipoParteInteresada,
    ParteInteresada,
    RequisitoParteInteresada,
    MatrizComunicacion
)
from apps.core.base_models.mixins import get_tenant_empresa


class EmpresaAutoAssignMixin:
    """Auto-asigna empresa y created_by al crear objetos BaseCompanyModel."""
    def perform_create(self, serializer):
        serializer.save(
            empresa=get_tenant_empresa(),
            created_by=self.request.user,
        )


from .serializers import (
    TipoAnalisisDOFASerializer,
    AnalisisDOFASerializer,
    FactorDOFASerializer,
    EstrategiaTOWSSerializer,
    TipoAnalisisPESTELSerializer,
    AnalisisPESTELSerializer,
    FactorPESTELSerializer,
    FuerzaPorterSerializer,
    GrupoParteInteresadaSerializer,
    TipoParteInteresadaSerializer,
    ParteInteresadaSerializer,
    RequisitoParteInteresadaSerializer,
    MatrizComunicacionSerializer
)


# ============================================================================
# CATÁLOGOS GLOBALES
# ============================================================================

class TipoAnalisisDOFAViewSet(StandardViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para gestión de Tipos de Análisis DOFA.

    Catálogo global de tipos de análisis (Organizacional, Anual, Por Área, etc.)
    No depende de empresa - es global para todo el sistema.
    """

    queryset = TipoAnalisisDOFA.objects.all()
    serializer_class = TipoAnalisisDOFASerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['is_active']
    search_fields = ['codigo', 'nombre', 'descripcion']
    ordering_fields = ['orden', 'nombre', 'codigo']
    ordering = ['orden', 'nombre']


class AnalisisDOFAViewSet(EmpresaAutoAssignMixin, ResumenRevisionMixin, StandardViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para gestión de análisis DOFA.

    Funcionalidad:
    - CRUD completo de análisis DOFA
    - Filtros por empresa, estado, periodo, fecha
    - Búsqueda por nombre, periodo, observaciones
    - Ordenamiento por fecha de análisis, creación
    - Acciones: aprobar(), archivar(), matriz_completa()
    - Funcionalidad estándar: toggle_active, bulk actions, auditoría

    Endpoints especiales:
    - POST /analisis-dofa/{id}/aprobar/ - Aprobar análisis
    - POST /analisis-dofa/{id}/archivar/ - Archivar análisis
    - GET /analisis-dofa/{id}/matriz-completa/ - Matriz DOFA completa con factores y estrategias
    - GET /analisis-dofa/estadisticas/ - Estadísticas generales
    """

    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'analisis_contexto'

    queryset = AnalisisDOFA.objects.select_related(
        'tipo_analisis',
        'responsable',
        'aprobado_por',
        'empresa'
    ).prefetch_related(
        'factores',
        'estrategias'
    )
    serializer_class = AnalisisDOFASerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['empresa', 'estado', 'periodo', 'responsable', 'tipo_analisis']
    search_fields = ['nombre', 'periodo', 'observaciones']
    ordering_fields = ['fecha_analisis', 'created_at', 'periodo']
    ordering = ['-fecha_analisis']

    # ResumenRevisionMixin config
    resumen_date_field = 'fecha_analisis'
    resumen_modulo_nombre = 'contexto_organizacional'

    def get_resumen_data(self, queryset, fecha_desde, fecha_hasta):
        """Resumen de contexto organizacional para Revisión por la Dirección."""
        # Análisis DOFA en período
        total_dofa = queryset.count()
        dofa_por_estado = list(
            queryset.values('estado').annotate(cantidad=Count('id')).order_by('estado')
        )

        # Factores DOFA (Debilidades, Oportunidades, Fortalezas, Amenazas)
        factores = FactorDOFA.objects.filter(
            analisis__in=queryset, is_active=True
        )
        factores_por_tipo = list(
            factores.values('tipo_factor').annotate(cantidad=Count('id')).order_by('tipo_factor')
        )

        # Análisis PESTEL en período
        pestel = AnalisisPESTEL.objects.filter(
            fecha_analisis__range=[fecha_desde, fecha_hasta]
        )
        total_pestel = pestel.count()

        # Factores PESTEL
        factores_pestel = FactorPESTEL.objects.filter(
            analisis__in=pestel, is_active=True
        )
        pestel_por_tipo = list(
            factores_pestel.values('tipo_factor')
            .annotate(cantidad=Count('id'))
            .order_by('tipo_factor')
        )

        # Partes interesadas
        total_partes_interesadas = ParteInteresada.objects.filter(is_active=True).count()
        nuevas_pi = ParteInteresada.objects.filter(
            created_at__date__range=[fecha_desde, fecha_hasta]
        ).count()

        return {
            'analisis_dofa': {
                'total': total_dofa,
                'por_estado': dofa_por_estado,
                'factores_por_tipo': factores_por_tipo,
            },
            'analisis_pestel': {
                'total': total_pestel,
                'factores_por_tipo': pestel_por_tipo,
            },
            'partes_interesadas': {
                'total': total_partes_interesadas,
                'nuevas_en_periodo': nuevas_pi,
            },
        }

    @action(detail=True, methods=['post'])
    def aprobar(self, request, pk=None) -> Response:
        """Aprobar un análisis DOFA."""
        analisis = self.get_object()

        if analisis.estado == 'aprobado':
            return Response(
                {'error': 'El análisis ya está aprobado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        analisis.estado = 'aprobado'
        analisis.aprobado_por = request.user
        analisis.fecha_aprobacion = date.today()
        analisis.save(update_fields=['estado', 'aprobado_por', 'fecha_aprobacion', 'updated_at'])

        serializer = self.get_serializer(analisis)
        return Response({
            'message': 'Análisis DOFA aprobado exitosamente',
            'data': serializer.data
        })

    @action(detail=True, methods=['post'])
    def archivar(self, request, pk=None) -> Response:
        """Archivar un análisis DOFA."""
        analisis = self.get_object()

        if analisis.estado == 'archivado':
            return Response(
                {'error': 'El análisis ya está archivado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        analisis.estado = 'archivado'
        analisis.save(update_fields=['estado', 'updated_at'])

        serializer = self.get_serializer(analisis)
        return Response({
            'message': 'Análisis DOFA archivado exitosamente',
            'data': serializer.data
        })

    @action(detail=True, methods=['get'], url_path='matriz-completa')
    def matriz_completa(self, request, pk=None) -> Response:
        """Retorna la matriz DOFA completa con factores y estrategias organizados."""
        analisis = self.get_object()

        # Factores por tipo
        factores = analisis.factores.all()
        fortalezas = FactorDOFASerializer(factores.filter(tipo='fortaleza'), many=True).data
        oportunidades = FactorDOFASerializer(factores.filter(tipo='oportunidad'), many=True).data
        debilidades = FactorDOFASerializer(factores.filter(tipo='debilidad'), many=True).data
        amenazas = FactorDOFASerializer(factores.filter(tipo='amenaza'), many=True).data

        # Estrategias por tipo
        estrategias = analisis.estrategias.all()
        estrategias_fo = EstrategiaTOWSSerializer(estrategias.filter(tipo='fo'), many=True).data
        estrategias_fa = EstrategiaTOWSSerializer(estrategias.filter(tipo='fa'), many=True).data
        estrategias_do = EstrategiaTOWSSerializer(estrategias.filter(tipo='do'), many=True).data
        estrategias_da = EstrategiaTOWSSerializer(estrategias.filter(tipo='da'), many=True).data

        return Response({
            'analisis': AnalisisDOFASerializer(analisis).data,
            'factores': {
                'fortalezas': fortalezas,
                'oportunidades': oportunidades,
                'debilidades': debilidades,
                'amenazas': amenazas,
            },
            'estrategias': {
                'fo': estrategias_fo,
                'fa': estrategias_fa,
                'do': estrategias_do,
                'da': estrategias_da,
            },
            'resumen': {
                'total_fortalezas': len(fortalezas),
                'total_oportunidades': len(oportunidades),
                'total_debilidades': len(debilidades),
                'total_amenazas': len(amenazas),
                'total_estrategias': len(estrategias),
                'total_estrategias_fo': len(estrategias_fo),
                'total_estrategias_fa': len(estrategias_fa),
                'total_estrategias_do': len(estrategias_do),
                'total_estrategias_da': len(estrategias_da),
            }
        })

    @action(detail=False, methods=['get'])
    def estadisticas(self, request) -> Response:
        """Retorna estadísticas generales de análisis DOFA."""
        queryset = self.filter_queryset(self.get_queryset())

        stats = {
            'total': queryset.count(),
            'por_estado': dict(
                queryset.values('estado').annotate(
                    total=Count('id')
                ).values_list('estado', 'total')
            ),
            'vigentes': queryset.filter(estado='vigente').count(),
        }

        return Response(stats)


class FactorDOFAViewSet(EmpresaAutoAssignMixin, StandardViewSetMixin, viewsets.ModelViewSet):
    """ViewSet para gestión de factores DOFA."""

    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'analisis_contexto'

    queryset = FactorDOFA.objects.select_related(
        'analisis',
        'empresa',
        'created_by'
    )
    serializer_class = FactorDOFASerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['empresa', 'analisis', 'tipo', 'impacto']
    search_fields = ['descripcion', 'area_afectada', 'evidencias']
    ordering_fields = ['analisis', 'tipo', 'orden', 'created_at']
    ordering = ['analisis', 'tipo', 'orden']


class EstrategiaTOWSViewSet(EmpresaAutoAssignMixin, StandardViewSetMixin, viewsets.ModelViewSet):
    """ViewSet para gestión de estrategias TOWS."""

    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'analisis_contexto'

    queryset = EstrategiaTOWS.objects.select_related(
        'analisis',
        'responsable',
        'empresa',
        'created_by'
    )
    serializer_class = EstrategiaTOWSSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['empresa', 'analisis', 'tipo', 'estado', 'prioridad', 'responsable']
    search_fields = ['descripcion', 'objetivo', 'recursos_necesarios']
    ordering_fields = ['prioridad', 'fecha_limite', 'estado', 'created_at']
    ordering = ['-prioridad', 'fecha_limite']

    @action(detail=True, methods=['post'])
    def aprobar(self, request, pk=None) -> Response:
        """Aprobar una estrategia TOWS."""
        estrategia = self.get_object()

        if estrategia.estado != 'propuesta':
            return Response(
                {'error': 'Solo se pueden aprobar estrategias propuestas'},
                status=status.HTTP_400_BAD_REQUEST
            )

        estrategia.estado = 'aprobada'
        estrategia.save(update_fields=['estado', 'updated_at'])

        serializer = self.get_serializer(estrategia)
        return Response({
            'message': 'Estrategia aprobada exitosamente',
            'data': serializer.data
        })

    @action(detail=True, methods=['post'])
    def ejecutar(self, request, pk=None) -> Response:
        """Marcar una estrategia como en ejecución."""
        estrategia = self.get_object()

        if estrategia.estado not in ['propuesta', 'aprobada']:
            return Response(
                {'error': 'La estrategia debe estar propuesta o aprobada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        estrategia.estado = 'en_ejecucion'
        if not estrategia.fecha_implementacion:
            estrategia.fecha_implementacion = date.today()
        estrategia.save(update_fields=['estado', 'fecha_implementacion', 'updated_at'])

        serializer = self.get_serializer(estrategia)
        return Response({
            'message': 'Estrategia marcada como en ejecución',
            'data': serializer.data
        })

    @action(detail=True, methods=['post'])
    def completar(self, request, pk=None) -> Response:
        """Marcar una estrategia como completada."""
        estrategia = self.get_object()

        if estrategia.estado != 'en_ejecucion':
            return Response(
                {'error': 'Solo se pueden completar estrategias en ejecución'},
                status=status.HTTP_400_BAD_REQUEST
            )

        estrategia.estado = 'completada'
        estrategia.progreso_porcentaje = 100
        estrategia.save(update_fields=['estado', 'progreso_porcentaje', 'updated_at'])

        serializer = self.get_serializer(estrategia)
        return Response({
            'message': 'Estrategia completada exitosamente',
            'data': serializer.data
        })

    @action(detail=False, methods=['get'], url_path='proximas-vencer')
    def proximas_vencer(self, request) -> Response:
        """Retorna estrategias próximas a vencer (30 días)."""
        from datetime import timedelta

        dias = int(request.query_params.get('dias', 30))
        fecha_limite = date.today() + timedelta(days=dias)

        estrategias = self.filter_queryset(self.get_queryset()).filter(
            fecha_limite__lte=fecha_limite,
            fecha_limite__gte=date.today(),
            estado__in=['propuesta', 'aprobada', 'en_ejecucion']
        ).order_by('fecha_limite')

        serializer = self.get_serializer(estrategias, many=True)
        return Response({
            'estrategias': serializer.data,
            'total': estrategias.count()
        })

    @action(detail=True, methods=['post'], url_path='convertir-objetivo')
    def convertir_objetivo(self, request, pk=None) -> Response:
        """Convierte una estrategia TOWS en un objetivo estratégico."""
        from apps.gestion_estrategica.planeacion.models import StrategicObjective, StrategicPlan

        estrategia = self.get_object()

        # Validar que no esté ya convertida (buscar si existe objetivo con esta estrategia como origen)
        objetivo_existente = StrategicObjective.objects.filter(estrategia_origen=estrategia).first()
        if objetivo_existente:
            return Response(
                {
                    'error': 'Esta estrategia ya fue convertida en objetivo',
                    'objetivo_id': str(objetivo_existente.id),
                    'objetivo_code': objetivo_existente.code
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que esté aprobada
        if estrategia.estado not in ['aprobada', 'en_ejecucion']:
            return Response(
                {'error': 'Solo se pueden convertir estrategias aprobadas o en ejecución'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener plan activo
        plan_activo = StrategicPlan.get_active()
        if not plan_activo:
            return Response(
                {'error': 'No hay un plan estratégico activo. Cree uno primero.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Datos del objetivo
        objetivo_data = request.data
        code = objetivo_data.get('code')
        name = objetivo_data.get('name', estrategia.descripcion[:300])
        bsc_perspective = objetivo_data.get('bsc_perspective')
        target_value = objetivo_data.get('target_value')
        unit = objetivo_data.get('unit', '%')

        # Validar campos requeridos
        if not code or not bsc_perspective:
            return Response(
                {'error': 'Los campos code y bsc_perspective son requeridos'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que el código no exista
        if StrategicObjective.objects.filter(plan=plan_activo, code=code).exists():
            return Response(
                {'error': f'Ya existe un objetivo con el código {code} en el plan actual'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Crear objetivo estratégico con referencia a la estrategia de origen
        objetivo = StrategicObjective.objects.create(
            plan=plan_activo,
            code=code,
            name=name,
            description=f"{estrategia.descripcion}\n\nObjetivo esperado: {estrategia.objetivo}",
            bsc_perspective=bsc_perspective,
            responsible=estrategia.responsable,
            target_value=target_value,
            unit=unit,
            status='PENDIENTE',
            estrategia_origen=estrategia,  # Nueva FK: objetivo apunta a estrategia
            created_by=request.user,
            updated_by=request.user
        )

        # Vincular área responsable si existe
        if estrategia.area_responsable:
            objetivo.areas_responsables.add(estrategia.area_responsable)

        # Serializar respuesta
        from apps.gestion_estrategica.planeacion.serializers import StrategicObjectiveSerializer
        objetivo_serializer = StrategicObjectiveSerializer(objetivo)
        estrategia_serializer = self.get_serializer(estrategia)

        return Response({
            'message': 'Objetivo estratégico creado exitosamente',
            'objetivo': objetivo_serializer.data,
            'estrategia': estrategia_serializer.data
        }, status=status.HTTP_201_CREATED)


class TipoAnalisisPESTELViewSet(StandardViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para gestión de Tipos de Análisis PESTEL.

    Catálogo global de tipos de análisis (Macro-Entorno, Sectorial, Por País, etc.)
    No depende de empresa - es global para todo el sistema.
    """

    queryset = TipoAnalisisPESTEL.objects.all()
    serializer_class = TipoAnalisisPESTELSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['is_active']
    search_fields = ['codigo', 'nombre', 'descripcion']
    ordering_fields = ['orden', 'nombre', 'codigo']
    ordering = ['orden', 'nombre']


class AnalisisPESTELViewSet(EmpresaAutoAssignMixin, StandardViewSetMixin, viewsets.ModelViewSet):
    """ViewSet para gestión de análisis PESTEL."""

    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'analisis_contexto'

    queryset = AnalisisPESTEL.objects.select_related(
        'tipo_analisis',
        'responsable',
        'empresa'
    ).prefetch_related(
        'factores'
    )
    serializer_class = AnalisisPESTELSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['empresa', 'estado', 'periodo', 'responsable', 'tipo_analisis']
    search_fields = ['nombre', 'periodo', 'conclusiones']
    ordering_fields = ['fecha_analisis', 'created_at', 'periodo']
    ordering = ['-fecha_analisis']

    @action(detail=True, methods=['post'])
    def aprobar(self, request, pk=None) -> Response:
        """Aprobar un análisis PESTEL."""
        analisis = self.get_object()

        if analisis.estado == 'aprobado':
            return Response(
                {'error': 'El análisis ya está aprobado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        analisis.estado = 'aprobado'
        analisis.save(update_fields=['estado', 'updated_at'])

        serializer = self.get_serializer(analisis)
        return Response({
            'message': 'Análisis PESTEL aprobado exitosamente',
            'data': serializer.data
        })

    @action(detail=True, methods=['post'])
    def archivar(self, request, pk=None) -> Response:
        """Archivar un análisis PESTEL."""
        analisis = self.get_object()

        if analisis.estado == 'archivado':
            return Response(
                {'error': 'El análisis ya está archivado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        analisis.estado = 'archivado'
        analisis.save(update_fields=['estado', 'updated_at'])

        serializer = self.get_serializer(analisis)
        return Response({
            'message': 'Análisis PESTEL archivado exitosamente',
            'data': serializer.data
        })

    @action(detail=False, methods=['get'])
    def estadisticas(self, request) -> Response:
        """Retorna estadísticas generales de análisis PESTEL."""
        queryset = self.filter_queryset(self.get_queryset())

        stats = {
            'total': queryset.count(),
            'por_estado': dict(
                queryset.values('estado').annotate(
                    total=Count('id')
                ).values_list('estado', 'total')
            ),
            'vigentes': queryset.filter(estado='vigente').count(),
        }

        return Response(stats)


class FactorPESTELViewSet(EmpresaAutoAssignMixin, StandardViewSetMixin, viewsets.ModelViewSet):
    """ViewSet para gestión de factores PESTEL."""

    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'analisis_contexto'

    queryset = FactorPESTEL.objects.select_related(
        'analisis',
        'empresa',
        'created_by'
    )
    serializer_class = FactorPESTELSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['empresa', 'analisis', 'tipo', 'impacto', 'probabilidad', 'tendencia']
    search_fields = ['descripcion', 'implicaciones', 'fuentes']
    ordering_fields = ['analisis', 'tipo', 'orden', 'created_at']
    ordering = ['analisis', 'tipo', 'orden']


class FuerzaPorterViewSet(EmpresaAutoAssignMixin, StandardViewSetMixin, viewsets.ModelViewSet):
    """ViewSet para gestión de fuerzas de Porter."""

    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'analisis_contexto'

    queryset = FuerzaPorter.objects.select_related(
        'empresa',
        'created_by'
    )
    serializer_class = FuerzaPorterSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['empresa', 'tipo', 'nivel', 'periodo']
    search_fields = ['descripcion', 'implicaciones_estrategicas']
    ordering_fields = ['tipo', 'nivel', 'fecha_analisis', 'created_at']
    ordering = ['tipo']

    @action(detail=False, methods=['get'])
    def resumen(self, request) -> Response:
        """Retorna resumen de las 5 fuerzas de Porter."""
        periodo = request.query_params.get('periodo')

        queryset = self.filter_queryset(self.get_queryset())
        if periodo:
            queryset = queryset.filter(periodo=periodo)

        serializer = self.get_serializer(queryset, many=True)

        distribucion = dict(
            queryset.values('nivel').annotate(
                total=Count('id')
            ).values_list('nivel', 'total')
        )

        return Response({
            'periodo': periodo,
            'fuerzas': serializer.data,
            'total_fuerzas': queryset.count(),
            'distribucion_niveles': distribucion
        })


# ============================================================================
# VIEWSETS PARTES INTERESADAS (Stakeholders) - ISO 9001:2015 Cláusula 4.2
# ============================================================================

class GrupoParteInteresadaViewSet(StandardViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para gestión de Grupos de Partes Interesadas.

    Funcionalidad:
    - CRUD completo de grupos
    - 10 grupos pre-seeded del sistema (es_sistema=True)
    - Grupos custom creados por la empresa (es_sistema=False)
    - Los grupos del sistema no pueden eliminarse (soft-delete protegido)
    - Filtros por estado, es_sistema
    - Ordenamiento por orden, nombre
    """

    queryset = GrupoParteInteresada.objects.all()
    serializer_class = GrupoParteInteresadaSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['is_active', 'es_sistema']
    search_fields = ['codigo', 'nombre', 'descripcion']
    ordering_fields = ['orden', 'nombre', 'codigo']
    ordering = ['orden', 'nombre']

    def perform_destroy(self, instance):
        """Proteger grupos del sistema de eliminación física."""
        if instance.es_sistema:
            # Soft delete only para grupos del sistema
            instance.is_active = False
            instance.save(update_fields=['is_active', 'updated_at'])
        else:
            # Permitir eliminación física de grupos custom
            super().perform_destroy(instance)


class TipoParteInteresadaViewSet(StandardViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para gestión de Tipos de Parte Interesada (Subgrupos).

    Funcionalidad:
    - CRUD completo de tipos/subgrupos
    - Filtros por grupo, categoria, es_sistema, estado
    - Búsqueda por código, nombre, descripción
    - Ordenamiento jerárquico: grupo.orden → tipo.orden → nombre
    """

    queryset = TipoParteInteresada.objects.select_related('grupo').all()
    serializer_class = TipoParteInteresadaSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['grupo', 'categoria', 'es_sistema', 'is_active']
    search_fields = ['codigo', 'nombre', 'descripcion']
    ordering_fields = ['orden', 'nombre', 'categoria']
    ordering = ['grupo__orden', 'orden', 'nombre']


class ParteInteresadaViewSet(EmpresaAutoAssignMixin, StandardViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para gestión de Partes Interesadas (Stakeholders).

    Funcionalidad:
    - CRUD completo de partes interesadas
    - Filtros por grupo, tipo, nivel_influencia_pi, nivel_influencia_empresa, responsable, área
    - Export Excel (4 hojas): Identificación, Caracterización, Modelos Relación, Matriz
    - Import Excel: Importar desde formato F-GD-04
    - Generación automática de Matriz de Comunicaciones
    - Matriz Poder-Interés (4 cuadrantes)
    - Estadísticas por grupo, tipo, sistema
    """

    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'partes_interesadas'

    queryset = ParteInteresada.objects.select_related(
        'tipo', 'tipo__grupo', 'cargo_responsable', 'area_responsable'
    ).all()
    serializer_class = ParteInteresadaSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = [
        'empresa', 'tipo', 'tipo__grupo', 'nivel_influencia_pi', 'nivel_influencia_empresa',
        'nivel_interes', 'cargo_responsable', 'area_responsable',
        'relacionado_sst', 'relacionado_ambiental', 'relacionado_calidad', 'relacionado_pesv',
        'is_active'
    ]
    search_fields = ['nombre', 'descripcion', 'representante', 'temas_interes_pi', 'temas_interes_empresa']
    ordering_fields = ['nombre', 'nivel_influencia_pi', 'nivel_influencia_empresa', 'nivel_interes', 'created_at']
    ordering = ['tipo__grupo__orden', 'tipo__orden', 'nombre']

    @action(detail=False, methods=['get'], url_path='matriz-poder-interes')
    def matriz_poder_interes(self, request) -> Response:
        """
        Retorna las partes interesadas organizadas por cuadrante.

        Cuadrantes basados en:
        - nivel_influencia_pi (PODER de la PI sobre la empresa): alta/media/baja
        - nivel_interes (INTERÉS de la PI en la empresa): alto/medio/bajo

        Cuadrantes:
        - gestionar_cerca: Alta influencia PI + Alto interés
        - mantener_satisfecho: Alta influencia PI + Medio/Bajo interés
        - mantener_informado: Media/Baja influencia PI + Alto interés
        - monitorear: Media/Baja influencia PI + Medio/Bajo interés
        """
        queryset = self.filter_queryset(self.get_queryset()).filter(is_active=True)

        cuadrantes = {
            'gestionar_cerca': [],
            'mantener_satisfecho': [],
            'mantener_informado': [],
            'monitorear': []
        }

        for parte in queryset:
            data = self.get_serializer(parte).data
            # nivel_influencia_pi: el PODER que la PI tiene sobre la empresa
            if parte.nivel_influencia_pi == 'alta' and parte.nivel_interes == 'alto':
                cuadrantes['gestionar_cerca'].append(data)
            elif parte.nivel_influencia_pi == 'alta':
                cuadrantes['mantener_satisfecho'].append(data)
            elif parte.nivel_interes == 'alto':
                cuadrantes['mantener_informado'].append(data)
            else:
                cuadrantes['monitorear'].append(data)

        return Response(cuadrantes)

    @action(detail=False, methods=['get'])
    def estadisticas(self, request) -> Response:
        """Retorna estadísticas de partes interesadas."""
        queryset = self.filter_queryset(self.get_queryset())

        stats = {
            'total': queryset.count(),
            'por_grupo': {
                k or 'Sin grupo': v
                for k, v in queryset.values('tipo__grupo__nombre').annotate(
                    total=Count('id')
                ).values_list('tipo__grupo__nombre', 'total')
            },
            'por_tipo': dict(
                queryset.values('tipo__nombre').annotate(
                    total=Count('id')
                ).values_list('tipo__nombre', 'total')
            ),
            'por_influencia_pi': dict(
                queryset.values('nivel_influencia_pi').annotate(
                    total=Count('id')
                ).values_list('nivel_influencia_pi', 'total')
            ),
            'por_influencia_empresa': dict(
                queryset.values('nivel_influencia_empresa').annotate(
                    total=Count('id')
                ).values_list('nivel_influencia_empresa', 'total')
            ),
            'por_interes': dict(
                queryset.values('nivel_interes').annotate(
                    total=Count('id')
                ).values_list('nivel_interes', 'total')
            ),
            'por_sistema': {
                'sst': queryset.filter(relacionado_sst=True).count(),
                'ambiental': queryset.filter(relacionado_ambiental=True).count(),
                'calidad': queryset.filter(relacionado_calidad=True).count(),
                'pesv': queryset.filter(relacionado_pesv=True).count(),
            }
        }

        return Response(stats)

    @action(detail=False, methods=['post'], url_path='generar-matriz-comunicacion')
    def generar_matriz_comunicacion(self, request) -> Response:
        """
        Genera la matriz de comunicaciones para una parte interesada específica.

        Body: { "parte_interesada_id": <int> }

        Lógica:
        - Usa el cuadrante de la matriz poder-interés para determinar frecuencia
        - Cuadrante → Frecuencia:
          - gestionar_cerca → mensual
          - mantener_satisfecho → trimestral
          - mantener_informado → bimestral
          - monitorear → semestral
        """
        parte_interesada_id = request.data.get('parte_interesada_id')

        if not parte_interesada_id:
            return Response(
                {'error': 'El campo parte_interesada_id es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            parte = ParteInteresada.objects.get(id=parte_interesada_id)
        except ParteInteresada.DoesNotExist:
            return Response(
                {'error': 'Parte interesada no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Llamar al método del modelo (retorna lista de tuplas)
        resultados = parte.generar_comunicacion_automatica()

        if not resultados:
            return Response(
                {'error': 'No se pudo generar la matriz. Verifique que la parte interesada tiene canales configurados.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        comunicaciones = [r[0] for r in resultados]
        any_created = any(r[1] for r in resultados)

        return Response({
            'message': f'Matriz de comunicación generada exitosamente ({len(comunicaciones)} registros)',
            'created': any_created,
            'total': len(comunicaciones),
            'data': MatrizComunicacionSerializer(comunicaciones, many=True).data
        }, status=status.HTTP_201_CREATED if any_created else status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='generar-matriz-comunicacion-masiva')
    def generar_matriz_comunicacion_masiva(self, request) -> Response:
        """
        Genera matrices de comunicación para todas las partes interesadas activas.

        Filtro opcional: ?grupo=<id> para generar solo para un grupo específico
        """
        grupo_id = request.query_params.get('grupo')

        queryset = ParteInteresada.objects.filter(is_active=True)
        if grupo_id:
            queryset = queryset.filter(tipo__grupo_id=grupo_id)

        created_count = 0
        updated_count = 0
        errors = []

        for parte in queryset:
            try:
                resultados = parte.generar_comunicacion_automatica()
                for _, created in resultados:
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
            except Exception as e:
                errors.append({
                    'parte_interesada_id': parte.id,
                    'parte_interesada_nombre': parte.nombre,
                    'error': str(e)
                })

        return Response({
            'message': f'Proceso completado: {created_count} creadas, {updated_count} actualizadas',
            'created': created_count,
            'updated': updated_count,
            'errors': errors,
            'total_procesadas': created_count + updated_count,
            'total_errores': len(errors)
        })

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request) -> Response:
        """
        Exporta las partes interesadas a formato Excel F-GD-04.

        4 hojas:
        1. Identificación (GRUPO → SUBGRUPO → PI)
        2. Caracterización (Temas de interés bidireccionales + Impacto bidireccional)
        3. Modelos de Relación (Responsable + Canal comunicación)
        4. Matriz Consolidada (resumen completo)

        Formato compatible con F-GD-04 MATRIZ PARTES INTERESADAS.xlsx
        """
        queryset = self.filter_queryset(self.get_queryset()).select_related(
            'tipo', 'tipo__grupo', 'cargo_responsable', 'area_responsable'
        ).order_by('tipo__grupo__orden', 'tipo__orden', 'nombre')

        # Crear workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remover sheet default

        # Estilos comunes
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ====================================================================
        # HOJA 1: IDENTIFICACIÓN
        # ====================================================================
        ws1 = wb.create_sheet('Identificación')
        ws1.append(['GRUPO', 'SUBGRUPO (TIPO)', 'NOMBRE PARTE INTERESADA', 'DESCRIPCIÓN', 'REPRESENTANTE'])

        # Aplicar estilos a headers
        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for parte in queryset:
            ws1.append([
                parte.tipo.grupo.nombre if parte.tipo and parte.tipo.grupo else '',
                parte.tipo.nombre if parte.tipo else '',
                parte.nombre,
                parte.descripcion or '',
                parte.representante or ''
            ])

        # Ajustar anchos
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 30
        ws1.column_dimensions['C'].width = 35
        ws1.column_dimensions['D'].width = 50
        ws1.column_dimensions['E'].width = 30

        # ====================================================================
        # HOJA 2: CARACTERIZACIÓN
        # ====================================================================
        ws2 = wb.create_sheet('Caracterización')
        ws2.append([
            'PARTE INTERESADA',
            'TEMAS DE INTERÉS PARA LA PI',
            'TEMAS DE INTERÉS PARA LA EMPRESA',
            'IMPACTO PI → EMPRESA (PODER)',
            'IMPACTO EMPRESA → PI',
            'NIVEL DE INTERÉS DE LA PI'
        ])

        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for parte in queryset:
            ws2.append([
                parte.nombre,
                parte.temas_interes_pi or '',
                parte.temas_interes_empresa or '',
                parte.get_nivel_influencia_pi_display(),
                parte.get_nivel_influencia_empresa_display(),
                parte.get_nivel_interes_display()
            ])

        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 45
        ws2.column_dimensions['C'].width = 45
        ws2.column_dimensions['D'].width = 25
        ws2.column_dimensions['E'].width = 25
        ws2.column_dimensions['F'].width = 25

        # ====================================================================
        # HOJA 3: MODELOS DE RELACIÓN
        # ====================================================================
        ws3 = wb.create_sheet('Modelos de Relación')
        ws3.append([
            'PARTE INTERESADA',
            'RESPONSABLE EN LA EMPRESA',
            'CARGO RESPONSABLE',
            'ÁREA RESPONSABLE',
            'CANAL PRINCIPAL DE COMUNICACIÓN'
        ])

        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for parte in queryset:
            ws3.append([
                parte.nombre,
                parte.responsable_empresa_nombre or '',
                parte.cargo_responsable.name if parte.cargo_responsable else '',
                parte.area_responsable.name if parte.area_responsable else '',
                parte.get_canal_principal_display() if parte.canal_principal else ''
            ])

        ws3.column_dimensions['A'].width = 35
        ws3.column_dimensions['B'].width = 30
        ws3.column_dimensions['C'].width = 30
        ws3.column_dimensions['D'].width = 30
        ws3.column_dimensions['E'].width = 30

        # ====================================================================
        # HOJA 4: MATRIZ CONSOLIDADA
        # ====================================================================
        ws4 = wb.create_sheet('Matriz Consolidada')
        ws4.append([
            'GRUPO',
            'SUBGRUPO',
            'PARTE INTERESADA',
            'TEMAS INTERÉS PI',
            'TEMAS INTERÉS EMPRESA',
            'IMPACTO PI→EMPRESA',
            'IMPACTO EMPRESA→PI',
            'NIVEL INTERÉS',
            'RESPONSABLE',
            'CARGO',
            'ÁREA',
            'CANAL',
            'SST',
            'AMBIENTAL',
            'CALIDAD',
            'PESV'
        ])

        for cell in ws4[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for parte in queryset:
            ws4.append([
                parte.tipo.grupo.nombre if parte.tipo and parte.tipo.grupo else '',
                parte.tipo.nombre if parte.tipo else '',
                parte.nombre,
                parte.temas_interes_pi or '',
                parte.temas_interes_empresa or '',
                parte.get_nivel_influencia_pi_display(),
                parte.get_nivel_influencia_empresa_display(),
                parte.get_nivel_interes_display(),
                parte.responsable_empresa_nombre or '',
                parte.cargo_responsable.name if parte.cargo_responsable else '',
                parte.area_responsable.name if parte.area_responsable else '',
                parte.get_canal_principal_display() if parte.canal_principal else '',
                'Sí' if parte.relacionado_sst else 'No',
                'Sí' if parte.relacionado_ambiental else 'No',
                'Sí' if parte.relacionado_calidad else 'No',
                'Sí' if parte.relacionado_pesv else 'No'
            ])

        # Anchos
        for col in range(1, 17):
            ws4.column_dimensions[get_column_letter(col)].width = 20

        # ====================================================================
        # GUARDAR Y RETORNAR
        # ====================================================================
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Matriz_Partes_Interesadas.xlsx"'

        return response

    @action(detail=False, methods=['get'], url_path='plantilla-importacion')
    def plantilla_importacion(self, request) -> Response:
        """
        Descarga plantilla Excel profesional para importación masiva.

        Patrón unificado: igual que proveedores, clientes, cargos.
        Hoja 1: Plantilla (headers + ejemplo + notas + 500 filas vacías)
        Hoja 2: Referencia (grupos, tipos, niveles, canales existentes)
        """
        from .import_partes_interesadas_utils import generate_partes_interesadas_template

        # Obtener catálogos para la hoja de referencia
        grupos = list(
            GrupoParteInteresada.objects.filter(
                is_active=True
            ).values('nombre').order_by('orden', 'nombre')
        )
        tipos = list(
            TipoParteInteresada.objects.filter(
                is_active=True
            ).values('grupo__nombre', 'nombre').order_by('grupo__orden', 'orden', 'nombre')
        )
        partes_existentes = list(
            ParteInteresada.objects.filter(
                is_active=True
            ).select_related('tipo', 'tipo__grupo').values(
                'nombre', 'tipo__grupo__nombre', 'tipo__nombre'
            ).order_by('nombre')[:200]
        )

        content = generate_partes_interesadas_template(
            grupos=grupos,
            tipos=tipos,
            partes_existentes=partes_existentes,
        )

        response = HttpResponse(
            content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            'attachment; filename="Plantilla_Partes_Interesadas.xlsx"'
        )
        return response

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser], url_path='import-excel')
    def import_excel(self, request) -> Response:
        """
        Importa partes interesadas desde Excel.

        Acepta:
        - Plantilla nueva (hoja "Partes Interesadas", datos desde fila 4)
        - Legacy F-GD-04 (hoja "Identificación" o "Matriz Consolidada", datos desde fila 2)

        Campo: 'archivo' (nuevo) o 'file' (legacy)
        """
        import logging
        logger = logging.getLogger(__name__)

        file_obj = request.FILES.get('archivo') or request.FILES.get('file')

        if not file_obj:
            return Response(
                {'error': 'No se recibió ningún archivo. Use el campo "archivo".'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
        except Exception as e:
            return Response(
                {'error': f'Error al leer el archivo Excel: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Detectar formato: nuevo (plantilla unificada) vs legacy (F-GD-04) ──
        use_new_format = 'Partes Interesadas' in wb.sheetnames

        if use_new_format:
            return self._import_new_format(wb, request)
        elif 'Identificación' in wb.sheetnames or 'Matriz Consolidada' in wb.sheetnames:
            return self._import_legacy_format(wb, request)
        else:
            return Response(
                {
                    'error': (
                        'El archivo debe contener la hoja "Partes Interesadas" '
                        '(plantilla nueva) o "Identificación"/"Matriz Consolidada" '
                        '(formato anterior).'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

    # ── Importación formato NUEVO (plantilla unificada, fila 4+) ─────────
    def _import_new_format(self, wb, request):
        """Importa usando la plantilla unificada (parse desde fila 4)."""
        from .import_partes_interesadas_utils import parse_partes_interesadas_excel
        from .import_partes_interesadas_serializer import ParteInteresadaImportRowSerializer
        from apps.core.base_models.mixins import get_tenant_empresa
        import io

        # Guardar a bytes para el parser
        output = io.BytesIO()
        wb.save(output)
        file_bytes = output.getvalue()

        rows = parse_partes_interesadas_excel(file_bytes)

        if not rows:
            return Response(
                {'error': 'El archivo no contiene datos (los datos deben empezar en la fila 4).'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        empresa = get_tenant_empresa()
        created_count = 0
        updated_count = 0
        errors = []

        with transaction.atomic():
            for row_data in rows:
                fila = row_data.pop('_fila', '?')
                nombre_display = str(row_data.get('nombre', '')).strip()

                serializer = ParteInteresadaImportRowSerializer(data=row_data)
                if not serializer.is_valid():
                    errors.append({
                        'fila': fila,
                        'nombre': nombre_display,
                        'errores': serializer.errors,
                    })
                    continue

                vd = serializer.validated_data
                try:
                    parte, created = ParteInteresada.objects.update_or_create(
                        nombre=vd['nombre'],
                        empresa=empresa,
                        defaults={
                            'tipo': vd['_tipo'],
                            'descripcion': vd.get('descripcion', ''),
                            'representante': vd.get('representante', ''),
                            'cargo_representante': vd.get('cargo_representante', ''),
                            'email': vd.get('email', ''),
                            'telefono': vd.get('telefono', ''),
                            'direccion': vd.get('direccion', ''),
                            'sitio_web': vd.get('sitio_web', ''),
                            'temas_interes_pi': vd.get('temas_interes_pi', ''),
                            'temas_interes_empresa': vd.get('temas_interes_empresa', ''),
                            'nivel_influencia_pi': vd['nivel_influencia_pi'],
                            'nivel_influencia_empresa': vd['nivel_influencia_empresa'],
                            'nivel_interes': vd['nivel_interes'],
                            'canal_principal': vd['canal_principal'],
                            'frecuencia_comunicacion': vd.get('frecuencia_comunicacion', 'mensual'),
                            'necesidades': vd.get('necesidades', ''),
                            'expectativas': vd.get('expectativas', ''),
                            'requisitos_pertinentes': vd.get('requisitos_pertinentes', ''),
                            'es_requisito_legal': vd.get('es_requisito_legal', False),
                            'relacionado_sst': vd['relacionado_sst'],
                            'relacionado_ambiental': vd['relacionado_ambiental'],
                            'relacionado_calidad': vd['relacionado_calidad'],
                            'relacionado_pesv': vd['relacionado_pesv'],
                        },
                    )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                except Exception as e:
                    errors.append({
                        'fila': fila,
                        'nombre': nombre_display,
                        'errores': {'detalle': str(e)},
                    })

        return Response({
            'message': f'Importación completada: {created_count} creadas, {updated_count} actualizadas',
            'created': created_count,
            'updated': updated_count,
            'errors': errors,
            'total_procesadas': created_count + updated_count,
            'total_errores': len(errors),
        })

    # ── Importación formato LEGACY (F-GD-04, fila 2+) ────────────────────
    def _import_legacy_format(self, wb, request):
        """Importa desde el formato antiguo F-GD-04 (compatibilidad)."""
        from apps.core.base_models.mixins import get_tenant_empresa

        if 'Identificación' in wb.sheetnames:
            ws = wb['Identificación']
        else:
            ws = wb['Matriz Consolidada']

        # Leer headers (fila 1)
        headers = [cell.value for cell in ws[1]]

        col_map = {}
        for idx, header in enumerate(headers, start=1):
            if header:
                header_lower = str(header).strip().lower()
                col_map[header_lower] = idx

        created_count = 0
        updated_count = 0
        errors = []

        with transaction.atomic():
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    def get_col(name_variants):
                        for variant in name_variants:
                            if variant in col_map:
                                idx = col_map[variant] - 1
                                if idx < len(row):
                                    val = row[idx]
                                    return str(val).strip() if val else None
                        return None

                    grupo_nombre = get_col(['grupo'])
                    subgrupo_nombre = get_col(['subgrupo', 'subgrupo (tipo)', 'tipo'])
                    nombre = get_col(['nombre parte interesada', 'parte interesada', 'nombre'])
                    descripcion = get_col(['descripción', 'descripcion'])
                    representante = get_col(['representante'])
                    temas_pi = get_col(['temas de interés para la pi', 'temas interes pi', 'temas interés pi'])
                    temas_empresa = get_col(['temas de interés para la empresa', 'temas interes empresa', 'temas interés empresa'])

                    if not nombre:
                        continue

                    if grupo_nombre:
                        grupo, _ = GrupoParteInteresada.objects.get_or_create(
                            nombre=grupo_nombre,
                            defaults={
                                'codigo': grupo_nombre.upper().replace(' ', '_')[:30],
                                'es_sistema': False,
                            }
                        )
                    else:
                        grupo = GrupoParteInteresada.objects.first()

                    if subgrupo_nombre and grupo:
                        tipo, _ = TipoParteInteresada.objects.get_or_create(
                            nombre=subgrupo_nombre,
                            grupo=grupo,
                            defaults={
                                'codigo': f"{grupo.codigo}_{subgrupo_nombre.upper().replace(' ', '_')[:20]}"[:30],
                                'categoria': 'externo',
                                'es_sistema': False,
                            }
                        )
                    else:
                        tipo = TipoParteInteresada.objects.filter(grupo=grupo).first()

                    parte, created = ParteInteresada.objects.update_or_create(
                        nombre=nombre,
                        empresa=get_tenant_empresa(),
                        defaults={
                            'tipo': tipo,
                            'descripcion': descripcion or '',
                            'representante': representante or '',
                            'temas_interes_pi': temas_pi or '',
                            'temas_interes_empresa': temas_empresa or '',
                        },
                    )

                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                except Exception as e:
                    errors.append({'fila': row_idx, 'error': str(e)})

        return Response({
            'message': f'Importación completada: {created_count} creadas, {updated_count} actualizadas',
            'created': created_count,
            'updated': updated_count,
            'errors': errors,
            'total_procesadas': created_count + updated_count,
            'total_errores': len(errors),
        })


class RequisitoParteInteresadaViewSet(EmpresaAutoAssignMixin, StandardViewSetMixin, viewsets.ModelViewSet):
    """ViewSet para gestión de Requisitos de Partes Interesadas."""

    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'partes_interesadas'

    queryset = RequisitoParteInteresada.objects.select_related('parte_interesada').all()
    serializer_class = RequisitoParteInteresadaSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = [
        'empresa', 'parte_interesada', 'tipo', 'prioridad', 'cumple', 'is_active'
    ]
    search_fields = ['descripcion', 'como_se_aborda', 'proceso_relacionado']
    ordering_fields = ['prioridad', 'tipo', 'created_at']
    ordering = ['-prioridad', 'tipo']


class MatrizComunicacionViewSet(EmpresaAutoAssignMixin, StandardViewSetMixin, viewsets.ModelViewSet):
    """ViewSet para gestión de Matriz de Comunicación."""

    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'partes_interesadas'

    queryset = MatrizComunicacion.objects.select_related(
        'parte_interesada', 'responsable'
    ).all()
    serializer_class = MatrizComunicacionSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = [
        'empresa', 'parte_interesada', 'cuando_comunicar', 'como_comunicar',
        'es_obligatoria', 'is_active'
    ]
    search_fields = ['que_comunicar', 'registro_evidencia']
    ordering_fields = ['parte_interesada', 'cuando_comunicar', 'created_at']
    ordering = ['parte_interesada', 'cuando_comunicar']
