"""
Views para Gestión Documental - Gestión Estratégica (N1)
Sistema de gestión documental transversal con control de versiones.

Migrado desde: apps.hseq_management.sistema_documental
NOTA: FirmaDocumentoViewSet ha sido ELIMINADO.
Las firmas digitales usan los endpoints de workflow_engine.firma_digital

Modelos migrados a TenantModel (SoftDeleteManager auto-filtra is_deleted).
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from apps.core.permissions import GranularActionPermission
from django.utils import timezone
from django.db.models import Q, Count
from django.contrib.contenttypes.models import ContentType
from apps.core.mixins import ExportMixin
from apps.core.base_models.mixins import get_tenant_empresa
from apps.audit_system.services import AuditSystemService
from .mixins import check_acceso_documento, verificar_acceso_documento
from .models import (
    TipoDocumento,
    PlantillaDocumento,
    Documento,
    VersionDocumento,
    CampoFormulario,
    ControlDocumental
)
from .serializers import (
    TipoDocumentoListSerializer,
    TipoDocumentoDetailSerializer,
    PlantillaDocumentoListSerializer,
    PlantillaDocumentoDetailSerializer,
    DocumentoListSerializer,
    DocumentoDetailSerializer,
    VersionDocumentoListSerializer,
    VersionDocumentoDetailSerializer,
    CampoFormularioListSerializer,
    CampoFormularioDetailSerializer,
    ControlDocumentalListSerializer,
    ControlDocumentalDetailSerializer,
)
from .services import DocumentoService


# =============================================================================
# HELPERS — Detección de duplicados (H-GD-A5)
# =============================================================================

def _calcular_hash_archivo(archivo):
    """Calcula SHA-256 de un UploadedFile leyendo en chunks (sin agotar memoria).

    Restaura el puntero del archivo a 0 para que el upload posterior funcione.
    """
    import hashlib

    sha256 = hashlib.sha256()
    archivo.seek(0)
    for chunk in archivo.chunks():
        sha256.update(chunk)
    archivo.seek(0)
    return sha256.hexdigest()


def _buscar_duplicado_por_hash(hash_sha256):
    """Busca un Documento del tenant actual que tenga el mismo hash.

    Retorna el primer match o None. Ignora documentos eliminados (soft delete
    via SoftDeleteManager).
    """
    if not hash_sha256:
        return None
    return Documento.objects.filter(archivo_hash_sha256=hash_sha256).first()


def _serializar_duplicado(documento):
    """Construye el payload 409 con datos esenciales del documento existente."""
    return {
        'id': documento.id,
        'codigo': documento.codigo,
        'titulo': documento.titulo,
        'fecha_creacion': (
            documento.fecha_creacion.isoformat()
            if documento.fecha_creacion else None
        ),
    }


# =============================================================================
# HELPERS — Distribución RBAC
# =============================================================================

def _auto_distribuir_documento(documento, asignado_por, empresa, fecha_limite=None):
    """
    Crea AceptacionDocumental para todos los usuarios que correspondan según
    la configuración RBAC del documento (aplica_a_todos / cargos_distribucion).

    Solo actúa si el documento tiene al menos uno de esos flags activo.
    Omite silenciosamente duplicados (get_or_create).

    Returns:
        int — número de nuevas AceptacionDocumental creadas.
    """
    from .models import AceptacionDocumental
    from django.contrib.auth import get_user_model

    if not documento.aplica_a_todos and not documento.cargos_distribucion.exists():
        return 0

    User = get_user_model()
    base_qs = User.objects.filter(is_active=True, deleted_at__isnull=True)

    if documento.aplica_a_todos:
        usuarios = base_qs
    else:
        cargo_ids = documento.cargos_distribucion.values_list('id', flat=True)
        usuarios = base_qs.filter(cargo_id__in=cargo_ids)

    creados = 0

    for usuario in usuarios.iterator(chunk_size=200):
        _, was_created = AceptacionDocumental.objects.get_or_create(
            documento=documento,
            version_documento=documento.version_actual,
            usuario=usuario,
            defaults={
                'estado': 'PENDIENTE',
                'asignado_por': asignado_por,
                'fecha_limite': fecha_limite,
            },
        )
        if was_created:
            creados += 1

    return creados


class TipoDocumentoViewSet(ExportMixin, viewsets.ModelViewSet):
    """ViewSet para Tipos de Documento"""
    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'configuracion'
    export_fields = [('codigo', 'Código'), ('nombre', 'Nombre'), ('nivel_documento', 'Nivel'), ('prefijo_codigo', 'Prefijo'), ('requiere_aprobacion', 'Req. Aprobación'), ('requiere_firma', 'Req. Firma')]
    export_filename = 'tipos_documento'

    def get_serializer_class(self):
        if self.action == 'list':
            return TipoDocumentoListSerializer
        return TipoDocumentoDetailSerializer

    def get_queryset(self):
        queryset = TipoDocumento.objects.select_related('created_by')

        # Filtros
        nivel = self.request.query_params.get('nivel')
        if nivel:
            queryset = queryset.filter(nivel_documento=nivel)

        return queryset.order_by('orden', 'codigo')

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=False, methods=['get'])
    def activos(self, request):
        """Obtiene tipos de documento activos (SoftDeleteManager ya filtra eliminados)"""
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class PlantillaDocumentoViewSet(viewsets.ModelViewSet):
    """ViewSet para Plantillas de Documento"""
    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'configuracion'

    def get_serializer_class(self):
        if self.action == 'list':
            return PlantillaDocumentoListSerializer
        return PlantillaDocumentoDetailSerializer

    def get_queryset(self):
        queryset = PlantillaDocumento.objects.select_related(
            'tipo_documento', 'created_by'
        )

        # Filtros
        tipo_documento = self.request.query_params.get('tipo_documento')
        if tipo_documento:
            queryset = queryset.filter(tipo_documento_id=tipo_documento)

        estado = self.request.query_params.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)

        tipo_plantilla = self.request.query_params.get('tipo_plantilla')
        if tipo_plantilla:
            queryset = queryset.filter(tipo_plantilla=tipo_plantilla)

        return queryset.order_by('-es_por_defecto', 'nombre')

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=True, methods=['post'])
    def activar(self, request, pk=None):
        """Activa una plantilla"""
        plantilla = self.get_object()
        plantilla.estado = 'ACTIVA'
        plantilla.save()
        return Response(PlantillaDocumentoDetailSerializer(plantilla).data)

    @action(detail=True, methods=['post'], url_path='marcar-obsoleta')
    def marcar_obsoleta(self, request, pk=None):
        """Marca plantilla como obsoleta"""
        plantilla = self.get_object()
        plantilla.estado = 'OBSOLETA'
        plantilla.save()
        return Response(PlantillaDocumentoDetailSerializer(plantilla).data)

    @action(detail=True, methods=['post'], url_path='establecer-por-defecto')
    def establecer_por_defecto(self, request, pk=None):
        """Establece plantilla como predeterminada"""
        plantilla = self.get_object()

        # Quitar flag de otras plantillas del mismo tipo
        PlantillaDocumento.objects.filter(
            tipo_documento=plantilla.tipo_documento,
            es_por_defecto=True
        ).update(es_por_defecto=False)

        plantilla.es_por_defecto = True
        plantilla.save()

        return Response(PlantillaDocumentoDetailSerializer(plantilla).data)

    @action(detail=True, methods=['get'], url_path='resolver-firmantes')
    def resolver_firmantes(self, request, pk=None):
        """
        Preview: resuelve firmantes_por_defecto a usuarios actuales.
        Muestra quién firmaría si se crea un documento hoy desde esta plantilla.
        """
        from django.apps import apps
        plantilla = self.get_object()
        firmantes_config = plantilla.firmantes_por_defecto or []

        if not firmantes_config:
            return Response({
                'firmantes': [],
                'mensaje': 'Esta plantilla no tiene firmantes por defecto configurados.',
            })

        Cargo = apps.get_model('core', 'Cargo')
        User = apps.get_model('core', 'User')
        resultado = []

        for config in firmantes_config:
            cargo_code = config.get('cargo_code', '')
            rol_firma = config.get('rol_firma', '')
            item = {
                'rol_firma': rol_firma,
                'cargo_code': cargo_code,
                'orden': config.get('orden', 0),
                'es_requerido': config.get('es_requerido', True),
                'cargo_nombre': None,
                'usuario_nombre': None,
                'usuario_id': None,
                'resuelto': False,
                'warning': None,
            }

            try:
                cargo = Cargo.objects.get(code=cargo_code, is_active=True)
                item['cargo_nombre'] = cargo.name
            except Cargo.DoesNotExist:
                item['warning'] = f'Cargo "{cargo_code}" no encontrado'
                resultado.append(item)
                continue

            usuario = User.objects.filter(
                cargo=cargo, is_active=True, deleted_at__isnull=True
            ).first()

            if usuario:
                item['usuario_nombre'] = usuario.get_full_name()
                item['usuario_id'] = usuario.pk
                item['resuelto'] = True
            else:
                item['warning'] = f'Sin usuario activo con cargo "{cargo.name}"'

            resultado.append(item)

        return Response({
            'firmantes': resultado,
            'total': len(resultado),
            'resueltos': sum(1 for f in resultado if f['resuelto']),
        })


class CampoFormularioViewSet(viewsets.ModelViewSet):
    """ViewSet para Campos de Formulario"""
    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'configuracion'

    def get_serializer_class(self):
        if self.action == 'list':
            return CampoFormularioListSerializer
        return CampoFormularioDetailSerializer

    def get_queryset(self):
        queryset = CampoFormulario.objects.select_related(
            'plantilla', 'tipo_documento', 'created_by'
        )

        # Filtros
        plantilla = self.request.query_params.get('plantilla')
        if plantilla:
            queryset = queryset.filter(plantilla_id=plantilla)

        tipo_documento = self.request.query_params.get('tipo_documento')
        if tipo_documento:
            queryset = queryset.filter(tipo_documento_id=tipo_documento)

        tipo_campo = self.request.query_params.get('tipo_campo')
        if tipo_campo:
            queryset = queryset.filter(tipo_campo=tipo_campo)

        return queryset.order_by('orden', 'nombre_campo')

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=False, methods=['post'])
    def reordenar(self, request):
        """Reordena campos de formulario.
        Body: {campos: [{id, orden}, ...]}
        """
        campos = request.data.get('campos', [])
        for item in campos:
            CampoFormulario.objects.filter(id=item['id']).update(orden=item['orden'])
        return Response({'message': 'Campos reordenados exitosamente'})


class DocumentoViewSet(ExportMixin, viewsets.ModelViewSet):
    """
    ViewSet para Documentos.

    Las firmas digitales se manejan a través del módulo workflow_engine.firma_digital.
    Use el método get_firmas_digitales() del documento o el endpoint /firmas/ para acceder a ellas.
    """
    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'repositorio'
    export_fields = [('codigo', 'Código'), ('titulo', 'Título'), ('tipo_documento__nombre', 'Tipo'), ('estado', 'Estado'), ('version_actual', 'Versión'), ('clasificacion', 'Clasificación'), ('fecha_publicacion', 'Fecha Publicación'), ('fecha_vigencia', 'Fecha Vigencia')]
    export_filename = 'documentos'

    # Self-service: endpoints accesibles sin RBAC (Mi Portal, Dashboard)
    SELF_SERVICE_ACTIONS = {'habeas_data_status', 'mis_lecturas_count'}

    def get_permissions(self):
        if self.action in self.SELF_SERVICE_ACTIONS:
            return [IsAuthenticated()]
        return super().get_permissions()

    def get_serializer_class(self):
        if self.action == 'list':
            return DocumentoListSerializer
        return DocumentoDetailSerializer

    def get_queryset(self):
        queryset = Documento.objects.select_related(
            'tipo_documento', 'plantilla', 'elaborado_por',
            'revisado_por', 'aprobado_por', 'documento_padre'
        ).prefetch_related('versiones')

        # Filtros
        tipo_documento = self.request.query_params.get('tipo_documento')
        if tipo_documento:
            queryset = queryset.filter(tipo_documento_id=tipo_documento)

        tipo_documento_codigo = self.request.query_params.get('tipo_documento_codigo')
        if tipo_documento_codigo:
            queryset = queryset.filter(tipo_documento__codigo=tipo_documento_codigo)

        estado = self.request.query_params.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)

        clasificacion = self.request.query_params.get('clasificacion')
        if clasificacion:
            queryset = queryset.filter(clasificacion=clasificacion)

        # Búsqueda full-text (Sprint 6.2 — PostgreSQL tsvector)
        # H-GD-A2: incluye texto_extraido (peso D) sólo si la confianza OCR
        # es ≥ 0.7. Documentos con OCR pobre se buscan vía LIKE/icontains
        # y por los demás campos del SearchVector, pero NO contaminan el
        # ranking tsvector con texto ruidoso.
        buscar = self.request.query_params.get('buscar')
        if buscar:
            if len(buscar) >= 3:
                from django.contrib.postgres.search import (
                    SearchQuery,
                    SearchRank,
                    SearchVector,
                )
                from django.db.models import (
                    Case,
                    F,
                    TextField,
                    Value,
                    When,
                )
                from django.db.models.functions import Coalesce

                # Vector base — siempre presente
                vector = (
                    SearchVector('codigo', weight='A', config='spanish') +
                    SearchVector('codigo_legacy', weight='A', config='spanish') +
                    SearchVector('titulo', weight='A', config='spanish') +
                    SearchVector('resumen', weight='B', config='spanish') +
                    SearchVector('proceso__name', weight='C', config='spanish') +
                    SearchVector('tipo_documento__nombre', weight='C', config='spanish')
                )

                # texto_extraido condicional — sólo se indexa si la
                # confianza OCR es buena (>= 0.7). Para documentos con
                # OCR malo (o sin metadatos), anotamos cadena vacía para
                # que el SearchVector no genere ruido en el ranking.
                # Usamos Q directo sobre ocr_metadatos__confianza__gte
                # (lookup nativo de JSONField) para evitar Cast manual.
                queryset = queryset.annotate(
                    _texto_indexable=Case(
                        When(
                            ocr_metadatos__confianza__gte=0.7,
                            then=Coalesce(F('texto_extraido'), Value('')),
                        ),
                        default=Value(''),
                        output_field=TextField(),
                    ),
                )
                vector = vector + SearchVector(
                    '_texto_indexable', weight='D', config='spanish'
                )

                query = SearchQuery(buscar, config='spanish')
                queryset = (
                    queryset
                    .annotate(rank=SearchRank(vector, query))
                    .filter(
                        Q(rank__gt=0.01) |
                        Q(codigo__icontains=buscar) |
                        Q(titulo__icontains=buscar) |
                        Q(codigo_legacy__icontains=buscar)
                    )
                    .order_by('-rank', '-fecha_publicacion', 'codigo')
                )
                return queryset
            else:
                queryset = queryset.filter(
                    Q(codigo__icontains=buscar) |
                    Q(titulo__icontains=buscar) |
                    Q(codigo_legacy__icontains=buscar)
                )

        return queryset.order_by('-created_at', '-fecha_publicacion', 'codigo')

    def retrieve(self, request, *args, **kwargs):
        """Verifica acceso a documentos CONFIDENCIAL/RESTRINGIDO y registra VISTA / ACCESO_DENEGADO."""
        from .mixins import check_acceso_documento
        from .services import EventoDocumentalService
        from rest_framework.exceptions import PermissionDenied

        instance = self.get_object()
        if not check_acceso_documento(request.user, instance):
            EventoDocumentalService.registrar(
                documento=instance,
                usuario=request.user,
                tipo='ACCESO_DENEGADO',
                request=request,
                metadatos={
                    'origen': 'documento_retrieve',
                    'clasificacion': instance.clasificacion,
                },
            )
            raise PermissionDenied(
                'No tiene permiso para acceder a este documento. '
                'Los documentos con clasificación Confidencial o Restringido '
                'requieren autorización explícita.'
            )

        EventoDocumentalService.registrar(
            documento=instance,
            usuario=request.user,
            tipo='VISTA',
            request=request,
            metadatos={'origen': 'documento_retrieve'},
        )
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def perform_create(self, serializer):
        from rest_framework.exceptions import PermissionDenied
        user = self.request.user

        # Solo usuarios con cargo asignado pueden crear documentos SGI.
        # EXCEPCIÓN TEMPORAL: superadmin puede crear mientras se habilita la
        # transición a consultores/proveedores externos con cargo propio.
        # Los documentos tienen elaborado_por=superadmin como responsable provisional.
        if not user.is_superuser and not getattr(user, 'cargo', None):
            raise PermissionDenied(
                'Debes tener un cargo asignado para crear documentos. '
                'Los documentos SGI requieren un responsable con cargo en la organización.'
            )

        empresa = get_tenant_empresa()

        # Auto-generar codigo TIPO-PROCESO-NNN si no viene en el request
        codigo = serializer.validated_data.get('codigo')
        if not codigo:
            tipo_documento = serializer.validated_data.get('tipo_documento')
            proceso = serializer.validated_data.get('proceso')
            if tipo_documento:
                from .services import DocumentoService
                codigo = DocumentoService.generar_codigo(tipo_documento, proceso)

        documento = serializer.save(
            created_by=self.request.user,
            elaborado_por=self.request.user,
            codigo=codigo,
        )

        # Registrar en LogCambio (Centro de Control)
        AuditSystemService.log_cambio(
            self.request.user, documento, 'crear',
            {'creado': {'old': None, 'new': f'{documento.codigo} - {documento.titulo}'}},
            self.request,
        )

        # Auto-interpolar variables de empresa en contenido de plantilla
        if documento.plantilla and documento.contenido:
            from .services.documento_service import interpolar_variables_empresa
            documento.contenido = interpolar_variables_empresa(documento.contenido)
            documento.save(update_fields=['contenido', 'updated_at'])

    @action(detail=True, methods=['get'])
    def firmas(self, request, pk=None):
        """
        Obtiene las firmas digitales del documento.
        Las firmas están en workflow_engine.firma_digital via GenericForeignKey.
        """
        documento = self.get_object()
        verificar_acceso_documento(request.user, documento)
        from apps.workflow_engine.firma_digital.serializers import FirmaDigitalSerializer
        firmas = documento.get_firmas_digitales()
        return Response(FirmaDigitalSerializer(firmas, many=True).data)

    @action(detail=False, methods=['get'], url_path='content-type-id')
    def content_type_id(self, request):
        """Retorna el ContentType ID del modelo Documento para uso en FirmaDigital."""
        ct = ContentType.objects.get_for_model(Documento)
        return Response({
            'content_type_id': ct.id,
            'app_label': ct.app_label,
            'model': ct.model,
        })

    @action(detail=False, methods=['get'], url_path='habeas-data-status')
    def habeas_data_status(self, request):
        """
        Estado de la Política de Habeas Data del tenant.
        Usado por dashboard para mostrar banner si no está publicada.
        """
        empresa = get_tenant_empresa()
        if not empresa:
            return Response({'tiene_politica': False, 'estado': None})
        resultado = DocumentoService.verificar_habeas_data_publicada()
        return Response(resultado)

    @action(detail=False, methods=['get'], url_path='mis-lecturas-count')
    def mis_lecturas_count(self, request):
        """
        Cuenta lecturas obligatorias pendientes del usuario autenticado.
        Self-service: solo requiere IsAuthenticated.
        """
        count = DocumentoService.contar_lecturas_pendientes_obligatorias(request.user)
        return Response({'count': count})

    @action(detail=True, methods=['get'], url_path='estado-firmas')
    def estado_firmas(self, request, pk=None):
        """Retorna estado de firmas digitales del documento."""
        from .services import DocumentoService
        documento = self.get_object()
        verificar_acceso_documento(request.user, documento)
        estado = DocumentoService.obtener_estado_firmas(documento)
        return Response(estado)

    @action(detail=True, methods=['post'])
    def aprobar(self, request, pk=None):
        """Aprueba un documento (valida firmas si requiere_firma)."""
        from .services import DocumentoService
        documento = self.get_object()
        empresa = get_tenant_empresa()
        try:
            doc = DocumentoService.aprobar_documento(
                documento_id=documento.id,
                usuario=request.user,
                observaciones=request.data.get('observaciones', ''),
            )
            return Response(DocumentoDetailSerializer(doc).data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def publicar(self, request, pk=None):
        """
        Publica un documento aprobado y aplica distribución RBAC.

        Campos opcionales en request.data:
          lectura_obligatoria (bool)  — auto-asignar a futuros usuarios
          aplica_a_todos (bool)       — distribuir a TODOS los usuarios activos ahora
          cargo_ids (list[int])       — distribuir a usuarios con esos cargos ahora
          fecha_vigencia (str)        — fecha ISO de vigencia
          fecha_limite_lectura (str)  — fecha límite para lecturas auto-distribuidas
        """
        from .services import DocumentoService
        documento = self.get_object()
        empresa = get_tenant_empresa()

        # ── Actualizar flags de distribución antes de publicar ────────────────
        update_fields = ['updated_at']

        lectura_obligatoria = request.data.get('lectura_obligatoria')
        if lectura_obligatoria is not None:
            documento.lectura_obligatoria = bool(lectura_obligatoria)
            update_fields.append('lectura_obligatoria')

        aplica_a_todos = request.data.get('aplica_a_todos')
        if aplica_a_todos is not None:
            documento.aplica_a_todos = bool(aplica_a_todos)
            update_fields.append('aplica_a_todos')

        cargo_ids = request.data.get('cargo_ids') or []

        if update_fields != ['updated_at']:
            documento.save(update_fields=update_fields)

        # Persistir cargos_distribucion si se pasaron
        if cargo_ids:
            from django.apps import apps as dj_apps
            try:
                Cargo = dj_apps.get_model('core', 'Cargo')
                cargos = Cargo.objects.filter(id__in=cargo_ids, is_active=True)
                documento.cargos_distribucion.set(cargos)
            except LookupError:
                pass

        try:
            doc = DocumentoService.publicar_documento(
                documento_id=documento.id,
                usuario=request.user,
                fecha_vigencia=request.data.get('fecha_vigencia'),
            )
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        # ── Auto-distribución RBAC post-publicación ───────────────────────────
        fecha_limite_lectura = request.data.get('fecha_limite_lectura')
        distribuidos = _auto_distribuir_documento(doc, request.user, empresa, fecha_limite_lectura)
        if distribuidos:
            import logging
            logging.getLogger('gestion_documental').info(
                'Auto-distribución: %d lectura(s) asignadas para Documento %s',
                distribuidos, doc.id,
            )

        return Response(DocumentoDetailSerializer(doc).data)

    @action(detail=True, methods=['post'], url_path='marcar-obsoleto')
    def marcar_obsoleto(self, request, pk=None):
        """Marca documento como obsoleto."""
        from .services import DocumentoService
        documento = self.get_object()
        empresa = get_tenant_empresa()
        try:
            doc = DocumentoService.marcar_obsoleto(
                documento_id=documento.id,
                usuario=request.user,
                motivo=request.data.get('motivo', 'Documento obsoleto'),
                sustituto_id=request.data.get('documento_sustituto'),
            )
            return Response(DocumentoDetailSerializer(doc).data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], url_path='digitalizar')
    def digitalizar(self, request, pk=None):
        """
        Digitaliza un documento externo ingestado.

        Flujo:
        1. Valida que el documento sea externo (es_externo=True) y esté en BORRADOR.
        2. Marca el documento original como OBSOLETO (queda en Archivo como trazabilidad).
        3. Crea un nuevo BORRADOR con el contenido estructurado HTML generado desde
           las secciones recibidas, vinculado al original via documento_padre.

        Body:
        {
            "titulo": "Procedimiento de Compras",
            "secciones": [
                {"id": "objetivo", "label": "Objetivo", "contenido": "<p>...</p>"},
                ...
            ],
            "responsables_cargo_ids": [1, 2]   // IDs de Cargo
        }
        """
        from .services import DocumentoService
        documento = self.get_object()
        empresa = get_tenant_empresa()
        empresa_id = empresa.id if empresa else None

        # Validaciones
        if not documento.es_externo:
            return Response(
                {'error': 'Solo se pueden digitalizar documentos ingestados externamente.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if documento.estado != 'BORRADOR':
            return Response(
                {'error': 'Solo se pueden digitalizar documentos en estado BORRADOR.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        titulo = request.data.get('titulo', '').strip()
        secciones = request.data.get('secciones', [])
        responsables_cargo_ids = request.data.get('responsables_cargo_ids', [])

        if not titulo:
            return Response({'error': 'El campo "titulo" es obligatorio.'}, status=status.HTTP_400_BAD_REQUEST)
        if not secciones:
            return Response({'error': 'Debe incluir al menos una sección de contenido.'}, status=status.HTTP_400_BAD_REQUEST)

        # Construir contenido HTML desde secciones
        contenido_html = ''
        for seccion in secciones:
            label = seccion.get('label', '')
            contenido_seccion = seccion.get('contenido', '')
            if label or contenido_seccion:
                contenido_html += f'<h2>{label}</h2>\n{contenido_seccion}\n'

        # Generar código para el nuevo documento
        from .services import DocumentoService as DS
        codigo = DS.generar_codigo(documento.tipo_documento, empresa_id)

        # 1. Marcar original como OBSOLETO
        try:
            DS.marcar_obsoleto(
                documento_id=documento.id,
                usuario=request.user,
                empresa_id=empresa_id,
                motivo='Documento digitalizado — versión original archivada como referencia.',
                sustituto_id=None,
            )
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        # 2. Crear nuevo BORRADOR digitalizado
        nuevo = Documento.objects.create(
            titulo=titulo,
            tipo_documento=documento.tipo_documento,
            plantilla=documento.plantilla,
            clasificacion=documento.clasificacion,
            areas_aplicacion=documento.areas_aplicacion,
            contenido=contenido_html,
            estado='BORRADOR',
            es_externo=False,
            elaborado_por=request.user,
            documento_padre=documento,
            codigo=codigo,
        )

        # Asignar responsable_cargo si viene un solo cargo, o guardar lista en areas_aplicacion
        if responsables_cargo_ids:
            try:
                from apps.gestion_estrategica.organizacion.models import Cargo
                primer_cargo = Cargo.objects.get(pk=responsables_cargo_ids[0])
                nuevo.responsable_cargo = primer_cargo
                nuevo.save(update_fields=['responsable_cargo', 'updated_at'])
            except Exception:
                pass

        AuditSystemService.log_cambio(
            request.user, nuevo, 'crear',
            {'creado': {'old': None, 'new': f'{nuevo.codigo} - {nuevo.titulo} (digitalizado desde {documento.codigo})'}},
            request,
        )

        return Response(DocumentoDetailSerializer(nuevo).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='enviar-revision')
    def enviar_revision(self, request, pk=None):
        """Envía documento a revisión."""
        from .services import DocumentoService
        documento = self.get_object()
        empresa = get_tenant_empresa()
        try:
            doc = DocumentoService.enviar_a_revision(
                documento_id=documento.id,
                usuario=request.user,
                revisor_id=request.data.get('revisor_id'),
            )
            return Response(DocumentoDetailSerializer(doc).data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], url_path='devolver-borrador')
    def devolver_borrador(self, request, pk=None):
        """Devuelve documento de EN_REVISION a BORRADOR."""
        from .services import DocumentoService
        documento = self.get_object()
        empresa = get_tenant_empresa()
        try:
            doc = DocumentoService.devolver_a_borrador(
                documento_id=documento.id,
                usuario=request.user,
                motivo=request.data.get('motivo', ''),
            )
            return Response(DocumentoDetailSerializer(doc).data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], url_path='incrementar-descarga')
    def incrementar_descarga(self, request, pk=None):
        """Registra evento de descarga (el signal incrementa el contador)."""
        from .services import EventoDocumentalService

        documento = self.get_object()
        verificar_acceso_documento(request.user, documento)
        EventoDocumentalService.registrar(
            documento=documento,
            usuario=request.user,
            tipo='DESCARGA_PDF',
            request=request,
            metadatos={'origen': 'incrementar_descarga'},
        )
        documento.refresh_from_db(fields=['numero_descargas'])
        return Response({'descargas': documento.numero_descargas})

    @action(detail=True, methods=['post'], url_path='incrementar-impresion')
    def incrementar_impresion(self, request, pk=None):
        """Registra evento de impresión (el signal incrementa el contador)."""
        from .services import EventoDocumentalService

        documento = self.get_object()
        verificar_acceso_documento(request.user, documento)
        EventoDocumentalService.registrar(
            documento=documento,
            usuario=request.user,
            tipo='IMPRESION',
            request=request,
            metadatos={'origen': 'incrementar_impresion'},
        )
        documento.refresh_from_db(fields=['numero_impresiones'])
        return Response({'impresiones': documento.numero_impresiones})

    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Estadísticas del sistema documental para dashboard"""
        from .services import DocumentoService
        empresa = get_tenant_empresa()
        stats = DocumentoService.obtener_estadisticas()
        return Response(stats)

    @action(detail=False, methods=['get'], url_path='pendientes-revision')
    def pendientes_revision(self, request):
        """Documentos pendientes de revisión programada"""
        hoy = timezone.now().date()

        queryset = Documento.objects.filter(
            estado='PUBLICADO',
            fecha_revision_programada__lte=hoy
        )

        return Response(DocumentoListSerializer(queryset, many=True).data)

    @action(detail=False, methods=['get'], url_path='listado-maestro')
    def listado_maestro(self, request):
        """Listado maestro de documentos. Soporta ?formato=pdf|json (default json) y ?estado=PUBLICADO,OBSOLETO"""
        from django.http import HttpResponse
        from .exporters.pdf_generator import DocumentoPDFGenerator

        estados = request.query_params.getlist('estado') or ['PUBLICADO']
        queryset = (
            Documento.objects
            .filter(estado__in=estados)
            .select_related('tipo_documento', 'proceso')
            .order_by('tipo_documento__codigo', 'codigo')
        )

        formato = request.query_params.get('formato', 'json')

        if formato == 'pdf':
            empresa = get_tenant_empresa()
            generator = DocumentoPDFGenerator(empresa=empresa)
            try:
                pdf_buffer = generator.generate_listado_maestro_pdf(queryset, usuario=request.user)
            except ImportError as exc:
                return Response({'error': str(exc)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

            from datetime import date as _date
            filename = f"listado_maestro_{_date.today().strftime('%Y%m%d')}.pdf"
            response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        # JSON (default)
        listado = {}
        for doc in queryset:
            tipo = doc.tipo_documento.codigo if doc.tipo_documento else 'SIN_TIPO'
            if tipo not in listado:
                listado[tipo] = {
                    'tipo': doc.tipo_documento.nombre if doc.tipo_documento else 'Sin tipo',
                    'documentos': []
                }
            listado[tipo]['documentos'].append(DocumentoListSerializer(doc).data)

        return Response(listado)

    @action(detail=False, methods=['get'], url_path='cobertura-documental')
    def cobertura_documental(self, request):
        """
        Dashboard de cobertura documental: qué tipos tienen documentos,
        cuáles no, y qué workflows carecen de procedimiento documentado.
        Útil para auditorías ISO y revisión por la dirección.
        """
        from .services import DocumentoService as _DocumentoService
        cobertura = _DocumentoService.obtener_cobertura_documental()
        return Response(cobertura)

    # =========================================================================
    # ARCHIVOS ANEXOS — Upload, listado y eliminación de evidencias
    # =========================================================================

    ALLOWED_EXTENSIONS = {
        '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx',
        '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp',
        '.csv', '.txt', '.zip', '.rar', '.7z',
    }
    MAX_ANEXO_SIZE = 20 * 1024 * 1024  # 20 MB por archivo

    @action(detail=True, methods=['post'], url_path='subir-anexo')
    def subir_anexo(self, request, pk=None):
        """
        Sube un archivo anexo a un documento existente.
        Acepta multipart/form-data con campo 'archivo'.
        Almacena metadata en archivos_anexos (JSONField).
        """
        documento = self.get_object()
        verificar_acceso_documento(request.user, documento)
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response(
                {'error': 'Se requiere un archivo'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validar extensión
        import os
        ext = os.path.splitext(archivo.name)[1].lower()
        if ext not in self.ALLOWED_EXTENSIONS:
            return Response(
                {'error': f'Extensión {ext} no permitida. '
                          f'Permitidas: {", ".join(sorted(self.ALLOWED_EXTENSIONS))}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validar tamaño
        if archivo.size > self.MAX_ANEXO_SIZE:
            return Response(
                {'error': 'El archivo excede el tamaño máximo de 20 MB'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Verificar cuota de almacenamiento del tenant
        from utils.storage import check_storage_quota
        puede_subir, usado_gb, limite_gb = check_storage_quota(archivo.size)
        if not puede_subir:
            return Response(
                {
                    'error': (
                        f'Cuota de almacenamiento excedida: '
                        f'{usado_gb:.2f} GB de {limite_gb:.2f} GB usados.'
                    ),
                    'usado_gb': usado_gb,
                    'limite_gb': limite_gb,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Guardar archivo — TenantFileStorage añade el prefix del schema automáticamente
        from django.core.files.storage import default_storage
        from django.utils import timezone
        from utils.storage import _safe_extension
        import uuid

        ahora = timezone.now()
        ext = _safe_extension(archivo.name)
        filename = f'{uuid.uuid4().hex}.{ext}'
        path = f'documentos/anexos/{ahora.year}/{ahora.month:02d}/{filename}'
        saved_path = default_storage.save(path, archivo)

        # Agregar metadata al JSONField
        anexos = documento.archivos_anexos or []
        anexo_meta = {
            'id': uuid.uuid4().hex[:12],
            'nombre': archivo.name,
            'path': saved_path,
            'url': default_storage.url(saved_path),
            'tipo_mime': archivo.content_type or 'application/octet-stream',
            'tamaño': archivo.size,
            'extension': ext,
            'subido_por': request.user.id,
            'subido_por_nombre': request.user.get_full_name(),
            'fecha_subida': ahora.isoformat(),
            'descripcion': request.data.get('descripcion', ''),
        }
        anexos.append(anexo_meta)
        documento.archivos_anexos = anexos
        documento.save(update_fields=['archivos_anexos'])

        return Response(
            {'mensaje': 'Archivo anexo subido', 'anexo': anexo_meta},
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=['delete'], url_path='eliminar-anexo/(?P<anexo_id>[\\w]+)')
    def eliminar_anexo(self, request, pk=None, anexo_id=None):
        """Elimina un archivo anexo por su ID interno."""
        documento = self.get_object()
        verificar_acceso_documento(request.user, documento)
        anexos = documento.archivos_anexos or []

        anexo_encontrado = None
        nuevos_anexos = []
        for anexo in anexos:
            if anexo.get('id') == anexo_id:
                anexo_encontrado = anexo
            else:
                nuevos_anexos.append(anexo)

        if not anexo_encontrado:
            return Response(
                {'error': 'Anexo no encontrado'},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Eliminar archivo físico
        from django.core.files.storage import default_storage
        try:
            if anexo_encontrado.get('path'):
                default_storage.delete(anexo_encontrado['path'])
        except Exception:
            pass  # Si el archivo ya no existe, no es error

        documento.archivos_anexos = nuevos_anexos
        documento.save(update_fields=['archivos_anexos'])

        return Response({'mensaje': 'Anexo eliminado'})

    # =========================================================================
    # OCR — Fase 5: Ingesta, reprocesamiento y búsqueda full-text
    # =========================================================================

    @action(detail=False, methods=['post'], url_path='ingestar-externo')
    def ingestar_externo(self, request):
        """
        Ingesta de PDF externo: crea Documento + dispara OCR async.
        Acepta multipart/form-data con archivo PDF.
        """
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response(
                {'error': 'Se requiere un archivo PDF'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar tipo de archivo
        if not archivo.name.lower().endswith('.pdf'):
            return Response(
                {'error': 'Solo se aceptan archivos PDF'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar tamaño (50MB máximo)
        max_size = 50 * 1024 * 1024
        if archivo.size > max_size:
            return Response(
                {'error': 'El archivo excede el tamaño máximo de 50 MB'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar cuota de almacenamiento del tenant
        from utils.storage import check_storage_quota
        puede_subir, usado_gb, limite_gb = check_storage_quota(archivo.size)
        if not puede_subir:
            return Response(
                {
                    'error': (
                        f'Cuota de almacenamiento excedida: '
                        f'{usado_gb:.2f} GB de {limite_gb:.2f} GB usados.'
                    ),
                    'usado_gb': usado_gb,
                    'limite_gb': limite_gb,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        titulo = request.data.get('titulo', archivo.name.rsplit('.', 1)[0])
        tipo_documento_id = request.data.get('tipo_documento')
        clasificacion = request.data.get('clasificacion', 'INTERNO')

        if not tipo_documento_id:
            return Response(
                {'error': 'Se requiere tipo_documento'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # H-GD-A5: detectar duplicados por SHA-256 ANTES de crear el documento.
        archivo_hash = _calcular_hash_archivo(archivo)
        duplicado = _buscar_duplicado_por_hash(archivo_hash)
        if duplicado:
            return Response(
                {
                    'error': (
                        'Ya existe un documento con el mismo contenido en este '
                        'tenant. Use el documento existente o adóptelo desde su '
                        'historial.'
                    ),
                    'documento_existente': _serializar_duplicado(duplicado),
                },
                status=status.HTTP_409_CONFLICT,
            )

        empresa = get_tenant_empresa()

        # Generar código automático
        from .services import DocumentoService
        tipo_doc = TipoDocumento.objects.get(id=tipo_documento_id)
        codigo = DocumentoService.generar_codigo(tipo_doc)

        documento = Documento.objects.create(
            codigo=codigo,
            titulo=titulo,
            tipo_documento=tipo_doc,
            estado='BORRADOR',
            clasificacion=clasificacion,
            elaborado_por=request.user,
            archivo_original=archivo,
            archivo_hash_sha256=archivo_hash,
            es_externo=True,
            ocr_estado='PENDIENTE',
        )

        # Palabras clave desde request (opcional)
        palabras = request.data.get('palabras_clave')
        if palabras:
            import json
            try:
                documento.palabras_clave = json.loads(palabras) if isinstance(
                    palabras, str
                ) else palabras
                documento.save(update_fields=['palabras_clave'])
            except (json.JSONDecodeError, TypeError):
                pass

        # Disparar OCR async
        from django.db import connection
        from .tasks import procesar_ocr_documento
        procesar_ocr_documento.delay(documento.id, connection.schema_name)

        serializer = DocumentoDetailSerializer(documento)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='adoptar-pdf')
    def adoptar_pdf(self, request):
        """
        Camino B: Adoptar PDF externo al ciclo de firmas → publicación.
        Crea documento BORRADOR con es_externo=True, asigna código StrateKaz,
        opcionalmente guarda codigo_legacy, asigna firmantes y dispara OCR.

        multipart/form-data:
          archivo (File, PDF) — obligatorio
          tipo_documento (int) — obligatorio
          proceso (int, FK Area) — obligatorio
          titulo (str) — opcional, default = nombre del archivo
          clasificacion (str) — opcional, default INTERNO
          codigo_legacy (str) — opcional, código original de la empresa
          firmantes (JSON str) — opcional, lista de {usuario_id, rol}
        """
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response(
                {'error': 'Se requiere un archivo PDF'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Solo PDF — política absoluta
        if not archivo.name.lower().endswith('.pdf'):
            return Response(
                {'error': 'Solo se aceptan archivos PDF. Si su documento está en otro formato, conviértalo a PDF antes de subirlo.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar magic bytes PDF (%PDF)
        header = archivo.read(5)
        archivo.seek(0)
        if not header.startswith(b'%PDF'):
            return Response(
                {'error': 'El archivo no es un PDF válido.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar tamaño (10 MB máximo para adopción)
        max_size = 10 * 1024 * 1024
        if archivo.size > max_size:
            return Response(
                {'error': 'El archivo excede el tamaño máximo de 10 MB'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar cuota de almacenamiento
        from utils.storage import check_storage_quota
        puede_subir, usado_gb, limite_gb = check_storage_quota(archivo.size)
        if not puede_subir:
            return Response(
                {
                    'error': (
                        f'Cuota de almacenamiento excedida: '
                        f'{usado_gb:.2f} GB de {limite_gb:.2f} GB usados.'
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        tipo_documento_id = request.data.get('tipo_documento')
        proceso_id = request.data.get('proceso')
        if not tipo_documento_id or not proceso_id:
            return Response(
                {'error': 'Se requiere tipo_documento y proceso'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # H-GD-A5: detectar duplicados por SHA-256 ANTES de crear el documento.
        archivo_hash = _calcular_hash_archivo(archivo)
        duplicado = _buscar_duplicado_por_hash(archivo_hash)
        if duplicado:
            return Response(
                {
                    'error': (
                        'Ya existe un documento con el mismo contenido en este '
                        'tenant. Revise el existente antes de volver a adoptarlo.'
                    ),
                    'documento_existente': _serializar_duplicado(duplicado),
                },
                status=status.HTTP_409_CONFLICT,
            )

        empresa = get_tenant_empresa()
        tipo_doc = TipoDocumento.objects.get(id=tipo_documento_id)

        # Resolver proceso (FK Area)
        from apps.gestion_estrategica.organizacion.models import Area
        try:
            proceso = Area.objects.get(id=proceso_id)
        except Area.DoesNotExist:
            return Response(
                {'error': 'Proceso no encontrado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Generar código StrateKaz
        from .services import DocumentoService
        codigo = DocumentoService.generar_codigo(tipo_doc, proceso=proceso)

        titulo = request.data.get('titulo', archivo.name.rsplit('.', 1)[0])
        clasificacion = request.data.get('clasificacion', 'INTERNO')
        codigo_legacy = request.data.get('codigo_legacy', '').strip() or None

        documento = Documento.objects.create(
            codigo=codigo,
            titulo=titulo,
            tipo_documento=tipo_doc,
            proceso=proceso,
            estado='BORRADOR',
            clasificacion=clasificacion,
            elaborado_por=request.user,
            archivo_original=archivo,
            archivo_hash_sha256=archivo_hash,
            es_externo=True,
            codigo_legacy=codigo_legacy,
            ocr_estado='PENDIENTE',
        )

        # Asignar firmantes si se proporcionan
        import json
        firmantes_raw = request.data.get('firmantes')
        if firmantes_raw:
            try:
                firmantes = json.loads(firmantes_raw) if isinstance(firmantes_raw, str) else firmantes_raw
                self._crear_firmantes(documento, firmantes, empresa)
            except (json.JSONDecodeError, TypeError, ValueError) as e:
                # Documento ya creado, pero firmantes fallaron — no fatal
                import logging
                logging.getLogger('gestion_documental').warning(
                    'adoptar_pdf: firmantes no válidos para doc %s: %s', documento.id, e
                )

        # Disparar OCR async
        from django.db import connection
        from .tasks import procesar_ocr_documento
        procesar_ocr_documento.delay(documento.id, connection.schema_name)

        serializer = DocumentoDetailSerializer(documento)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def _crear_firmantes(self, documento, firmantes, empresa):
        """
        Crea instancias FirmaDigital para el documento adoptado.
        firmantes: [{usuario_id: int, rol: str}]
        Valida firma única por usuario.
        """
        from django.contrib.contenttypes.models import ContentType
        from django.apps import apps as dj_apps

        FirmaDigital = dj_apps.get_model('firma_digital', 'FirmaDigital')
        User = dj_apps.get_model('core', 'User')
        ct = ContentType.objects.get_for_model(documento)

        # Validar unicidad de usuarios
        usuario_ids = [f['usuario_id'] for f in firmantes]
        if len(usuario_ids) != len(set(usuario_ids)):
            raise ValueError(
                'Un mismo usuario no puede ocupar dos roles de firma en el mismo documento.'
            )

        for idx, f in enumerate(firmantes):
            usuario = User.objects.get(id=f['usuario_id'])
            rol = f.get('rol', 'APROBO')
            FirmaDigital.objects.create(
                content_type=ct,
                object_id=str(documento.id),
                firmante=usuario,
                rol=rol,
                orden=idx + 1,
                estado='PENDIENTE',
                empresa_id=empresa.id,
            )

    @action(detail=False, methods=['post'], url_path='ingestar-lote')
    def ingestar_lote(self, request):
        """
        Ingesta masiva de PDFs: crea múltiples documentos + OCR para cada uno.
        Acepta multipart/form-data con múltiples archivos en 'archivos'.
        Comparte tipo_documento y clasificacion para todo el lote.
        """
        archivos = request.FILES.getlist('archivos')
        if not archivos:
            return Response(
                {'error': 'Se requiere al menos un archivo PDF'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if len(archivos) > 20:
            return Response(
                {'error': 'Máximo 20 archivos por lote'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        tipo_documento_id = request.data.get('tipo_documento')
        clasificacion = request.data.get('clasificacion', 'INTERNO')
        if not tipo_documento_id:
            return Response(
                {'error': 'Se requiere tipo_documento'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        empresa = get_tenant_empresa()
        from .services import DocumentoService
        tipo_doc = TipoDocumento.objects.get(id=tipo_documento_id)

        from django.db import connection
        from .tasks import procesar_ocr_documento

        max_size = 50 * 1024 * 1024
        creados = []
        errores = []

        # Verificar cuota considerando el tamaño total del lote
        from utils.storage import check_storage_quota
        total_lote = sum(a.size for a in archivos)
        puede_subir, usado_gb, limite_gb = check_storage_quota(total_lote)
        if not puede_subir:
            return Response(
                {
                    'error': (
                        f'Cuota de almacenamiento excedida para este lote: '
                        f'{usado_gb:.2f} GB de {limite_gb:.2f} GB usados.'
                    ),
                    'usado_gb': usado_gb,
                    'limite_gb': limite_gb,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        for archivo in archivos:
            if not archivo.name.lower().endswith('.pdf'):
                errores.append({'archivo': archivo.name, 'error': 'No es PDF'})
                continue
            if archivo.size > max_size:
                errores.append({'archivo': archivo.name, 'error': 'Excede 50 MB'})
                continue

            try:
                codigo = DocumentoService.generar_codigo(tipo_doc)
                titulo = archivo.name.rsplit('.', 1)[0]
                documento = Documento.objects.create(
                    codigo=codigo,
                    titulo=titulo,
                    tipo_documento=tipo_doc,
                    estado='BORRADOR',
                    clasificacion=clasificacion,
                    elaborado_por=request.user,
                    archivo_original=archivo,
                    es_externo=True,
                    ocr_estado='PENDIENTE',
                )
                procesar_ocr_documento.delay(documento.id, connection.schema_name)
                creados.append({
                    'id': documento.id,
                    'codigo': codigo,
                    'titulo': titulo,
                })
            except Exception as e:
                errores.append({'archivo': archivo.name, 'error': str(e)})

        return Response({
            'creados': len(creados),
            'errores': len(errores),
            'documentos': creados,
            'detalle_errores': errores,
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='reprocesar-ocr')
    def reprocesar_ocr(self, request, pk=None):
        """Re-dispara OCR para un documento (útil si falló antes)."""
        documento = self.get_object()
        verificar_acceso_documento(request.user, documento)

        archivo = documento.archivo_original or documento.archivo_pdf
        if not archivo:
            return Response(
                {'error': 'El documento no tiene archivo PDF para procesar'},
                status=status.HTTP_400_BAD_REQUEST
            )

        documento.ocr_estado = 'PENDIENTE'
        documento.ocr_metadatos = {}
        documento.save(update_fields=['ocr_estado', 'ocr_metadatos'])

        from django.db import connection
        from .tasks import procesar_ocr_documento
        procesar_ocr_documento.delay(documento.id, connection.schema_name)

        return Response({
            'mensaje': 'OCR reprogramado. Recibirá una notificación al completar.',
            'ocr_estado': 'PENDIENTE',
        })

    @action(detail=False, methods=['get'], url_path='busqueda-texto')
    def busqueda_texto(self, request):
        """
        Búsqueda full-text usando PostgreSQL tsvector.
        Query param: q (mínimo 3 caracteres).
        """
        query = request.query_params.get('q', '').strip()
        if len(query) < 3:
            return Response(
                {'error': 'La búsqueda requiere al menos 3 caracteres'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from django.contrib.postgres.search import (
            SearchQuery, SearchRank, SearchVector
        )

        search_vector = SearchVector(
            'codigo', 'codigo_legacy', 'titulo', 'resumen', 'texto_extraido',
            config='spanish'
        )
        search_query = SearchQuery(query, config='spanish')

        resultados = (
            Documento.objects
            .annotate(
                relevancia=SearchRank(search_vector, search_query),
            )
            .filter(relevancia__gt=0.01)
            .order_by('-relevancia')[:50]
        )

        data = []
        for doc in resultados:
            # Extraer fragmento relevante del texto
            extracto = ''
            if doc.texto_extraido:
                # Buscar contexto alrededor de la primera ocurrencia
                texto_lower = doc.texto_extraido.lower()
                query_lower = query.lower()
                pos = texto_lower.find(query_lower)
                if pos >= 0:
                    start = max(0, pos - 100)
                    end = min(len(doc.texto_extraido), pos + len(query) + 100)
                    extracto = ('...' if start > 0 else '') + \
                        doc.texto_extraido[start:end] + \
                        ('...' if end < len(doc.texto_extraido) else '')

            data.append({
                'id': doc.id,
                'codigo': doc.codigo,
                'titulo': doc.titulo,
                'resumen': doc.resumen[:200] if doc.resumen else '',
                'estado': doc.estado,
                'estado_display': doc.get_estado_display(),
                'clasificacion': doc.clasificacion,
                'relevancia': round(float(doc.relevancia), 4),
                'texto_extracto': extracto,
                'ocr_estado': doc.ocr_estado,
                'es_externo': doc.es_externo,
                'codigo_legacy': doc.codigo_legacy or '',
            })

        return Response(data)

    # =========================================================================
    # SCORING — Fase 6: Score de cumplimiento heurístico
    # =========================================================================

    @action(detail=True, methods=['post'], url_path='calcular-score')
    def calcular_score(self, request, pk=None):
        """Calcula y guarda el score de cumplimiento del documento."""
        documento = self.get_object()
        from .services_scoring import ScoringService
        resultado = ScoringService.actualizar_score(documento)
        return Response(resultado)

    @action(detail=False, methods=['get'], url_path='score-resumen')
    def score_resumen(self, request):
        """Dashboard de scores: promedio, distribución, incompletos."""
        from .services_scoring import ScoringService
        resumen = ScoringService.obtener_resumen(self.get_queryset())
        return Response(resumen)

    # =========================================================================
    # SELLADO PDF — Mejora 2: Firma digital X.509 con pyHanko
    # =========================================================================

    @action(detail=True, methods=['post'], url_path='sellar-pdf')
    def sellar_pdf(self, request, pk=None):
        """Inicia sellado PDF con firma digital X.509 via Celery."""
        documento = self.get_object()
        verificar_acceso_documento(request.user, documento)

        if documento.estado != 'PUBLICADO':
            return Response(
                {'error': 'Solo se pueden sellar documentos PUBLICADOS'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if documento.sellado_estado not in ('NO_APLICA', 'ERROR'):
            return Response(
                {'error': f'El documento ya está en estado de sellado: {documento.sellado_estado}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        documento.sellado_estado = 'PENDIENTE'
        documento.sellado_metadatos = {}
        documento.save(update_fields=['sellado_estado', 'sellado_metadatos'])

        from django.db import connection
        from .tasks import sellar_pdf_pyhanko
        sellar_pdf_pyhanko.delay(
            documento.id,
            connection.schema_name,
            request.user.id,
        )

        return Response(
            {
                'mensaje': 'Sellado iniciado. Recibirá una notificación al completar.',
                'sellado_estado': 'PENDIENTE',
            },
            status=status.HTTP_202_ACCEPTED,
        )

    @action(detail=True, methods=['get'], url_path='verificar-sellado')
    def verificar_sellado(self, request, pk=None):
        """Verifica integridad del PDF sellado recalculando SHA-256."""
        documento = self.get_object()
        verificar_acceso_documento(request.user, documento)

        if documento.sellado_estado != 'COMPLETADO':
            return Response(
                {'error': 'El documento no tiene un PDF sellado para verificar'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from .services.pdf_sealing import PDFSealingService
        resultado = PDFSealingService.verificar_integridad(documento)
        return Response(resultado)

    # =========================================================================
    # GOOGLE DRIVE — Fase 7: Exportación con Habeas Data
    # =========================================================================

    @action(detail=True, methods=['post'], url_path='exportar-drive')
    def exportar_drive(self, request, pk=None):
        """Exporta un documento individual a Google Drive."""
        documento = self.get_object()
        verificar_acceso_documento(request.user, documento)
        folder_id = request.data.get('folder_id')

        from django.apps import apps as django_apps
        IntegracionExterna = django_apps.get_model(
            'configuracion', 'IntegracionExterna'
        )
        integracion = IntegracionExterna.objects.filter(
            proveedor='GOOGLE_DRIVE',
            is_active=True,
        ).first()

        if not integracion:
            return Response(
                {'error': 'No hay integración de Google Drive activa configurada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from .services_drive import GoogleDriveService
            resultado = GoogleDriveService.exportar_documento(
                documento_id=documento.id,
                integracion_id=integracion.id,
                folder_id=folder_id,
                usuario=request.user,
            )
            return Response(resultado)
        except PermissionError as e:
            return Response({'error': str(e)}, status=status.HTTP_403_FORBIDDEN)
        except ImportError as e:
            return Response({'error': str(e)}, status=status.HTTP_501_NOT_IMPLEMENTED)

    @action(detail=False, methods=['post'], url_path='exportar-drive-lote')
    def exportar_drive_lote(self, request):
        """Exportación batch de documentos a Google Drive (async via Celery)."""
        from django.apps import apps as django_apps
        IntegracionExterna = django_apps.get_model(
            'configuracion', 'IntegracionExterna'
        )
        integracion = IntegracionExterna.objects.filter(
            proveedor='GOOGLE_DRIVE', is_active=True,
        ).first()

        if not integracion:
            return Response(
                {'error': 'No hay integración de Google Drive activa configurada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        empresa = get_tenant_empresa()
        from django.db import connection
        from .tasks import exportar_drive_lote
        exportar_drive_lote.delay(
            empresa_id=empresa.id,
            integracion_id=integracion.id,
            folder_id=request.data.get('folder_id'),
            usuario_id=request.user.id,
            filtros=request.data.get('filtros', {}),
            tenant_schema=connection.schema_name,
        )

        return Response({
            'mensaje': 'Exportación a Google Drive iniciada. '
                       'Recibirá una notificación al completar.',
        })

    def _get_client_ip(self, request):
        """Obtiene IP del cliente"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip


class VersionDocumentoViewSet(viewsets.ModelViewSet):
    """ViewSet para Versiones de Documento"""
    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'archivo'

    def get_serializer_class(self):
        if self.action == 'list':
            return VersionDocumentoListSerializer
        return VersionDocumentoDetailSerializer

    def get_queryset(self):
        queryset = VersionDocumento.objects.select_related(
            'documento', 'creado_por', 'aprobado_por'
        )

        documento = self.request.query_params.get('documento')
        if documento:
            queryset = queryset.filter(documento_id=documento)

        return queryset.order_by('-fecha_version')

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, creado_por=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=False, methods=['get'], url_path='por-documento')
    def por_documento(self, request):
        """Historial de versiones de un documento específico"""
        documento_id = request.query_params.get('documento_id')

        if not documento_id:
            return Response(
                {'error': 'Debe especificar documento_id'},
                status=status.HTTP_400_BAD_REQUEST
            )

        queryset = self.get_queryset().filter(documento_id=documento_id)
        serializer = VersionDocumentoListSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def comparar(self, request, pk=None):
        """Compara esta versión con la anterior"""
        version_actual = self.get_object()

        version_anterior = VersionDocumento.objects.filter(
            documento=version_actual.documento,
            fecha_version__lt=version_actual.fecha_version
        ).order_by('-fecha_version').first()

        if not version_anterior:
            return Response({
                'mensaje': 'No hay versión anterior para comparar',
                'version_actual': VersionDocumentoDetailSerializer(version_actual).data
            })

        return Response({
            'version_actual': VersionDocumentoDetailSerializer(version_actual).data,
            'version_anterior': VersionDocumentoDetailSerializer(version_anterior).data,
            'cambios': version_actual.cambios_detectados
        })


# NOTA: FirmaDocumentoViewSet ha sido ELIMINADO
# Las firmas digitales se manejan a través de:
# - workflow_engine.firma_digital.views.FirmaDigitalViewSet
# - Endpoints: /api/workflows/firma-digital/firmas/
# - Documento.get_firmas_digitales() para obtener firmas de un documento específico


class ControlDocumentalViewSet(viewsets.ModelViewSet):
    """ViewSet para Control Documental"""
    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'archivo'

    def get_serializer_class(self):
        if self.action == 'list':
            return ControlDocumentalListSerializer
        return ControlDocumentalDetailSerializer

    def get_queryset(self):
        queryset = ControlDocumental.objects.select_related(
            'documento', 'version_documento', 'documento_sustituto',
            'responsable_destruccion', 'created_by'
        ).prefetch_related('usuarios_distribucion')

        documento = self.request.query_params.get('documento')
        if documento:
            queryset = queryset.filter(documento_id=documento)

        tipo_control = self.request.query_params.get('tipo_control')
        if tipo_control:
            queryset = queryset.filter(tipo_control=tipo_control)

        return queryset.order_by('-fecha_distribucion')

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=True, methods=['post'], url_path='confirmar-recepcion')
    def confirmar_recepcion(self, request, pk=None):
        """Confirma recepción de documento por usuario"""
        control = self.get_object()

        confirmacion = {
            'usuario_id': request.user.id,
            'usuario_nombre': request.user.get_full_name(),
            'fecha': timezone.now().isoformat(),
            'ip_address': self._get_client_ip(request)
        }

        confirmaciones = control.confirmaciones_recepcion or []
        confirmaciones.append(confirmacion)
        control.confirmaciones_recepcion = confirmaciones
        control.save()

        return Response({'message': 'Recepción confirmada exitosamente'})

    @action(detail=False, methods=['get'], url_path='distribuciones-activas')
    def distribuciones_activas(self, request):
        """Controles de distribución activos"""
        queryset = self.get_queryset().filter(
            tipo_control='DISTRIBUCION',
            documento__estado='PUBLICADO'
        )
        return Response(ControlDocumentalListSerializer(queryset, many=True).data)

    @action(detail=False, methods=['get'], url_path='documentos-obsoletos')
    def documentos_obsoletos(self, request):
        """Controles de documentos retirados/obsoletos"""
        queryset = self.get_queryset().filter(
            tipo_control='RETIRO'
        )
        return Response(ControlDocumentalListSerializer(queryset, many=True).data)

    def _get_client_ip(self, request):
        """Obtiene IP del cliente"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip


# =============================================================================
# ACEPTACIÓN DOCUMENTAL (Mejora 3 — Lectura Verificada)
# =============================================================================

class AceptacionDocumentalViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de lecturas verificadas de documentos.
    ISO 7.3 Toma de Conciencia + Decreto 1072 Art. 2.2.4.6.10/12.
    """
    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'repositorio'
    http_method_names = ['get', 'post', 'patch']

    # Self-service: acciones personales del empleado en Mi Portal. Cada acción
    # valida internamente que `aceptacion.usuario == request.user`, por lo que
    # un empleado solo puede operar sobre SUS propias lecturas sin necesidad
    # de acceso RBAC al módulo "repositorio" (ese es para admin documental).
    SELF_SERVICE_ACTIONS = frozenset({
        'mis_pendientes',       # GET lista personal
        'registrar_progreso',   # POST avance de scroll/tiempo
        'aceptar',              # POST marca como leído (cumplimiento ISO)
    })

    # Acciones que requieren rol administrativo (can_edit en repositorio).
    # exportar-evidencia produce evidencia ISO firmada — solo admin SGI.
    granular_action_map = {
        'exportar_evidencia': 'can_edit',
    }

    def get_permissions(self):
        if self.action in self.SELF_SERVICE_ACTIONS:
            return [IsAuthenticated()]
        return super().get_permissions()

    def get_serializer_class(self):
        from .serializers import (
            AceptacionDocumentalListSerializer,
            AceptacionDocumentalDetailSerializer,
        )
        if self.action in ('list', 'mis_pendientes'):
            return AceptacionDocumentalListSerializer
        return AceptacionDocumentalDetailSerializer

    def get_queryset(self):
        from .models import AceptacionDocumental
        qs = AceptacionDocumental.objects.select_related(
            'documento', 'usuario', 'asignado_por'
        )
        estado = self.request.query_params.get('estado')
        if estado:
            qs = qs.filter(estado=estado)
        documento_id = self.request.query_params.get('documento')
        if documento_id:
            qs = qs.filter(documento_id=documento_id)
        return qs

    @action(detail=False, methods=['get'], url_path='mis-pendientes')
    def mis_pendientes(self, request):
        """
        Documentos pendientes de lectura del usuario autenticado.

        Filtra defensivamente: si una AceptacionDocumental quedó asignada a un
        documento CONFIDENCIAL/RESTRINGIDO al cual el usuario ya no tiene
        acceso (por reasignación de cargos_distribucion, p. ej.), no se incluye
        en la lista. Esto evita que la auto-distribución legacy filtre
        documentos sensibles a usuarios que no deberían verlos.
        """
        from .models import AceptacionDocumental
        from .serializers import AceptacionDocumentalListSerializer

        qs = AceptacionDocumental.objects.select_related(
            'documento', 'asignado_por'
        ).prefetch_related(
            'documento__usuarios_autorizados',
            'documento__cargos_distribucion',
        ).filter(
            usuario=request.user,
            estado__in=['PENDIENTE', 'EN_PROGRESO'],
            invalidada=False,
        ).order_by('fecha_limite', '-fecha_asignacion')

        # Filtrar fuera asignaciones a docs confidenciales sin acceso vigente (H-GD-C3).
        visibles = [a for a in qs if check_acceso_documento(request.user, a.documento)]

        return Response(
            AceptacionDocumentalListSerializer(
                visibles, many=True, context={'request': request}
            ).data
        )

    @action(detail=False, methods=['post'], url_path='asignar')
    def asignar(self, request):
        """
        Asigna lectura verificada soportando distribución RBAC.

        Modos (combinables):
          usuario_ids  — usuarios individuales
          cargo_ids    — todos los usuarios con esos cargos
          aplica_a_todos — todos los usuarios activos del tenant
        """
        from .serializers import AsignarLecturaSerializer
        from .models import AceptacionDocumental, Documento
        from django.contrib.auth import get_user_model

        serializer = AsignarLecturaSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        documento_id = serializer.validated_data['documento_id']
        usuario_ids = serializer.validated_data.get('usuario_ids') or []
        cargo_ids = serializer.validated_data.get('cargo_ids') or []
        aplica_a_todos = serializer.validated_data.get('aplica_a_todos', False)
        fecha_limite = serializer.validated_data.get('fecha_limite')

        try:
            documento = Documento.objects.get(id=documento_id, estado='PUBLICADO')
        except Documento.DoesNotExist:
            return Response(
                {'error': 'Documento no encontrado o no está publicado'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        User = get_user_model()
        base_qs = User.objects.filter(is_active=True, deleted_at__isnull=True)

        # ── Resolver conjunto final de usuarios ───────────────────────────────
        user_pk_set: set[int] = set()

        # Modo 1: usuarios individuales (legado)
        if usuario_ids:
            user_pk_set.update(
                base_qs.filter(id__in=usuario_ids).values_list('id', flat=True)
            )

        # Modo 2: por cargo RBAC
        if cargo_ids:
            user_pk_set.update(
                base_qs.filter(cargo_id__in=cargo_ids).values_list('id', flat=True)
            )

        # Modo 3: todos los usuarios del tenant
        if aplica_a_todos:
            user_pk_set.update(base_qs.values_list('id', flat=True))

        creados = 0
        omitidos = 0

        for uid in user_pk_set:
            try:
                usuario = User.objects.get(id=uid)
            except User.DoesNotExist:
                omitidos += 1
                continue

            _, created = AceptacionDocumental.objects.get_or_create(
                documento=documento,
                version_documento=documento.version_actual,
                usuario=usuario,
                defaults={
                    'asignado_por': request.user,
                    'fecha_limite': fecha_limite,
                    'estado': 'PENDIENTE',
                },
            )
            if created:
                creados += 1
            else:
                omitidos += 1

        return Response({
            'mensaje': f'Lectura asignada a {creados} usuario(s)',
            'creados': creados,
            'omitidos': omitidos,
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='registrar-progreso')
    def registrar_progreso(self, request, pk=None):
        """Actualiza progreso de scroll y tiempo de lectura."""
        aceptacion = self.get_object()

        if aceptacion.usuario != request.user:
            return Response(
                {'error': 'Solo el usuario asignado puede registrar progreso'},
                status=status.HTTP_403_FORBIDDEN,
            )

        if aceptacion.estado in ('ACEPTADO', 'RECHAZADO'):
            return Response(
                {'error': 'Esta lectura ya fue finalizada'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not aceptacion.fecha_inicio_lectura:
            aceptacion.fecha_inicio_lectura = timezone.now()
            aceptacion.estado = 'EN_PROGRESO'

        porcentaje = request.data.get('porcentaje_lectura', aceptacion.porcentaje_lectura)
        tiempo = request.data.get('tiempo_lectura_seg', aceptacion.tiempo_lectura_seg)
        scroll_data = request.data.get('scroll_data', aceptacion.scroll_data)

        # Nunca decrementar: tomar el máximo entre el progreso actual y el nuevo
        aceptacion.porcentaje_lectura = max(min(int(porcentaje), 100), aceptacion.porcentaje_lectura)
        aceptacion.tiempo_lectura_seg = int(tiempo)
        aceptacion.scroll_data = scroll_data
        aceptacion.save(update_fields=[
            'porcentaje_lectura', 'tiempo_lectura_seg', 'scroll_data',
            'fecha_inicio_lectura', 'estado', 'updated_at',
        ])

        return Response({
            'porcentaje_lectura': aceptacion.porcentaje_lectura,
            'tiempo_lectura_seg': aceptacion.tiempo_lectura_seg,
            'estado': aceptacion.estado,
        })

    @action(detail=True, methods=['post'], url_path='aceptar')
    def aceptar(self, request, pk=None):
        """Acepta el documento tras lectura verificada (requiere >= 90%).

        Validaciones (en orden):
          1. usuario == request.user
          2. estado != ACEPTADO/RECHAZADO
          3. fecha_limite no vencida (si vencida, marca VENCIDO + 400)
          4. Si documento.archivo_original existe (PDF externo):
                scroll_data['total_paginas'] y scroll_data['paginas_vistas']
                deben existir y paginas_vistas/total_paginas >= 0.9
          5. porcentaje_lectura >= 90 (fallback para HTML / PDFs ya validados)
        """
        aceptacion = self.get_object()

        if aceptacion.usuario != request.user:
            return Response(
                {'error': 'Solo el usuario asignado puede aceptar'},
                status=status.HTTP_403_FORBIDDEN,
            )

        if aceptacion.estado in ('ACEPTADO', 'RECHAZADO'):
            return Response(
                {'error': 'Esta lectura ya fue finalizada'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # H-GD-M1 — Validar fecha_limite ANTES de aceptar (evita race con Celery)
        if aceptacion.fecha_limite:
            hoy = timezone.now().date()
            if aceptacion.fecha_limite < hoy and aceptacion.estado in (
                'PENDIENTE', 'EN_PROGRESO'
            ):
                aceptacion.estado = 'VENCIDO'
                aceptacion.save(update_fields=['estado', 'updated_at'])
                fecha_str = aceptacion.fecha_limite.strftime('%d/%m/%Y')
                return Response(
                    {
                        'error': (
                            f'El plazo de lectura venció el {fecha_str}. '
                            'No es posible aceptar este documento.'
                        ),
                        'estado': aceptacion.estado,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # H-GD-C1/C2 — Validar tracking real para PDFs externos
        documento = aceptacion.documento
        tiene_pdf_externo = bool(
            getattr(documento, 'archivo_original', None)
            and documento.archivo_original.name
        )
        if tiene_pdf_externo:
            scroll_data = aceptacion.scroll_data or {}
            total_paginas = scroll_data.get('total_paginas')
            paginas_vistas = scroll_data.get('paginas_vistas')

            if total_paginas is None or paginas_vistas is None:
                return Response(
                    {
                        'error': (
                            'Este documento es un PDF externo. Debe abrirlo y '
                            'recorrer al menos el 90% de sus páginas antes de '
                            'aceptar (tracking de páginas no registrado).'
                        ),
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            try:
                total_int = int(total_paginas)
                vistas_int = int(paginas_vistas)
            except (TypeError, ValueError):
                return Response(
                    {
                        'error': (
                            'Datos de progreso del PDF inválidos. '
                            'Vuelva a abrir el documento y reintente.'
                        ),
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if total_int <= 0:
                return Response(
                    {
                        'error': (
                            'No fue posible determinar el número de páginas '
                            'del PDF. Vuelva a abrir el documento y reintente.'
                        ),
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            ratio = vistas_int / total_int
            if ratio < 0.9:
                pct = round(ratio * 100)
                return Response(
                    {
                        'error': (
                            f'Debe recorrer al menos el 90% de las páginas del '
                            f'PDF (actual: {vistas_int}/{total_int} = {pct}%).'
                        ),
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if aceptacion.porcentaje_lectura < 90:
            return Response(
                {
                    'error': (
                        f'Debe leer al menos el 90% del documento '
                        f'(actual: {aceptacion.porcentaje_lectura}%)'
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        aceptacion.estado = 'ACEPTADO'
        aceptacion.fecha_aceptacion = timezone.now()
        aceptacion.texto_aceptacion = request.data.get(
            'texto_aceptacion',
            'He leído y comprendido el contenido de este documento.'
        )
        aceptacion.ip_address = _get_client_ip(request)
        aceptacion.user_agent = request.META.get('HTTP_USER_AGENT', '')[:500]
        aceptacion.save(update_fields=[
            'estado', 'fecha_aceptacion', 'texto_aceptacion',
            'ip_address', 'user_agent', 'updated_at',
        ])

        from .serializers import AceptacionDocumentalDetailSerializer
        return Response(
            AceptacionDocumentalDetailSerializer(
                aceptacion, context={'request': request}
            ).data
        )

    @action(detail=True, methods=['post'], url_path='rechazar')
    def rechazar(self, request, pk=None):
        """Rechaza la lectura del documento con motivo."""
        aceptacion = self.get_object()

        if aceptacion.usuario != request.user:
            return Response(
                {'error': 'Solo el usuario asignado puede rechazar'},
                status=status.HTTP_403_FORBIDDEN,
            )

        motivo = request.data.get('motivo_rechazo', '')
        if not motivo.strip():
            return Response(
                {'error': 'Debe indicar un motivo de rechazo'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        aceptacion.estado = 'RECHAZADO'
        aceptacion.fecha_rechazo = timezone.now()
        aceptacion.motivo_rechazo = motivo
        aceptacion.ip_address = _get_client_ip(request)
        aceptacion.user_agent = request.META.get('HTTP_USER_AGENT', '')[:500]
        aceptacion.save(update_fields=[
            'estado', 'fecha_rechazo', 'motivo_rechazo',
            'ip_address', 'user_agent', 'updated_at',
        ])

        from .serializers import AceptacionDocumentalDetailSerializer
        return Response(AceptacionDocumentalDetailSerializer(aceptacion).data)

    @action(detail=False, methods=['get'], url_path='resumen')
    def resumen(self, request):
        """Dashboard de aceptaciones: pendientes, completados, vencidos."""
        from .models import AceptacionDocumental
        from django.db.models import Avg

        qs = AceptacionDocumental.objects.all()
        documento_id = request.query_params.get('documento')
        if documento_id:
            qs = qs.filter(documento_id=documento_id)

        stats = qs.aggregate(
            total=Count('id'),
            pendientes=Count('id', filter=Q(estado='PENDIENTE')),
            en_progreso=Count('id', filter=Q(estado='EN_PROGRESO')),
            aceptados=Count('id', filter=Q(estado='ACEPTADO')),
            rechazados=Count('id', filter=Q(estado='RECHAZADO')),
            vencidos=Count('id', filter=Q(estado='VENCIDO')),
            promedio_tiempo=Avg('tiempo_lectura_seg', filter=Q(estado='ACEPTADO')),
            promedio_porcentaje=Avg('porcentaje_lectura'),
        )
        return Response(stats)

    @action(detail=False, methods=['get'], url_path='exportar-evidencia')
    def exportar_evidencia(self, request):
        """
        Exporta el log de aceptaciones documentales como evidencia ISO firmada
        con hash SHA-256.

        Query params:
            documento (int, opcional)  — filtrar por documento.
            desde (YYYY-MM-DD, opcional) — fecha mínima de fecha_asignacion.
            hasta (YYYY-MM-DD, opcional) — fecha máxima de fecha_asignacion.
            formato (xlsx|pdf, default xlsx) — formato de salida.

        Solo accesible a usuarios con can_edit en sección 'repositorio'
        (admin SGI). El export incluye el hash SHA-256 del cuerpo de datos
        en el footer / metadata para verificación de integridad.
        """
        import hashlib
        from datetime import datetime, time

        from django.utils.dateparse import parse_date
        from .models import AceptacionDocumental

        # Filtrado del queryset
        qs = AceptacionDocumental.objects.select_related(
            'documento', 'documento__tipo_documento', 'usuario',
        ).order_by('documento__codigo', 'fecha_asignacion')

        documento_id = request.query_params.get('documento')
        if documento_id:
            qs = qs.filter(documento_id=documento_id)

        desde_raw = request.query_params.get('desde')
        if desde_raw:
            desde = parse_date(desde_raw)
            if desde:
                qs = qs.filter(fecha_asignacion__gte=datetime.combine(desde, time.min))

        hasta_raw = request.query_params.get('hasta')
        if hasta_raw:
            hasta = parse_date(hasta_raw)
            if hasta:
                qs = qs.filter(fecha_asignacion__lte=datetime.combine(hasta, time.max))

        formato = (request.query_params.get('formato') or 'xlsx').lower()
        if formato not in ('xlsx', 'pdf'):
            return Response(
                {'error': 'Formato inválido. Use xlsx o pdf.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Construir filas — orden estable para que el hash sea reproducible
        filas = []
        for ac in qs.iterator(chunk_size=200):
            filas.append({
                'usuario': ac.usuario.get_full_name() if ac.usuario else '',
                'usuario_email': ac.usuario.email if ac.usuario else '',
                'documento_codigo': ac.documento.codigo,
                'documento_titulo': ac.documento.titulo,
                'version': ac.version_documento or '',
                'fecha_asignacion': ac.fecha_asignacion.isoformat() if ac.fecha_asignacion else '',
                'fecha_aceptacion': ac.fecha_aceptacion.isoformat() if ac.fecha_aceptacion else '',
                'porcentaje_lectura': ac.porcentaje_lectura,
                'tiempo_lectura_seg': ac.tiempo_lectura_seg,
                'ip_address': ac.ip_address or '',
                'user_agent': ac.user_agent or '',
                'estado': ac.estado,
                'motivo_rechazo': ac.motivo_rechazo or '',
            })

        # Hash SHA-256 sobre la representación canónica de las filas
        canonical = '\n'.join(
            '|'.join(str(row[k]) for k in (
                'usuario', 'usuario_email', 'documento_codigo', 'version',
                'fecha_asignacion', 'fecha_aceptacion', 'porcentaje_lectura',
                'tiempo_lectura_seg', 'ip_address', 'estado', 'motivo_rechazo',
            ))
            for row in filas
        )
        sha256 = hashlib.sha256(canonical.encode('utf-8')).hexdigest()
        generado_en = timezone.now()
        filtros = {
            'documento': documento_id or '—',
            'desde': desde_raw or '—',
            'hasta': hasta_raw or '—',
            'total_registros': len(filas),
            'generado_por': request.user.get_full_name() or request.user.email,
            'generado_en': generado_en.isoformat(),
            'sha256': sha256,
        }

        if formato == 'xlsx':
            return _build_evidencia_xlsx(filas, filtros)
        return _build_evidencia_pdf(filas, filtros)


def _build_evidencia_xlsx(filas, filtros):
    """Construye un XLSX con las filas + metadata de integridad."""
    from io import BytesIO
    from django.http import HttpResponse

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:  # pragma: no cover
        return Response(
            {'error': 'openpyxl no disponible en el servidor.'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    wb = Workbook()

    # Hoja 1: metadata + hash de integridad
    meta = wb.active
    meta.title = 'Evidencia ISO'
    bold = Font(bold=True)
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    white_bold = Font(bold=True, color='FFFFFF')

    meta['A1'] = 'Evidencia de Aceptación Documental — ISO 7.3 / Decreto 1072'
    meta['A1'].font = Font(bold=True, size=14)
    meta.merge_cells('A1:B1')

    meta_rows = [
        ('Documento', filtros['documento']),
        ('Período desde', filtros['desde']),
        ('Período hasta', filtros['hasta']),
        ('Total registros', filtros['total_registros']),
        ('Generado por', filtros['generado_por']),
        ('Generado en', filtros['generado_en']),
        ('Hash SHA-256', filtros['sha256']),
    ]
    for idx, (k, v) in enumerate(meta_rows, start=3):
        meta[f'A{idx}'] = k
        meta[f'A{idx}'].font = bold
        meta[f'B{idx}'] = v
    meta.column_dimensions['A'].width = 22
    meta.column_dimensions['B'].width = 80

    # Hoja 2: detalle
    ws = wb.create_sheet('Aceptaciones')
    headers = [
        'Usuario', 'Email', 'Código', 'Título', 'Versión',
        'Fecha asignación', 'Fecha aceptación',
        '% Lectura', 'Tiempo (seg)',
        'IP', 'User Agent', 'Estado', 'Motivo rechazo',
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = white_bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for fila in filas:
        ws.append([
            fila['usuario'], fila['usuario_email'],
            fila['documento_codigo'], fila['documento_titulo'], fila['version'],
            fila['fecha_asignacion'], fila['fecha_aceptacion'],
            fila['porcentaje_lectura'], fila['tiempo_lectura_seg'],
            fila['ip_address'], fila['user_agent'],
            fila['estado'], fila['motivo_rechazo'],
        ])

    widths = [22, 28, 16, 36, 10, 22, 22, 10, 12, 16, 40, 14, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="evidencia-aceptaciones-{filtros["sha256"][:8]}.xlsx"'
    )
    response['X-Evidencia-SHA256'] = filtros['sha256']
    return response


def _build_evidencia_pdf(filas, filtros):
    """Construye un PDF con WeasyPrint con metadata + tabla + footer firmado."""
    from io import BytesIO
    from django.http import HttpResponse

    try:
        from weasyprint import HTML
    except ImportError:  # pragma: no cover
        return Response(
            {'error': 'WeasyPrint no disponible en el servidor.'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    rows_html = ''.join(
        f"<tr>"
        f"<td>{(f['usuario'] or '').replace('<', '&lt;')}</td>"
        f"<td>{(f['usuario_email'] or '').replace('<', '&lt;')}</td>"
        f"<td>{f['documento_codigo']}</td>"
        f"<td>{f['version']}</td>"
        f"<td>{f['fecha_asignacion'][:19]}</td>"
        f"<td>{f['fecha_aceptacion'][:19]}</td>"
        f"<td>{f['porcentaje_lectura']}%</td>"
        f"<td>{f['tiempo_lectura_seg']}s</td>"
        f"<td>{f['ip_address']}</td>"
        f"<td>{f['estado']}</td>"
        f"</tr>"
        for f in filas
    )

    html = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
      <meta charset="utf-8" />
      <title>Evidencia ISO — Aceptaciones</title>
      <style>
        @page {{ size: A4 landscape; margin: 1.5cm; @bottom-center {{
            content: "Hash SHA-256: {filtros['sha256']} · Página " counter(page) " de " counter(pages);
            font-size: 8pt; color: #555;
        }} }}
        body {{ font-family: 'Helvetica', sans-serif; color: #1f2937; font-size: 9pt; }}
        h1 {{ font-size: 16pt; margin: 0 0 8px 0; color: #111; }}
        .meta {{ background: #f3f4f6; padding: 12px; border-radius: 6px; margin-bottom: 16px; font-size: 9pt; }}
        .meta dl {{ display: grid; grid-template-columns: 160px 1fr; gap: 4px 12px; margin: 0; }}
        .meta dt {{ font-weight: bold; }}
        table {{ width: 100%; border-collapse: collapse; font-size: 8pt; }}
        thead {{ background: #1f2937; color: white; }}
        th, td {{ border: 1px solid #d1d5db; padding: 4px 6px; text-align: left; vertical-align: top; }}
        tbody tr:nth-child(even) {{ background: #f9fafb; }}
        .hash-footer {{ margin-top: 18px; padding: 10px; border: 1px dashed #6b7280; font-family: monospace; font-size: 8pt; word-break: break-all; }}
      </style>
    </head>
    <body>
      <h1>Evidencia de Aceptación Documental</h1>
      <div class="meta">
        <dl>
          <dt>Documento</dt><dd>{filtros['documento']}</dd>
          <dt>Período desde</dt><dd>{filtros['desde']}</dd>
          <dt>Período hasta</dt><dd>{filtros['hasta']}</dd>
          <dt>Total registros</dt><dd>{filtros['total_registros']}</dd>
          <dt>Generado por</dt><dd>{filtros['generado_por']}</dd>
          <dt>Generado en</dt><dd>{filtros['generado_en']}</dd>
        </dl>
      </div>
      <table>
        <thead><tr>
          <th>Usuario</th><th>Email</th><th>Código</th><th>Versión</th>
          <th>Asignación</th><th>Aceptación</th><th>% Lec.</th><th>Tiempo</th>
          <th>IP</th><th>Estado</th>
        </tr></thead>
        <tbody>{rows_html or '<tr><td colspan="10" style="text-align:center;color:#6b7280;">Sin registros</td></tr>'}</tbody>
      </table>
      <div class="hash-footer">
        <strong>Hash SHA-256 (integridad):</strong><br />
        {filtros['sha256']}
      </div>
    </body>
    </html>
    """

    buf = BytesIO()
    HTML(string=html).write_pdf(buf)
    buf.seek(0)

    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="evidencia-aceptaciones-{filtros["sha256"][:8]}.pdf"'
    )
    response['X-Evidencia-SHA256'] = filtros['sha256']
    return response


class TablaRetencionDocumentalViewSet(viewsets.ModelViewSet):
    """
    CRUD para la Tabla de Retención Documental (TRD).
    Sprint 2 — Arquitectura GD v5 §9.
    """
    permission_classes = [IsAuthenticated, GranularActionPermission]
    module_code = 'gestion_documental'
    section_code = 'configuracion'

    def get_queryset(self):
        from .models import TablaRetencionDocumental
        return TablaRetencionDocumental.objects.select_related(
            'tipo_documento', 'proceso'
        ).order_by('tipo_documento__codigo', 'proceso__code')

    def get_serializer_class(self):
        from .serializers import TablaRetencionDocumentalSerializer
        return TablaRetencionDocumentalSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)


def _get_client_ip(request):
    """Obtiene IP del cliente (función standalone para reutilizar)."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0]
    return request.META.get('REMOTE_ADDR')
