"""
ViewSets para el módulo de Organización

Mixins aplicados:
- StandardViewSetMixin = ToggleActiveMixin + FilterInactiveMixin + BulkActionsMixin + AuditMixin
- OrderingMixin = Reordenamiento de registros con campo 'orden'

Endpoints generados automáticamente:
- POST /{resource}/{id}/toggle-active/ - Toggle is_active
- POST /{resource}/bulk-activate/ - Activar múltiples
- POST /{resource}/bulk-deactivate/ - Desactivar múltiples
- POST /{resource}/bulk-delete/ - Eliminar múltiples (con confirmación)
- POST /{resource}/reorder/ - Reordenar (solo AreaViewSet)
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from django.http import HttpResponse
from django.db import transaction

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from .models import Area, OrganigramaNodePosition, CaracterizacionProceso
from .models_caracterizacion import (
    CaracterizacionProveedor, CaracterizacionEntrada,
    CaracterizacionActividad, CaracterizacionSalida,
    CaracterizacionCliente, CaracterizacionRecurso,
)
from .serializers import (
    AreaSerializer, AreaTreeSerializer, AreaListSerializer,
    OrganigramaNodePositionSerializer,
    CaracterizacionProcesoListSerializer,
    CaracterizacionProcesoDetailSerializer,
)
from apps.core.models import Cargo, User
from apps.core.permissions import GranularActionPermission
from apps.core.mixins import (
    StandardViewSetMixin,
    OrderingMixin,
)


class AreaViewSet(StandardViewSetMixin, OrderingMixin, viewsets.ModelViewSet):
    """
    ViewSet para gestión de Áreas/Departamentos

    Mixins aplicados:
    - StandardViewSetMixin: toggle_active, filtros, bulk actions, audit
    - OrderingMixin: reordenamiento de áreas

    Endpoints estándar:
    - GET /api/organizacion/areas/ - Lista todas las áreas
    - POST /api/organizacion/areas/ - Crea una nueva área
    - GET /api/organizacion/areas/{id}/ - Detalle de un área
    - PUT/PATCH /api/organizacion/areas/{id}/ - Actualiza un área
    - DELETE /api/organizacion/areas/{id}/ - Elimina un área

    Endpoints automáticos (mixins):
    - POST /api/organizacion/areas/{id}/toggle-active/ - Activa/desactiva área
    - POST /api/organizacion/areas/bulk-activate/ - Activar múltiples
    - POST /api/organizacion/areas/bulk-deactivate/ - Desactivar múltiples
    - POST /api/organizacion/areas/bulk-delete/ - Eliminar múltiples
    - POST /api/organizacion/areas/reorder/ - Reordenar áreas

    Endpoints custom:
    - GET /api/organizacion/areas/tree/ - Árbol jerárquico de áreas
    - GET /api/organizacion/areas/root/ - Solo áreas raíz (sin padre)
    - GET /api/organizacion/areas/{id}/children/ - Subáreas directas

    Parámetros de filtrado:
    - is_active: true/false
    - parent: ID del área padre
    - include_inactive: true (para incluir inactivas, por defecto solo activas)
    """
    queryset = Area.objects.all()
    serializer_class = AreaSerializer
    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'areas'

    granular_action_map = {
        'tree': 'can_view',
        'root': 'can_view',
        'children': 'can_view',
        'reorder': 'can_edit',
        'toggle_active': 'can_edit',
        'bulk_activate': 'can_edit',
        'bulk_deactivate': 'can_edit',
        'bulk_delete': 'can_delete',
    }

    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['is_active', 'parent']
    search_fields = ['code', 'name', 'description', 'cost_center']
    ordering_fields = ['orden', 'name', 'code', 'created_at']
    ordering = ['orden', 'name']

    # Configuración para ValidateBeforeDeleteMixin
    protected_relations = ['children']
    custom_error_messages = {
        'children': 'No se puede eliminar un área con subáreas. Elimine primero las subáreas.'
    }

    def get_queryset(self):
        """
        Filtra áreas según parámetros.
        FilterInactiveMixin ya maneja el filtro por is_active con param include_inactive.
        """
        queryset = super().get_queryset()

        # Para acciones de detalle (retrieve, update, destroy, toggle, children)
        # NO aplicar filtro de activas para poder operar sobre áreas inactivas
        if self.action in ['retrieve', 'update', 'partial_update', 'destroy', 'toggle_active', 'children']:
            return Area.objects.all()

        return queryset

    @action(detail=False, methods=['get'])
    def tree(self, request):
        """
        Retorna el árbol jerárquico completo de áreas.
        Solo incluye áreas raíz, las subáreas vienen anidadas.
        """
        root_areas = self.get_queryset().filter(parent__isnull=True)
        serializer = AreaTreeSerializer(root_areas, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def root(self, request):
        """
        Retorna solo las áreas raíz (sin padre).
        Útil para selects de área padre.
        """
        root_areas = self.get_queryset().filter(parent__isnull=True)
        serializer = AreaSerializer(root_areas, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def children(self, request, pk=None):
        """
        Retorna las subáreas directas de un área.
        """
        area = self.get_object()
        children_qs = area.children.all()

        if request.query_params.get('include_inactive') != 'true':
            children_qs = children_qs.filter(is_active=True)

        children = children_qs.order_by('orden', 'name')
        serializer = AreaSerializer(children, many=True)
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        """
        Elimina un área solo si no tiene subáreas activas.
        """
        area = self.get_object()

        if area.children.filter(is_active=True).exists():
            return Response(
                {'error': 'No se puede eliminar un área con subáreas activas. Desactive o elimine primero las subáreas.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        return super().destroy(request, *args, **kwargs)

    def get_serializer_class(self):
        """Usa serializer simplificado para listados"""
        if self.action == 'list':
            return AreaListSerializer
        return AreaSerializer


# =============================================================================
# ORGANIGRAMA VIEW - Datos para visualización interactiva
# =============================================================================

class OrganigramaView(APIView):
    """
    Vista para obtener datos del organigrama visual interactivo.

    Endpoint: GET /api/organizacion/organigrama/

    Retorna:
    - areas: Lista de áreas con jerarquía
    - cargos: Lista de cargos con relaciones y usuarios asignados
    - usuarios: Lista resumida de usuarios (opcional)
    - stats: Estadísticas del organigrama
    """
    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'organigrama'

    def _get_initials(self, name: str) -> str:
        """Genera iniciales a partir del nombre completo"""
        if not name:
            return '??'
        parts = name.strip().split()
        if len(parts) >= 2:
            return f"{parts[0][0]}{parts[1][0]}".upper()
        return name[:2].upper()

    def get(self, request):
        """Obtiene todos los datos necesarios para el organigrama"""
        include_usuarios = request.query_params.get('include_usuarios', 'false').lower() == 'true'
        solo_activos = request.query_params.get('solo_activos', 'true').lower() == 'true'

        # Obtener áreas
        areas_qs = Area.objects.all()
        if solo_activos:
            areas_qs = areas_qs.filter(is_active=True)

        areas_data = []
        for area in areas_qs.order_by('orden', 'name'):
            area_dict = {
                'id': area.id,
                'code': area.code,
                'name': area.name,
                'description': area.description,
                'parent': area.parent_id,
                'parent_name': area.parent.name if area.parent else None,
                'cost_center': area.cost_center,
                'manager': area.manager_id,
                'manager_name': area.manager.get_full_name() if area.manager else None,
                'icon': area.icon,
                'color': area.color,
                'is_active': area.is_active,
                'orden': area.orden,
                'level': area.level,
                'children_count': area.children.filter(is_active=True).count() if solo_activos else area.children.count(),
            }
            # Contar cargos en esta área (excluir cargos del sistema)
            cargos_area = Cargo.objects.filter(area=area, is_system=False)
            if solo_activos:
                cargos_area = cargos_area.filter(is_active=True)
            area_dict['cargos_count'] = cargos_area.count()

            # Contar usuarios en esta área (excluir superusuarios y cargos del sistema)
            usuarios_area = User.objects.filter(
                cargo__area=area,
                cargo__is_system=False,
                is_active=True,
                deleted_at__isnull=True
            ).exclude(is_superuser=True).count()
            area_dict['usuarios_count'] = usuarios_area

            areas_data.append(area_dict)

        # Obtener cargos (excluir cargos del sistema: ADMIN, USUARIO)
        cargos_qs = Cargo.objects.filter(is_system=False).select_related('area', 'parent_cargo')
        if solo_activos:
            cargos_qs = cargos_qs.filter(is_active=True)

        cargos_data = []
        for cargo in cargos_qs.order_by('nivel_jerarquico', 'name'):
            cargo_dict = {
                'id': cargo.id,
                'code': cargo.code,
                'name': cargo.name,
                'description': cargo.description,
                'area': cargo.area_id,
                'area_name': cargo.area.name if cargo.area else None,
                'parent_cargo': cargo.parent_cargo_id,
                'parent_cargo_name': cargo.parent_cargo.name if cargo.parent_cargo else None,
                'nivel_jerarquico': cargo.nivel_jerarquico,
                'is_jefatura': cargo.is_jefatura,
                'is_externo': cargo.is_externo,
                'is_active': cargo.is_active,
                'cantidad_posiciones': cargo.cantidad_posiciones,
            }
            # Obtener usuarios asignados (excluir superusuario del sistema)
            usuarios_cargo = User.objects.filter(
                cargo=cargo,
                is_active=True,
                deleted_at__isnull=True
            ).exclude(is_superuser=True)
            cargo_dict['usuarios_count'] = usuarios_cargo.count()

            cargo_dict['usuarios_asignados'] = [
                {
                    'id': u.id,
                    'full_name': u.get_full_name() or u.username,
                    'photo_url': u.photo.url if u.photo else None,
                    'initials': self._get_initials(u.get_full_name() or u.username),
                }
                for u in usuarios_cargo[:5]
            ]

            # Contar subordinados directos (excluir cargos del sistema)
            subordinados = Cargo.objects.filter(parent_cargo=cargo, is_system=False)
            if solo_activos:
                subordinados = subordinados.filter(is_active=True)
            cargo_dict['subordinados_count'] = subordinados.count()

            cargos_data.append(cargo_dict)

        # Usuarios (opcional, excluir superusuario del sistema)
        usuarios_data = []
        if include_usuarios:
            usuarios_qs = User.objects.filter(
                is_active=True,
                deleted_at__isnull=True
            ).exclude(is_superuser=True).select_related('cargo', 'cargo__area')

            for user in usuarios_qs:
                usuarios_data.append({
                    'id': user.id,
                    'username': user.username,
                    'full_name': user.get_full_name() or user.username,
                    'email': user.email,
                    'cargo_id': user.cargo_id,
                    'cargo_name': user.cargo.name if user.cargo else None,
                    'area_id': user.cargo.area_id if user.cargo else None,
                    'photo_url': user.photo.url if user.photo else None,
                    'initials': self._get_initials(user.get_full_name() or user.username),
                })

        # Estadísticas (excluir superusuarios y cargos del sistema)
        stats = {
            'total_areas': Area.objects.count(),
            'total_cargos': Cargo.objects.filter(is_system=False).count(),
            'total_usuarios': User.objects.filter(
                is_active=True, deleted_at__isnull=True
            ).exclude(is_superuser=True).count(),
            'areas_activas': Area.objects.filter(is_active=True).count(),
            'cargos_activos': Cargo.objects.filter(is_active=True, is_system=False).count(),
        }

        response_data = {
            'areas': areas_data,
            'cargos': cargos_data,
            'stats': stats,
        }

        if include_usuarios:
            response_data['usuarios'] = usuarios_data

        return Response(response_data)


# =============================================================================
# POSICIONES DE NODOS DEL ORGANIGRAMA
# =============================================================================

class OrganigramaNodePositionView(APIView):
    """
    Gestión de posiciones personalizadas de nodos del organigrama.

    GET  /api/organizacion/organigrama/positions/?view_mode=X&direction=Y
    POST /api/organizacion/organigrama/positions/ (bulk upsert)
    DELETE /api/organizacion/organigrama/positions/?view_mode=X&direction=Y (reset)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Cargar posiciones guardadas para un view_mode y direction"""
        view_mode = request.query_params.get('view_mode', 'cargos')
        direction = request.query_params.get('direction', 'TB')

        positions = OrganigramaNodePosition.objects.filter(
            view_mode=view_mode,
            direction=direction,
        )
        serializer = OrganigramaNodePositionSerializer(positions, many=True)
        return Response(serializer.data)

    def post(self, request):
        """Bulk upsert de posiciones de nodos"""
        positions = request.data if isinstance(request.data, list) else [request.data]

        for pos_data in positions:
            OrganigramaNodePosition.objects.update_or_create(
                node_type=pos_data['node_type'],
                node_id=pos_data['node_id'],
                view_mode=pos_data['view_mode'],
                direction=pos_data['direction'],
                defaults={'x': pos_data['x'], 'y': pos_data['y']},
            )

        return Response({'saved': len(positions)}, status=status.HTTP_200_OK)

    def delete(self, request):
        """Reset posiciones para un view_mode y direction"""
        view_mode = request.query_params.get('view_mode')
        direction = request.query_params.get('direction')

        qs = OrganigramaNodePosition.objects.all()
        if view_mode:
            qs = qs.filter(view_mode=view_mode)
        if direction:
            qs = qs.filter(direction=direction)

        count, _ = qs.delete()
        return Response({'deleted': count}, status=status.HTTP_200_OK)


# ==============================================================================
# CARACTERIZACIÓN DE PROCESOS
# ==============================================================================


class CaracterizacionProcesoViewSet(StandardViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para Caracterización de Procesos (SIPOC)

    Endpoints estándar:
    - GET /api/organizacion/caracterizaciones/ - Lista
    - POST /api/organizacion/caracterizaciones/ - Crear
    - GET /api/organizacion/caracterizaciones/{id}/ - Detalle
    - PUT/PATCH /api/organizacion/caracterizaciones/{id}/ - Actualizar
    - DELETE /api/organizacion/caracterizaciones/{id}/ - Eliminar

    Endpoints custom:
    - GET /api/organizacion/caracterizaciones/by-area/{area_id}/ - Por área
    """
    queryset = CaracterizacionProceso.objects.select_related(
        'area', 'lider_proceso', 'created_by'
    ).prefetch_related(
        'caracterizacionproveedors',
        'caracterizacionentradas',
        'caracterizacionactividads',
        'caracterizacionsalidas',
        'caracterizacionclientes',
        'caracterizacionrecursos',
        'caracterizacionindicadors',
        'caracterizacionriesgos',
        'caracterizaciondocumentos',
    ).all()
    serializer_class = CaracterizacionProcesoDetailSerializer
    permission_classes = [IsAuthenticated, GranularActionPermission]
    section_code = 'caracterizaciones'

    granular_action_map = {
        'by_area': 'can_view',
        'export_excel': 'can_view',
        'plantilla_importacion': 'can_view',
        'import_excel': 'can_create',
        'toggle_active': 'can_edit',
        'bulk_activate': 'can_edit',
        'bulk_deactivate': 'can_edit',
        'bulk_delete': 'can_delete',
    }

    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['estado', 'area', 'is_active']
    search_fields = ['area__name', 'area__code', 'objetivo', 'alcance']
    ordering_fields = ['area__name', 'version', 'created_at', 'estado']
    ordering = ['area__name']

    def get_serializer_class(self):
        if self.action == 'list':
            return CaracterizacionProcesoListSerializer
        return CaracterizacionProcesoDetailSerializer

    @action(detail=False, methods=['get'], url_path='by-area/(?P<area_id>[^/.]+)')
    def by_area(self, request, area_id=None):
        """Obtener caracterización de un área específica"""
        try:
            caracterizacion = self.get_queryset().get(area_id=area_id)
            serializer = self.get_serializer(caracterizacion)
            return Response(serializer.data)
        except CaracterizacionProceso.DoesNotExist:
            return Response(
                {'detail': 'No existe caracterización para esta área.'},
                status=status.HTTP_404_NOT_FOUND
            )

    # ------------------------------------------------------------------
    # EXCEL: export / plantilla / import
    # ------------------------------------------------------------------

    def _excel_header_style(self, color_hex='#1565C0'):
        """Estilo reutilizable para headers de Excel."""
        return {
            'font': Font(bold=True, color='FFFFFF', size=11),
            'fill': PatternFill(start_color=color_hex.replace('#', ''), end_color=color_hex.replace('#', ''), fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                bottom=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
            ),
        }

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """Exportar caracterizaciones con SIPOC a Excel multi-hoja."""
        qs = self.filter_queryset(self.get_queryset())
        wb = openpyxl.Workbook()

        # --- HOJA 1: Caracterizaciones ---
        ws = wb.active
        ws.title = 'Caracterizaciones'
        headers_main = [
            'Código Área', 'Nombre Área', 'Estado', 'Versión',
            'Objetivo', 'Alcance', 'Líder del Proceso',
            'Requisitos Normativos', 'Observaciones',
        ]
        style = self._excel_header_style('#1565C0')
        for col_idx, h in enumerate(headers_main, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = style['font']
            cell.fill = style['fill']
            cell.alignment = style['alignment']
            cell.border = style['border']

        for row_idx, c in enumerate(qs, 2):
            ws.cell(row=row_idx, column=1, value=c.area.code if c.area else '')
            ws.cell(row=row_idx, column=2, value=c.area.name if c.area else '')
            ws.cell(row=row_idx, column=3, value=c.get_estado_display())
            ws.cell(row=row_idx, column=4, value=c.version)
            ws.cell(row=row_idx, column=5, value=c.objetivo or '')
            ws.cell(row=row_idx, column=6, value=c.alcance or '')
            ws.cell(row=row_idx, column=7, value=str(c.lider_proceso) if c.lider_proceso else '')
            ws.cell(row=row_idx, column=8, value=c.requisitos_normativos or '')
            ws.cell(row=row_idx, column=9, value=c.observaciones or '')

        for col_idx in range(1, len(headers_main) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

        # --- HOJA 2: SIPOC (todos los items) ---
        ws_sipoc = wb.create_sheet('SIPOC')
        headers_sipoc = [
            'Código Área', 'Tipo Item', 'Nombre / Descripción',
            'Detalle (Origen/Destino/Tipo/Responsable)',
            'Campos Extra',
        ]
        style_sipoc = self._excel_header_style('#2E7D32')
        for col_idx, h in enumerate(headers_sipoc, 1):
            cell = ws_sipoc.cell(row=1, column=col_idx, value=h)
            cell.font = style_sipoc['font']
            cell.fill = style_sipoc['fill']
            cell.alignment = style_sipoc['alignment']
            cell.border = style_sipoc['border']

        sipoc_row = 2
        for c in qs:
            area_code = c.area.code if c.area else ''

            # Proveedores (S)
            for item in c.caracterizacionproveedors.all():
                ws_sipoc.cell(row=sipoc_row, column=1, value=area_code)
                ws_sipoc.cell(row=sipoc_row, column=2, value='Proveedor')
                ws_sipoc.cell(row=sipoc_row, column=3, value=item.nombre)
                ws_sipoc.cell(row=sipoc_row, column=4, value=item.tipo)
                sipoc_row += 1

            # Entradas (I)
            for item in c.caracterizacionentradas.all():
                ws_sipoc.cell(row=sipoc_row, column=1, value=area_code)
                ws_sipoc.cell(row=sipoc_row, column=2, value='Entrada')
                ws_sipoc.cell(row=sipoc_row, column=3, value=item.descripcion)
                ws_sipoc.cell(row=sipoc_row, column=4, value=item.origen)
                sipoc_row += 1

            # Actividades (P)
            for item in c.caracterizacionactividads.all():
                ws_sipoc.cell(row=sipoc_row, column=1, value=area_code)
                ws_sipoc.cell(row=sipoc_row, column=2, value='Actividad')
                ws_sipoc.cell(row=sipoc_row, column=3, value=item.descripcion)
                ws_sipoc.cell(row=sipoc_row, column=4, value=item.responsable)
                sipoc_row += 1

            # Salidas (O)
            for item in c.caracterizacionsalidas.all():
                ws_sipoc.cell(row=sipoc_row, column=1, value=area_code)
                ws_sipoc.cell(row=sipoc_row, column=2, value='Salida')
                ws_sipoc.cell(row=sipoc_row, column=3, value=item.descripcion)
                ws_sipoc.cell(row=sipoc_row, column=4, value=item.destino)
                sipoc_row += 1

            # Clientes (C)
            for item in c.caracterizacionclientes.all():
                ws_sipoc.cell(row=sipoc_row, column=1, value=area_code)
                ws_sipoc.cell(row=sipoc_row, column=2, value='Cliente')
                ws_sipoc.cell(row=sipoc_row, column=3, value=item.nombre)
                ws_sipoc.cell(row=sipoc_row, column=4, value=item.tipo)
                sipoc_row += 1

            # Recursos
            for item in c.caracterizacionrecursos.all():
                ws_sipoc.cell(row=sipoc_row, column=1, value=area_code)
                ws_sipoc.cell(row=sipoc_row, column=2, value='Recurso')
                ws_sipoc.cell(row=sipoc_row, column=3, value=item.descripcion)
                ws_sipoc.cell(row=sipoc_row, column=4, value=item.get_tipo_display())
                sipoc_row += 1

        for col_idx in range(1, len(headers_sipoc) + 1):
            ws_sipoc.column_dimensions[get_column_letter(col_idx)].width = 28

        # Respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="Caracterizaciones_SIPOC.xlsx"'
        return response

    @action(detail=False, methods=['get'], url_path='plantilla-importacion')
    def plantilla_importacion(self, request):
        """Descargar plantilla Excel para importar caracterizaciones."""
        wb = openpyxl.Workbook()

        # --- HOJA 1: Caracterizaciones ---
        ws = wb.active
        ws.title = 'Caracterizaciones'
        headers = [
            'Código Área *', 'Objetivo', 'Alcance',
            'Cargo Líder', 'Requisitos Normativos', 'Observaciones',
        ]
        style = self._excel_header_style('#1565C0')
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = style['font']
            cell.fill = style['fill']
            cell.alignment = style['alignment']
            cell.border = style['border']

        # Instrucciones
        instructions = [
            '= INSTRUCCIONES =',
            'Complete desde la fila 4. El Código Área debe existir en el sistema.',
            'Estado se crea como BORRADOR. Si ya existe una caracterización para el área, se actualiza.',
        ]
        for i, txt in enumerate(instructions, 2):
            ws.cell(row=i, column=1, value=txt)
            ws.cell(row=i, column=1).font = Font(italic=True, color='666666')

        # Fila ejemplo
        ws.cell(row=4, column=1, value='GER')
        ws.cell(row=4, column=2, value='Dirigir y controlar la organización...')
        ws.cell(row=4, column=3, value='Desde la planeación estratégica hasta la revisión por la dirección')
        ws.cell(row=4, column=4, value='Gerente General')
        ws.cell(row=4, column=5, value='ISO 9001:2015 §5 Liderazgo')
        ws.cell(row=4, column=6, value='')

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 30

        # --- HOJA 2: SIPOC ---
        ws_sipoc = wb.create_sheet('SIPOC')
        headers_sipoc = [
            'Código Área *', 'Tipo Item *', 'Nombre / Descripción *',
            'Detalle (Origen/Destino/Tipo/Responsable)',
        ]
        style_sipoc = self._excel_header_style('#2E7D32')
        for col_idx, h in enumerate(headers_sipoc, 1):
            cell = ws_sipoc.cell(row=1, column=col_idx, value=h)
            cell.font = style_sipoc['font']
            cell.fill = style_sipoc['fill']
            cell.alignment = style_sipoc['alignment']
            cell.border = style_sipoc['border']

        instructions_sipoc = [
            '= INSTRUCCIONES =',
            'Tipo Item: Proveedor, Entrada, Actividad, Salida, Cliente, Recurso',
            'Complete desde la fila 4. El Código Área debe coincidir con la hoja Caracterizaciones.',
        ]
        for i, txt in enumerate(instructions_sipoc, 2):
            ws_sipoc.cell(row=i, column=1, value=txt)
            ws_sipoc.cell(row=i, column=1).font = Font(italic=True, color='666666')

        # Ejemplos SIPOC
        ejemplos = [
            ('GER', 'Proveedor', 'Junta Directiva', 'interno'),
            ('GER', 'Entrada', 'Informes de gestión', 'Todos los procesos'),
            ('GER', 'Actividad', 'Planificación estratégica', 'Gerente General'),
            ('GER', 'Salida', 'Plan estratégico', 'Todos los procesos'),
            ('GER', 'Cliente', 'Accionistas', 'externo'),
            ('GER', 'Recurso', 'Sistema de información gerencial', 'Tecnológico'),
        ]
        for i, (area, tipo, nombre, detalle) in enumerate(ejemplos, 4):
            ws_sipoc.cell(row=i, column=1, value=area)
            ws_sipoc.cell(row=i, column=2, value=tipo)
            ws_sipoc.cell(row=i, column=3, value=nombre)
            ws_sipoc.cell(row=i, column=4, value=detalle)

        for col_idx in range(1, len(headers_sipoc) + 1):
            ws_sipoc.column_dimensions[get_column_letter(col_idx)].width = 32

        # --- HOJA 3: Referencia (áreas disponibles) ---
        ws_ref = wb.create_sheet('Áreas Disponibles')
        ref_headers = ['Código', 'Nombre Área', '¿Tiene Caracterización?']
        style_ref = self._excel_header_style('#6A1B9A')
        for col_idx, h in enumerate(ref_headers, 1):
            cell = ws_ref.cell(row=1, column=col_idx, value=h)
            cell.font = style_ref['font']
            cell.fill = style_ref['fill']
            cell.alignment = style_ref['alignment']
            cell.border = style_ref['border']

        areas = Area.objects.filter(is_active=True).order_by('name')
        for i, area in enumerate(areas, 2):
            ws_ref.cell(row=i, column=1, value=area.code)
            ws_ref.cell(row=i, column=2, value=area.name)
            tiene = 'Sí' if hasattr(area, 'caracterizacion') and CaracterizacionProceso.objects.filter(area=area, is_active=True).exists() else 'No'
            ws_ref.cell(row=i, column=3, value=tiene)

        for col_idx in range(1, 4):
            ws_ref.column_dimensions[get_column_letter(col_idx)].width = 25

        # Respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="Plantilla_Caracterizaciones_SIPOC.xlsx"'
        return response

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        """Importar caracterizaciones + SIPOC desde Excel."""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Archivo requerido'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
        except Exception:
            return Response({'error': 'Archivo Excel inválido'}, status=status.HTTP_400_BAD_REQUEST)

        errors = []
        created = 0
        updated = 0
        sipoc_created = 0

        # Mapear áreas por código
        areas_map = {a.code.upper(): a for a in Area.objects.filter(is_active=True)}

        # Mapear cargos por nombre (case-insensitive)
        cargos_map = {c.name.lower(): c for c in Cargo.objects.filter(is_active=True)}

        # --- HOJA 1: Caracterizaciones ---
        if 'Caracterizaciones' in wb.sheetnames:
            ws = wb['Caracterizaciones']
            for row_idx in range(4, ws.max_row + 1):
                area_code = str(ws.cell(row=row_idx, column=1).value or '').strip().upper()
                if not area_code:
                    continue

                area = areas_map.get(area_code)
                if not area:
                    errors.append(f'Fila {row_idx}: Código de área "{area_code}" no encontrado')
                    continue

                objetivo = str(ws.cell(row=row_idx, column=2).value or '').strip()
                alcance = str(ws.cell(row=row_idx, column=3).value or '').strip()
                cargo_nombre = str(ws.cell(row=row_idx, column=4).value or '').strip()
                requisitos = str(ws.cell(row=row_idx, column=5).value or '').strip()
                observaciones = str(ws.cell(row=row_idx, column=6).value or '').strip()

                # Buscar cargo líder
                lider = None
                if cargo_nombre:
                    lider = cargos_map.get(cargo_nombre.lower())

                defaults = {
                    'objetivo': objetivo,
                    'alcance': alcance,
                    'lider_proceso': lider,
                    'requisitos_normativos': requisitos,
                    'observaciones': observaciones,
                }

                with transaction.atomic():
                    obj, was_created = CaracterizacionProceso.objects.update_or_create(
                        area=area,
                        defaults=defaults,
                    )
                    if was_created:
                        obj.created_by = request.user
                        obj.save(update_fields=['created_by'])
                        created += 1
                    else:
                        updated += 1

        # --- HOJA 2: SIPOC ---
        TIPO_MAP = {
            'proveedor': 'proveedor',
            'entrada': 'entrada',
            'actividad': 'actividad',
            'salida': 'salida',
            'cliente': 'cliente',
            'recurso': 'recurso',
        }

        TIPO_RECURSO_MAP = {
            'humano': 'humano',
            'tecnológico': 'tecnologico',
            'tecnologico': 'tecnologico',
            'físico': 'fisico',
            'fisico': 'fisico',
            'financiero': 'financiero',
        }

        if 'SIPOC' in wb.sheetnames:
            ws_sipoc = wb['SIPOC']

            # Agrupar items por área para hacer bulk create
            sipoc_by_area = {}

            for row_idx in range(4, ws_sipoc.max_row + 1):
                area_code = str(ws_sipoc.cell(row=row_idx, column=1).value or '').strip().upper()
                tipo_raw = str(ws_sipoc.cell(row=row_idx, column=2).value or '').strip().lower()
                nombre = str(ws_sipoc.cell(row=row_idx, column=3).value or '').strip()
                detalle = str(ws_sipoc.cell(row=row_idx, column=4).value or '').strip()

                if not area_code or not tipo_raw or not nombre:
                    continue

                tipo = TIPO_MAP.get(tipo_raw)
                if not tipo:
                    errors.append(f'SIPOC fila {row_idx}: Tipo "{tipo_raw}" no válido')
                    continue

                if area_code not in sipoc_by_area:
                    sipoc_by_area[area_code] = []
                sipoc_by_area[area_code].append({
                    'tipo': tipo,
                    'nombre': nombre,
                    'detalle': detalle,
                    'row': row_idx,
                })

            # Procesar items agrupados por área
            for area_code, items in sipoc_by_area.items():
                area = areas_map.get(area_code)
                if not area:
                    errors.append(f'SIPOC: Código de área "{area_code}" no encontrado')
                    continue

                try:
                    caract = CaracterizacionProceso.objects.get(area=area)
                except CaracterizacionProceso.DoesNotExist:
                    errors.append(f'SIPOC: No existe caracterización para área "{area_code}". Créela primero en la hoja Caracterizaciones.')
                    continue

                with transaction.atomic():
                    for idx, item in enumerate(items):
                        tipo = item['tipo']
                        nombre = item['nombre']
                        detalle = item['detalle']

                        if tipo == 'proveedor':
                            tipo_valor = 'externo' if detalle.lower() in ('externo', 'ext', '') else 'interno'
                            CaracterizacionProveedor.objects.create(
                                caracterizacion=caract, nombre=nombre,
                                tipo=tipo_valor if detalle.lower() in ('interno', 'int', 'externo', 'ext', '') else 'externo',
                                orden=idx,
                            )
                        elif tipo == 'entrada':
                            CaracterizacionEntrada.objects.create(
                                caracterizacion=caract, descripcion=nombre,
                                origen=detalle, orden=idx,
                            )
                        elif tipo == 'actividad':
                            CaracterizacionActividad.objects.create(
                                caracterizacion=caract, descripcion=nombre,
                                responsable=detalle, orden=idx,
                            )
                        elif tipo == 'salida':
                            CaracterizacionSalida.objects.create(
                                caracterizacion=caract, descripcion=nombre,
                                destino=detalle, orden=idx,
                            )
                        elif tipo == 'cliente':
                            tipo_valor = 'interno' if detalle.lower() in ('interno', 'int') else 'externo'
                            CaracterizacionCliente.objects.create(
                                caracterizacion=caract, nombre=nombre,
                                tipo=tipo_valor, orden=idx,
                            )
                        elif tipo == 'recurso':
                            tipo_recurso = TIPO_RECURSO_MAP.get(detalle.lower(), 'humano')
                            CaracterizacionRecurso.objects.create(
                                caracterizacion=caract, descripcion=nombre,
                                tipo=tipo_recurso, orden=idx,
                            )

                        sipoc_created += 1

        return Response({
            'message': f'{created} caracterizaciones creadas, {updated} actualizadas, {sipoc_created} items SIPOC importados.',
            'created': created,
            'updated': updated,
            'sipoc_created': sipoc_created,
            'errors': errors,
        })
