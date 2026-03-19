"""
Utilidades para importación masiva de Colaboradores desde Excel/CSV.
Sistema de Gestión StrateKaz — Talent Hub
"""
import io
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# =============================================================================
# MAPEOS DE VALORES LEGIBLES → CODES DEL MODELO
# =============================================================================

TIPO_DOCUMENTO_MAP = {
    'CC': 'CC',
    'C.C.': 'CC',
    'CEDULA DE CIUDADANIA': 'CC',
    'CÉDULA DE CIUDADANÍA': 'CC',
    'CE': 'CE',
    'C.E.': 'CE',
    'CEDULA EXTRANJERIA': 'CE',
    'CÉDULA EXTRANJERÍA': 'CE',
    'TI': 'TI',
    'T.I.': 'TI',
    'TARJETA DE IDENTIDAD': 'TI',
    'PA': 'PA',
    'PASAPORTE': 'PA',
    'PEP': 'PEP',
    'PPT': 'PPT',
    'NIT': 'NIT',
}

TIPO_CONTRATO_MAP = {
    'INDEFINIDO': 'indefinido',
    'FIJO': 'fijo',
    'TERMINO FIJO': 'fijo',
    'TÉRMINO FIJO': 'fijo',
    'OBRA LABOR': 'obra_labor',
    'OBRA O LABOR': 'obra_labor',
    'OBRA_LABOR': 'obra_labor',
    'APRENDIZAJE': 'aprendizaje',
    'PRESTACION DE SERVICIOS': 'prestacion_servicios',
    'PRESTACIÓN DE SERVICIOS': 'prestacion_servicios',
    'PRESTACION_SERVICIOS': 'prestacion_servicios',
}

ESTADO_MAP = {
    'ACTIVO': 'activo',
    'INACTIVO': 'inactivo',
    'SUSPENDIDO': 'suspendido',
    'RETIRADO': 'retirado',
}

SI_NO_MAP = {
    'SI': True, 'SÍ': True, 'S': True, '1': True, 'TRUE': True, 'YES': True, 'Y': True,
    'NO': False, 'N': False, '0': False, 'FALSE': False,
}

# Columnas del archivo de importación (índice 0-based)
COLUMNAS = [
    'tipo_documento',
    'numero_identificacion',
    'primer_nombre',
    'segundo_nombre',
    'primer_apellido',
    'segundo_apellido',
    'cargo_nombre',
    'area_nombre',
    'fecha_ingreso',
    'tipo_contrato',
    'fecha_fin_contrato',
    'salario',
    'auxilio_transporte',
    'horas_semanales',
    'email_personal',
    'telefono_movil',
    'estado',
    'crear_acceso',
    'email_corporativo',
    'username',
    'observaciones',
]

# =============================================================================
# PARSER DE EXCEL
# =============================================================================

def parse_excel_file(file_content: bytes) -> list[dict]:
    """
    Parsea un archivo Excel y retorna lista de dicts con los datos de cada fila.
    Ignora la primera fila (cabeceras) y filas completamente vacías.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Ignorar filas completamente vacías
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        row_data = {'_fila': row_idx}
        for col_idx, col_name in enumerate(COLUMNAS):
            if col_idx < len(row):
                val = row[col_idx]
                row_data[col_name] = val if val is not None else ''
            else:
                row_data[col_name] = ''

        rows.append(row_data)

    return rows


def normalizar_valor(valor) -> str:
    """Convierte un valor de celda a string normalizado en mayúsculas."""
    if valor is None:
        return ''
    return str(valor).strip().upper()


def parsear_fecha(valor) -> str | None:
    """Convierte fecha de celda (date, datetime o string) a string YYYY-MM-DD."""
    if not valor:
        return None
    if hasattr(valor, 'strftime'):
        return valor.strftime('%Y-%m-%d')
    s = str(valor).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def parsear_decimal(valor) -> str | None:
    """Parsea un valor de celda a string decimal."""
    if not valor and valor != 0:
        return None
    try:
        val = float(str(valor).replace(',', '.').replace(' ', ''))
        return str(val)
    except (ValueError, TypeError):
        return None


def parsear_entero(valor, default: int = 0) -> int:
    """Parsea un valor de celda a entero."""
    if not valor and valor != 0:
        return default
    try:
        return int(float(str(valor).replace(',', '').replace(' ', '')))
    except (ValueError, TypeError):
        return default


def parsear_bool(valor) -> bool:
    """Parsea un valor de celda a booleano."""
    return SI_NO_MAP.get(normalizar_valor(valor), False)


# =============================================================================
# GENERADOR DE PLANTILLA EXCEL
# =============================================================================

# Colores de la plantilla
COLOR_HEADER_BG = '4472C4'
COLOR_HEADER_FONT = 'FFFFFF'
COLOR_REQUIRED_BG = 'FFF2CC'
COLOR_EXAMPLE_BG = 'F2F2F2'
COLOR_SECTION_BG = 'D6E4F0'

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

# Definición de columnas para la plantilla
TEMPLATE_COLUMNS = [
    # (header, ancho, requerido, ejemplo, nota)
    ('Tipo Documento', 18, True, 'CC', 'CC, CE, TI, PA, PEP, PPT'),
    ('Número Identificación', 22, True, '1234567890', 'Único por empresa'),
    ('Primer Nombre', 18, True, 'Juan', ''),
    ('Segundo Nombre', 18, False, 'Carlos', 'Opcional'),
    ('Primer Apellido', 18, True, 'Pérez', ''),
    ('Segundo Apellido', 18, False, 'González', 'Opcional'),
    ('Cargo', 25, True, 'Coordinador SST', 'Nombre exacto del cargo'),
    ('Área / Proceso', 25, True, 'Seguridad Industrial', 'Nombre exacto del área'),
    ('Fecha Ingreso', 16, True, '2026-01-15', 'Formato: AAAA-MM-DD'),
    ('Tipo Contrato', 22, True, 'indefinido', 'indefinido, fijo, obra_labor, aprendizaje, prestacion_servicios'),
    ('Fecha Fin Contrato', 18, False, '2027-01-15', 'Requerido si contrato=fijo'),
    ('Salario', 16, True, '3000000', 'Valor numérico sin puntos ni comas'),
    ('Auxilio Transporte', 20, True, 'Si', 'Si o No'),
    ('Horas Semanales', 18, True, '48', 'Número (default 48)'),
    ('Email Personal', 28, False, 'juan@gmail.com', 'Opcional'),
    ('Celular', 16, False, '3105551234', 'Opcional'),
    ('Estado', 14, True, 'activo', 'activo, inactivo, suspendido, retirado'),
    ('Crear Acceso Sistema', 20, True, 'Si', 'Si = crea cuenta + envía email'),
    ('Email Corporativo', 28, False, 'juan.perez@empresa.com', 'Requerido si Crear Acceso=Si'),
    ('Username', 22, False, 'juan.perez', 'Sin espacios. Requerido si Crear Acceso=Si'),
    ('Observaciones', 30, False, '', 'Opcional'),
]


def generate_import_template(cargos: list[dict] = None, areas: list[dict] = None) -> bytes:
    """
    Genera el archivo Excel de plantilla para importación masiva de colaboradores.
    Incluye:
    - Hoja 1: Plantilla con cabeceras coloreadas y fila de ejemplo
    - Hoja 2: Valores válidos (cargos, áreas, tipos de documento, etc.)
    """
    wb = openpyxl.Workbook()

    # ── HOJA 1: Plantilla ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Colaboradores'

    # Fila de cabeceras
    for col_idx, (header, width, required, _, _nota) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
        cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        # Marcador de requerido (fondo amarillo tenue en cabecera si es requerido)
        if required:
            cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)

    ws.row_dimensions[1].height = 30

    # Fila de ejemplo
    for col_idx, (_, _, _, ejemplo, nota) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=ejemplo)
        cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_EXAMPLE_BG)
        cell.font = Font(italic=True, color='888888', size=9)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = THIN_BORDER

    # Fila de notas (fila 3 como guía)
    for col_idx, (_, _, required, _, nota) in enumerate(TEMPLATE_COLUMNS, start=1):
        prefix = '* ' if required else ''
        cell = ws.cell(row=3, column=col_idx, value=f'{prefix}{nota}' if nota else ('* Requerido' if required else ''))
        cell.fill = PatternFill(fill_type='solid', fgColor='FFFBE6' if required else 'F9F9F9')
        cell.font = Font(italic=True, color='777777', size=8)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 25

    # Filas vacías para datos (4 al 103 = 100 filas)
    for row_idx in range(4, 104):
        for col_idx in range(1, len(TEMPLATE_COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx, value='')
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Fijar primera fila
    ws.freeze_panes = 'A4'

    # ── HOJA 2: Valores válidos ───────────────────────────────────────────────
    ws2 = wb.create_sheet('Valores Válidos')

    def write_section(ws, title: str, headers: list[str], rows: list[list], start_row: int, start_col: int) -> int:
        """Escribe una sección de valores en la hoja de referencia."""
        # Título
        title_cell = ws.cell(row=start_row, column=start_col, value=title)
        title_cell.font = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
        title_cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        title_cell.alignment = Alignment(horizontal='center')
        if len(headers) > 1:
            ws.merge_cells(
                start_row=start_row, start_column=start_col,
                end_row=start_row, end_column=start_col + len(headers) - 1
            )
        start_row += 1
        # Cabeceras
        for h_idx, h in enumerate(headers, start=start_col):
            hc = ws.cell(row=start_row, column=h_idx, value=h)
            hc.font = Font(bold=True, size=9)
            hc.fill = PatternFill(fill_type='solid', fgColor=COLOR_SECTION_BG)
            hc.border = THIN_BORDER
        start_row += 1
        # Datos
        for r in rows:
            for c_idx, val in enumerate(r, start=start_col):
                rc = ws.cell(row=start_row, column=c_idx, value=val)
                rc.border = THIN_BORDER
                rc.font = Font(size=9)
            start_row += 1
        return start_row + 1  # fila de separación

    # Secciones de referencia
    row = 1
    row = write_section(ws2, 'TIPOS DE DOCUMENTO', ['Código', 'Descripción'], [
        ['CC', 'Cédula de Ciudadanía'],
        ['CE', 'Cédula de Extranjería'],
        ['TI', 'Tarjeta de Identidad'],
        ['PA', 'Pasaporte'],
        ['PEP', 'Permiso Especial de Permanencia'],
        ['PPT', 'Permiso por Protección Temporal'],
    ], row, 1)

    row = write_section(ws2, 'TIPOS DE CONTRATO', ['Código', 'Descripción'], [
        ['indefinido', 'Contrato Indefinido'],
        ['fijo', 'Término Fijo (requiere Fecha Fin)'],
        ['obra_labor', 'Obra o Labor'],
        ['aprendizaje', 'Contrato de Aprendizaje'],
        ['prestacion_servicios', 'Prestación de Servicios'],
    ], row, 1)

    row = write_section(ws2, 'ESTADOS', ['Código', 'Descripción'], [
        ['activo', 'Empleado activo'],
        ['inactivo', 'Inactivo (suspensión temporal)'],
        ['suspendido', 'Suspendido'],
        ['retirado', 'Retirado (requiere Fecha Retiro)'],
    ], row, 1)

    # Cargos disponibles en el tenant
    cargo_rows = [[c['name']] for c in (cargos or [])]
    if cargo_rows:
        row_cargos = 1
        row_cargos = write_section(ws2, 'CARGOS DISPONIBLES', ['Nombre del Cargo'], cargo_rows, row_cargos, 4)

    # Áreas disponibles en el tenant
    area_rows = [[a['name']] for a in (areas or [])]
    if area_rows:
        row_areas = 1
        row_areas = write_section(ws2, 'ÁREAS DISPONIBLES', ['Nombre del Área'], area_rows, row_areas, 6)

    # Ajustar anchos hoja 2
    for col_idx in range(1, 8):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 30

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
