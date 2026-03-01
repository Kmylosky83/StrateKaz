"""
Utilidades para importación masiva de Clientes desde Excel.
Sistema de Gestión StrateKaz — Sales CRM

Sigue el mismo patrón que core.import_cargos_utils
"""
import io
import logging

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# =============================================================================
# MAPEOS DE VALORES LEGIBLES → CODES DEL MODELO
# =============================================================================

TIPO_DOCUMENTO_MAP = {
    'NIT': 'NIT',
    'CC': 'CC',
    'CEDULA': 'CC',
    'CÉDULA': 'CC',
    'CEDULA DE CIUDADANIA': 'CC',
    'CÉDULA DE CIUDADANÍA': 'CC',
    'CE': 'CE',
    'CEDULA DE EXTRANJERIA': 'CE',
    'CÉDULA DE EXTRANJERÍA': 'CE',
    'PASAPORTE': 'PASAPORTE',
}

# Columnas del archivo de importación (índice 0-based)
CLIENTE_COLUMNAS = [
    'tipo_documento',
    'numero_documento',
    'razon_social',
    'nombre_comercial',
    'tipo_cliente_nombre',
    'canal_venta_nombre',
    'telefono',
    'email',
    'direccion',
    'ciudad',
    'departamento',
    'plazo_pago_dias',
    'cupo_credito',
    'descuento_comercial',
    'observaciones',
]


# =============================================================================
# PARSER DE EXCEL
# =============================================================================

def parse_cliente_excel(file_content: bytes) -> list[dict]:
    """
    Parsea un archivo Excel y retorna lista de dicts con los datos de cada fila.
    Ignora la primera fila (cabeceras), la fila de ejemplo y la fila de notas.
    Los datos empiezan en la fila 4.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        # Ignorar filas completamente vacías
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        row_data = {'_fila': row_idx}
        for col_idx, col_name in enumerate(CLIENTE_COLUMNAS):
            if col_idx < len(row):
                val = row[col_idx]
                row_data[col_name] = val if val is not None else ''
            else:
                row_data[col_name] = ''

        rows.append(row_data)

    return rows


def normalizar_valor(valor) -> str:
    """Convierte un valor de celda a string normalizado en mayúsculas."""
    if valor is None:
        return ''
    return str(valor).strip().upper()


def parsear_entero(valor, default: int = 0) -> int:
    """Parsea un valor de celda a entero."""
    if not valor and valor != 0:
        return default
    try:
        return int(float(str(valor).replace(',', '').replace(' ', '')))
    except (ValueError, TypeError):
        return default


def parsear_decimal(valor, default: str = '0.00') -> str:
    """Parsea un valor de celda a string decimal."""
    if not valor and valor != 0:
        return default
    try:
        val = float(str(valor).replace(',', '').replace(' ', '').replace('$', '').replace('%', ''))
        return f'{val:.2f}'
    except (ValueError, TypeError):
        return default


# =============================================================================
# GENERADOR DE PLANTILLA EXCEL
# =============================================================================

# Colores de la plantilla
COLOR_HEADER_BG = '1565C0'
COLOR_HEADER_FONT = 'FFFFFF'
COLOR_EXAMPLE_BG = 'F2F2F2'
COLOR_SECTION_BG = 'BBDEFB'

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

# Definición de columnas para la plantilla
TEMPLATE_COLUMNS = [
    # (header, ancho, requerido, ejemplo, nota)
    ('Tipo Documento', 22, True, 'NIT', 'NIT, CC, CE, PASAPORTE'),
    ('Número Documento', 20, True, '900123456-1', 'Número de identificación'),
    ('Razón Social', 30, True, 'Distribuciones ABC S.A.S.', 'Nombre legal del cliente'),
    ('Nombre Comercial', 28, False, 'ABC Distribuciones', 'Nombre comercial o marca'),
    ('Tipo Cliente', 22, True, 'EMPRESA', 'Nombre del tipo de cliente (ver hoja Referencia)'),
    ('Canal de Venta', 22, True, 'DIRECTO', 'Canal de venta (ver hoja Referencia)'),
    ('Teléfono', 18, False, '3001234567', 'Teléfono de contacto'),
    ('Email', 28, False, 'ventas@abc.com', 'Correo electrónico'),
    ('Dirección', 35, False, 'Cra 45 #12-34', 'Dirección completa'),
    ('Ciudad', 18, False, 'Bogotá', 'Ciudad del cliente'),
    ('Departamento', 20, False, 'Cundinamarca', 'Departamento'),
    ('Plazo Pago (días)', 18, False, '30', 'Días de plazo para pago (default: 30)'),
    ('Cupo Crédito', 18, False, '5000000', 'Cupo máximo de crédito en pesos'),
    ('Descuento %', 16, False, '5', 'Descuento comercial (0-100%)'),
    ('Observaciones', 35, False, 'Cliente mayorista zona norte', 'Notas adicionales'),
]


def generate_cliente_import_template(
    tipos_cliente: list[dict] | None = None,
    estados_cliente: list[dict] | None = None,
    canales_venta: list[dict] | None = None,
    clientes_existentes: list[dict] | None = None,
) -> bytes:
    """
    Genera el archivo Excel de plantilla para importación masiva de clientes.
    Incluye:
    - Hoja 1: Plantilla con cabeceras coloreadas, fila de ejemplo y notas
    - Hoja 2: Valores de referencia (tipos, canales, clientes existentes)
    """
    wb = openpyxl.Workbook()

    # ── HOJA 1: Plantilla ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Clientes'

    # Fila de cabeceras (fila 1)
    for col_idx, (header, width, required, _, _nota) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
        cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # Fila de ejemplo (fila 2)
    for col_idx, (_, _, _, ejemplo, _nota) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=ejemplo)
        cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_EXAMPLE_BG)
        cell.font = Font(italic=True, color='888888', size=9)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = THIN_BORDER

    # Fila de notas (fila 3)
    for col_idx, (_, _, required, _, nota) in enumerate(TEMPLATE_COLUMNS, start=1):
        prefix = '* ' if required else ''
        cell = ws.cell(
            row=3, column=col_idx,
            value=f'{prefix}{nota}' if nota else ('* Requerido' if required else '')
        )
        cell.fill = PatternFill(fill_type='solid', fgColor='FFFBE6' if required else 'F9F9F9')
        cell.font = Font(italic=True, color='777777', size=8)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 25

    # Filas vacías para datos (4 al 503 = 500 filas)
    for row_idx in range(4, 504):
        for col_idx in range(1, len(TEMPLATE_COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx, value='')
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Fijar filas de cabecera
    ws.freeze_panes = 'A4'

    # ── HOJA 2: Valores de Referencia ────────────────────────────────────────
    ws2 = wb.create_sheet('Referencia')

    def write_section(ws, title, headers, rows, start_row, start_col):
        """Escribe una sección de valores en la hoja de referencia."""
        title_cell = ws.cell(row=start_row, column=start_col, value=title)
        title_cell.font = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
        title_cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        title_cell.alignment = Alignment(horizontal='center')
        if len(headers) > 1:
            ws.merge_cells(
                start_row=start_row, start_column=start_col,
                end_row=start_row, end_column=start_col + len(headers) - 1
            )
        start_row += 1
        # Cabeceras
        for h_idx, h in enumerate(headers, start=start_col):
            hc = ws.cell(row=start_row, column=h_idx, value=h)
            hc.font = Font(bold=True, size=9)
            hc.fill = PatternFill(fill_type='solid', fgColor=COLOR_SECTION_BG)
            hc.border = THIN_BORDER
        start_row += 1
        # Datos
        for r in rows:
            for c_idx, val in enumerate(r, start=start_col):
                rc = ws.cell(row=start_row, column=c_idx, value=val)
                rc.border = THIN_BORDER
                rc.font = Font(size=9)
            start_row += 1
        return start_row + 1  # fila de separación

    # Secciones de referencia - columna 1
    row = 1

    # Tipos de documento
    row = write_section(ws2, 'TIPOS DE DOCUMENTO', ['Código', 'Descripción'], [
        ['NIT', 'NIT'],
        ['CC', 'Cédula de Ciudadanía'],
        ['CE', 'Cédula de Extranjería'],
        ['PASAPORTE', 'Pasaporte'],
    ], row, 1)

    # Tipos de cliente
    tipo_rows = [[t['nombre'], t.get('codigo', '')] for t in (tipos_cliente or [])]
    if tipo_rows:
        row = write_section(ws2, 'TIPOS DE CLIENTE', ['Nombre', 'Código'], tipo_rows, row, 1)

    # Canales de venta
    canal_rows = [[c['nombre'], c.get('codigo', '')] for c in (canales_venta or [])]
    if canal_rows:
        row = write_section(ws2, 'CANALES DE VENTA', ['Nombre', 'Código'], canal_rows, row, 1)

    # Clientes existentes - columna 4
    cliente_rows = [[c['razon_social'], c.get('codigo_cliente', '')] for c in (clientes_existentes or [])]
    if cliente_rows:
        write_section(
            ws2, 'CLIENTES EXISTENTES',
            ['Razón Social', 'Código'], cliente_rows, 1, 4
        )

    # Ajustar anchos hoja 2
    for col_idx in range(1, 7):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 30

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
